from flask import Blueprint
from flask.json import jsonify
from flask import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
import math, random
import os
import requests
from api.sendMail import * 
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import simpleSplit

# import datetime
# import mysql.connector




course = Blueprint('course', __name__)


courseFolderPath = ''



@course.route('/getCourseList',methods=['POST'])
@cross_origin()
def getCourseList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json

        print(req)
        # mentorName = req['mentorName']
        mentorName = ''
        # status = req['status']
        # if mentorName != '':

        #     courseSql = " SELECT DISTINCT(courseType) FROM `batchMaster` WHERE mentorName LIKE '"+str(mentorName)+"' "
        # else:
        courseSql = " SELECT id,`courseName`, `code`, `description`,`folderName` ,activeFlag FROM `courseMaster` WHERE 1 ORDER BY `id` DESC"
        print(courseSql)
        cursor.execute(courseSql)
        courseList = cursor.fetchall()

        folder_names = [course['folderName'] for course in courseList]
        print("courseList['folderName']", folder_names)

       
        for course in courseList:
            folder_name = course['folderName']
            if folder_name:
                folders_and_files = folder_name.split(',')  # Split by comma to get individual entries
                folder_file_list = []
                for entry in folders_and_files:
                    parts = entry.split('/')  # Split each entry by slash to get folder and file
                    if len(parts) == 2:
                        folder_file_dict = {
                            'path': entry.strip(),
                            'file': parts[1].strip()
                        }
                        folder_file_list.append(folder_file_dict)

                course['folderName'] = folder_file_list

    

        resp =  jsonify({"status":1, "success":True, "result": courseList, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp



@course.route('/getCourseContent',methods=['POST'])
@cross_origin()
def getCourseContent():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json

        print(req)

       
       
        courseSql = " SELECT id,`courseName`, `code`,`folderName` ,`contentStatus`,`updatedBy`,`lastUpdated` FROM `courseMaster` WHERE 1"
        print(courseSql)
        cursor.execute(courseSql)
        courseList = cursor.fetchall()
        print(courseList)
        
        folder_names = [course['folderName'] for course in courseList]
        print("courseList['folderName']", folder_names)

        # result_list = []

        # for course in courseList:
        #     folder_name = course['folderName']
        #     if folder_name:
        #         folders_and_files = folder_name.split(',')  # Split by comma to get individual entries
        #         folder_file_list = [entry.split('/') for entry in folders_and_files]  # Split each entry by slash to get folder and file
        #         course['folderName'] = folder_file_list

        # print(courseList)
        for course in courseList:
            folder_name = course['folderName']
            if folder_name:
                folders_and_files = folder_name.split(',')  # Split by comma to get individual entries
                folder_file_list = []
                for entry in folders_and_files:
                    parts = entry.split('/')  # Split each entry by slash to get folder and file
                    if len(parts) == 2:
                        folder_file_dict = {
                            'path': entry.strip(),
                            'file': parts[1].strip()
                        }
                        folder_file_list.append(folder_file_dict)

                course['folderName'] = folder_file_list

        print(courseList)
   
        

        resp =  jsonify({"status":1, "success":True, "result": courseList,"message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp

@course.route('/addCourse',methods=['POST'])
@cross_origin()
def addCourse():
    global courseFolderPath
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json

        print(req)
        print("courseFolderPath---------->", courseFolderPath)

        courseName = req['courseName']
        code = req['code']
        description = req['description']

        insertSql = "INSERT INTO `courseMaster`(`courseName`, `code`, `description`,`folderName`) VALUES (%s,%s,%s,%s)"

        data = (courseName,code,description,courseFolderPath)
        cursor.execute(insertSql,data)


        resp =  jsonify({"status":1, "success":True, "message": " Inserted successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp




@course.route('/editDeleteCourse',methods=['POST'])
@cross_origin()
def editDeleteCourse():
    global courseFolderPath
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        # courseFolderPath = 'courseDetails/new course 7_2024-07-18-14-33-56.pdf'
        print("courseFolderPath", courseFolderPath)
        req =request.json
        print(req)
        courseName = req['courseName']
        code = req['code']
        description = req['description']
        delete =req['delete']
        id = req ['id']
        activeFlag = req['activeFlag']

        # Retrieve existing folderName from the database
        selectSql = f"SELECT folderName FROM courseMaster WHERE id = {id}"
        cursor.execute(selectSql)
        result = cursor.fetchone()

        print("45678987654",result)
        
        if result and result['folderName'] is not None:
            existingFolderName = result['folderName']
        else:
            existingFolderName = ''

        # Initialize courseFolderName
        if existingFolderName:
            courseFolderName = f"{existingFolderName},{courseFolderPath}"
        else:
            courseFolderName = courseFolderPath

        # Update database based on delete flag
        if delete == 1:
            updateSql = f"DELETE FROM courseMaster WHERE id = {id}"
        else:
            updateSql = f"UPDATE courseMaster SET courseName = '{courseName}', code = '{code}', description = '{description}', activeFlag = '{activeFlag}', folderName = '{courseFolderName}' WHERE id = {id}"

        cursor.execute(updateSql)
        conn.commit()  # Commit the transaction

        resp = jsonify({"status": 1, "success": True, "message": "Updated successfully" })


        # # Retrieve existing folderName from the database
        # selectSql = "SELECT folderName FROM courseMaster WHERE id = "+str(id)
        # cursor.execute(selectSql)
        # result = cursor.fetchone()
        # print(result)

        # existingFolderName = result if result else ""  
        # # existingFolderName = result['folderName'] if result else ""      
        # print("existingFolderName", existingFolderName)

        # # Append existingFolderName to the new courseFolderName if it's not empty
        # if existingFolderName:
        #     courseFolderName = f"{existingFolderName},{courseFolderPath}"
        #     print("courseFolderName", courseFolderName)

        # if delete == 1:
        #     updateSql = " delete from courseMaster where id ="+str(id)
        # else :
        #     updateSql = " update courseMaster set courseName ='"+str(courseName)+"' , code = '"+str(code)+"' , description = '" +str(description)+"', activeFlag = '"+str(activeFlag)+"' , folderName ='"+str(courseFolderName)+"' where id ="+str(id)
        
        # cursor.execute(updateSql)


        # resp = jsonify({"status":1, "success":True,  "message": " Updated successfully" })



    except Exception as e:
        print("cghdhbjn",e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp

@course.route('/getBatchList',methods=['POST'])
@cross_origin()
def getBatchList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json

        print(req)
        course = req['course']
        if course != 'all':
            sql = "SELECT DISTINCT(batchName) FROM `batchMaster` where courseType = '"+str(course)+"'"
        else:
            sql = " SELECT DISTINCT(batchName) FROM `batchMaster` "
        cursor.execute(sql)
        batchList = cursor.fetchall()

    
        resp =  jsonify({"status":1, "success":True, "result": batchList, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp










#Rishu 

@course.route('/courseDocumentUpload', methods=['POST'])
@cross_origin()
def uploadCourseDocument():
    global courseFolderPath
    try:
        conn = connect_mysql()
        print("Hellooo", conn)
        cursor = conn.cursor()
        print("-------->",request.form)
        # Extract form data
        contentFile = request.form.get('courseName')
        content = contentFile.split('.')[0]
        # fileType = request.form.get('fileType')
        
        file = request.files.get('file')
        contentExtension = file.filename.split('.')[-1].lower()

        print("File---> ", file)
        if not file:
            return {"error": "No file uploaded"}, 400
        

        path = os.path.join(UPLOAD_COURSE_PATH, content)
        # Generate current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        if not os.path.exists(path):
            os.makedirs(path)
        
        fileName = f"{content}_{timestamp}.{contentExtension}"
        

        

  
        file.save(os.path.join(path, fileName))

        folderName = f"{content}/{fileName}"

        print("folderName-->", folderName)

        courseFolderPath = folderName
        print("courseFolderPath", courseFolderPath)




        return {"success": "File uploaded successfully"}, 200
        
    except Exception as e:
        print(e)
        return {"error": "An error occurred during file upload"}, 500





# @course.route('/getAllBatchList', methods=['POST'])
# @cross_origin()
# def getAllBatchList():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # sql = "SELECT id, batchName, mentorName, courseType, activeFlag, updatedOn FROM `batchMaster`"
#         # sql = "SELECT id, batchName, mentorName, courseType, activeFlag, updatedOn FROM `batchMaster`ORDER BY CASE WHEN activeFlag = 'pending' THEN 0 ELSE 1 END, updatedOn DESC"
#         sql = """
#            SELECT id, batchName, mentorName, courseType, activeFlag, updatedOn 
# FROM `batchMaster`
# ORDER BY 
#     CASE activeFlag 
#         WHEN '2' THEN 0 
#         WHEN '1' THEN 1 
#         WHEN '0' THEN 2 
#         ELSE 3  -- Handle any other cases if necessary
#     END,
#     updatedOn DESC
#               """
#         cursor.execute(sql)
#         batchList = cursor.fetchall()
        
#         processed_batch_list = []
#         for batch in batchList:
#             print("batcchhh", batch)
#             try:
#                 # for mentor in batch['mentorName']:
#                 #     print("mentorrrrrrrr", mentor)
#                 mentor_names = [name.strip() for name in batch['mentorName'].split('/')]
#                 # mentor_names = [mentor['name'].strip() for mentor in batch['mentorName']]
#                 print("menotrr", mentor_names)
#                 batch_dict = {
#                     "batchName": batch['batchName'],
#                     "mentorName": mentor_names,
#                     "courseType": batch['courseType'],
#                     "activeFlag": batch['activeFlag'],
#                     "id":batch['id'],
#                     "updatedOn": batch['updatedOn'].strftime('%Y-%m-%d') if batch['updatedOn'] else None
                    
#                 }
#                 processed_batch_list.append(batch_dict)
#             except Exception as e:
#                 print("Error processing batch:", e)

#         resp = jsonify({
#             "status": 1,
#             "success": True,
#             "result": processed_batch_list,
#             "message": "Fetched successfully"
#         })

#     except Exception as e:
#         print("Error occurred:", e)
#         resp = jsonify({"status": 0, "success": False, "message": "An error occurred"})

#     finally:
#         cursor.close()
#         conn.close()
#         return resp


@course.route('/getAllBatchList', methods=['POST'])
@cross_origin()
def getAllBatchList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        sql = """
            SELECT id, batchName, mentorName ,mentorDetails, courseType, activeFlag, updatedOn 
            FROM `batchMaster`
            ORDER BY 
                CASE activeFlag 
                    WHEN '2' THEN 0 
                    WHEN '1' THEN 1 
                    WHEN '0' THEN 2 
                    ELSE 3
                END,
                updatedOn DESC
        """
        cursor.execute(sql)
        batchList = cursor.fetchall()
        

        processed_batch_list = []
        for batch in batchList:
            try:
                batchName = batch['batchName']
                sqlForNoOfStudent = """
                SELECT COUNT(id) as NoOfStudent FROM `converted_student_data` WHERE batch=%s;
                """
                cursor.execute(sqlForNoOfStudent, (batchName,))
                NoOfStudentResult = cursor.fetchall()

                if NoOfStudentResult:
                    NoOfStudent = NoOfStudentResult[0]['NoOfStudent']
                else:
                    NoOfStudent = 0 

                print(NoOfStudent)
                
                mentor_names = []
                mentorDetail = []

                if batch['mentorDetails']:
                    print("batch['mentorDetails']", batch['mentorDetails'])
                    mentorDetail = json.loads(batch['mentorDetails'])
            
                if batch['mentorName']:
                    previous_mentor_names = [name.strip() for name in batch['mentorName'].split('/')]
                    mentor_names.extend(previous_mentor_names)
                
                batch_dict = {
                    "batchName": batch['batchName'],
                    "mentorName": mentor_names,
                    "courseType": batch['courseType'],
                    "activeFlag": batch['activeFlag'],
                    "id": batch['id'],
                    "mentorDetails" : mentorDetail,
                    "NoOfStudent" : NoOfStudent,
                    "updatedOn": batch['updatedOn'].strftime('%Y-%m-%d') if batch['updatedOn'] else None
                }
                processed_batch_list.append(batch_dict)
            except Exception as e:
                print("Error processing batch:", e)

        resp = jsonify({
            "status": 1,
            "success": True,
            "result": processed_batch_list,
            "message": "Fetched successfully"
        })

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred"})

    finally:
        cursor.close()
        conn.close()
        return resp# @course.route('/addBatch', methods=['POST'])
# @cross_origin()
# def addBatch():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.get_json()
#         print("data", data)
#         batchName = data.get('batchName')
#         print("batchName", batchName)

#         mentorNames_list = data.get('mentorName', []) 
#         print("mentorNames_list", mentorNames_list)
        
#         courseType = data.get('courseName')
#         print(courseType)
#         # activeFlag = data.get('activeFlag')
#         activeFlag = data.get('activeFlag')
#         updatedOn = datetime.now().strftime('%Y-%m-%d')

#         if not (batchName and mentorNames_list and courseType):
#             raise ValueError("BatchName, mentorNames, and courseType are required fields.")

#         mentorName = '/'.join(mentorNames_list)

      
#         sql = "INSERT INTO `batchMaster` (batchName, mentorName, courseType, activeFlag, updatedOn) VALUES (%s, %s, %s, %s, %s)"
#         cursor.execute(sql, (batchName, mentorName, courseType, activeFlag, updatedOn))
#         conn.commit()

#         resp = jsonify({
#             "status": 1,
#             "success": True,
#             "message": "Batch added successfully"
#         })

#     except ValueError as ve:
#         print("Validation Error:", ve)
#         resp = jsonify({"status": 0, "success": False, "message": str(ve)})

#     except Exception as e:
#         print("Error occurred:", e)
#         resp = jsonify({"status": 0, "success": False, "message": "An error occurred while adding batch"})

#     finally:
#         cursor.close()
#         conn.close()
#         return resp


# @course.route('/addBatch', methods=['POST'])
# @cross_origin()
# def addBatch():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.get_json()
#         print("data", data)
#         batchName = data.get('batchName')
#         print("batchName", batchName)

#         mentorDetails = data.get('mentorName')  # assuming this is the correct key
#         print("------------------>",mentorDetails)

#         # Constructing JSON object for mentorDetails
#         mentorNameD = []
#         mentor_details_json = []
#         for mentor in mentorDetails:
#             mentorNameD.append(mentor['name'])

#             mentor_details_json.append({'id': mentor['id'], 'name': mentor['name']})
#         mentorName = '/'.join(mentorNameD)
#         mentor_details_json_str = json.dumps(mentor_details_json)

#         courseType = data.get('courseName')
#         activeFlag = data.get('activeFlag', '2')
#         updatedOn = datetime.now().strftime('%Y-%m-%d')


      
#         sql = "INSERT INTO `batchMaster` (batchName, mentorDetails, mentorName, courseType, activeFlag, updatedOn) VALUES (%s, %s, %s, %s, %s, %s)"
#         cursor.execute(sql, (batchName, mentor_details_json_str, mentorName, courseType, activeFlag, updatedOn))
#         conn.commit()

#         resp = jsonify({
#             "status": 1,
#             "success": True,
#             "message": "Batch added successfully"
#         })

#     except ValueError as ve:
#         print("Validation Error:", ve)
#         resp = jsonify({"status": 0, "success": False, "message": str(ve)})

#     except Exception as e:
#         print("Error occurred:", e)
#         resp = jsonify({"status": 0, "success": False, "message": "An error occurred while adding batch"})

#     finally:
#         cursor.close()
#         conn.close()
#         return resp


@course.route('/addBatch', methods=['POST'])
@cross_origin()
def addBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.get_json()
        print("data", data)
        batchName = data.get('batchName')
        print("batchName", batchName)
        startDate = data.get('startDate')
        endDate = data.get('endDate')
        mentorDetails = data.get('mentorName')  # assuming this is the correct key
        print("------------------>",mentorDetails)

        # Constructing JSON object for mentorDetails
        mentorNameD = []
        mentor_details_json = []
        for mentor in mentorDetails:
            mentorNameD.append(mentor['name'])

            mentor_details_json.append({'id': mentor['id'], 'name': mentor['name']})
        mentorName = '/'.join(mentorNameD)
        mentor_details_json_str = json.dumps(mentor_details_json)

        courseType = data.get('courseName')
        activeFlag = data.get('activeFlag', '2')
        updatedOn = datetime.now().strftime('%Y-%m-%d')


      
        sql = "INSERT INTO `batchMaster` (batchName, mentorDetails, mentorName, courseType, activeFlag, updatedOn, startingDate, batchTentiveEndingDate) VALUES (%s, %s, %s, %s, %s, %s,%s, %s)"
        cursor.execute(sql, (batchName, mentor_details_json_str, mentorName, courseType, activeFlag, updatedOn, startDate, endDate))
        conn.commit()

        resp = jsonify({
            "status": 1,
            "success": True,
            "message": "Batch added successfully"
        })

    except ValueError as ve:
        print("Validation Error:", ve)
        resp = jsonify({"status": 0, "success": False, "message": str(ve)})

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred while adding batch"})

    finally:
        cursor.close()
        conn.close()
        return resp


@course.route('/updateMentorBatch', methods=['POST'])
@cross_origin()
def addMentorBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.get_json()
        print("data", data)
        batchName = data.get('batchName')
        print("batchName", batchName)

        email = data.get('email', []) 
        print("emails_list", email)
        
        courseType = data.get('courseName')
        print(courseType)
        
        # activeFlag = data.get('activeFlag')
        updatedOn = datetime.now().strftime('%Y-%m-%d')

        if not (batchName and email and courseType):
            raise ValueError("BatchName, email, and courseType are required fields.")


        mentor_details = []
        # for email in emails_list:
        sql_mentor = "SELECT name FROM `mentorMaster` WHERE emailId = %s"
        cursor.execute(sql_mentor, (email))
        print(sql_mentor)
        mentor_result = cursor.fetchone()
        if mentor_result:
            mentor_details.append(mentor_result)
        else:
            # Handle case where mentor with given email is not found
            raise ValueError(f"Mentor with email '{email}' not found.")

        # Prepare mentor names and IDs for insertion
        # mentor_ids = [str(mentor['id']) for mentor in mentor_details]
        mentor_names = [mentor['name'] for mentor in mentor_details]
        mentor_name_str = '/'.join(mentor_names)
      
        sql_insert = "INSERT INTO `batchMaster` (batchName, mentorName, courseType,  updatedOn) VALUES (%s, %s, %s, %s, %s)"
        cursor.execute(sql_insert, (batchName, mentor_name_str, courseType, updatedOn))
        conn.commit()

        resp = jsonify({
            "status": 1,
            "success": True,
            "message": "Batch added successfully"
        })

    except ValueError as ve:
        print("Validation Error:", ve)
        resp = jsonify({"status": 0, "success": False, "message": str(ve)})

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred while adding batch"})

    finally:
        cursor.close()
        conn.close()
        return resp


@course.route('/deleteBatch', methods=['POST'])
@cross_origin()
def deleteBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.get_json()
        batchId = data.get('id')

        if not batchId:
            raise ValueError("BatchId is required for deletion.")

        sql = "DELETE FROM `batchMaster` WHERE id = %s"
        cursor.execute(sql, (batchId,))
        conn.commit()

        if cursor.rowcount == 0:
            raise ValueError(f"Batch with BatchId {batchId} does not exist.")

        resp = jsonify({
            "status": 1,
            "success": True,
            "message": f"Batch with BatchId {batchId} deleted successfully"
        })

    except ValueError as ve:
        print("Validation Error:", ve)
        resp = jsonify({"status": 0, "success": False, "message": str(ve)})

    # except mysql.connector.Error as mysql_error:
    #     print("MySQL Error:", mysql_error)
    #     resp = jsonify({"status": 0, "success": False, "message": "MySQL error occurred while deleting batch"})

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred while deleting batch"})

    finally:
        # Close cursor and connection
        cursor.close()
        conn.close()

        return resp

# @course.route('/editBatch', methods=['POST'])
# @cross_origin()
# def editBatch():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.get_json()
#         print("data", data)
#         batchId = data.get('id')
#         batchName = data.get('batchName')
#         mentorNames_list = data.get('mentorName', [])
#         courseType = data.get('courseName')
#         # activeFlag = data.get('activeFlag')
#         activeFlag = data.get('activeFlag', '2')
#         updatedOn = datetime.now().strftime('%Y-%m-%d')
#         print("The Time is ---> ", updatedOn)

#         if not (batchId and batchName and mentorNames_list and courseType):
#             raise ValueError("BatchId, batchName, mentorNames, and courseType are required fields.")

#         mentorName = '/'.join(mentorNames_list)

#         sql = "UPDATE `batchMaster` SET batchName=%s, mentorName=%s, courseType=%s, activeFlag=%s, updatedOn=%s WHERE id=%s"
#         cursor.execute(sql, (batchName, mentorName, courseType, activeFlag, updatedOn, batchId))
#         conn.commit()

#         if cursor.rowcount > 0:
#             resp = jsonify({
#                 "status": 1,
#                 "success": True,
#                 "message": "Batch updated successfully"
#             })
#         else:
#             resp = jsonify({
#                 "status": 0,
#                 "success": False,
#                 "message": "Batch with ID {} not found".format(batchId)
#             })

#     except ValueError as ve:
#         print("Validation Error:", ve)
#         resp = jsonify({"status": 0, "success": False, "message": str(ve)})

#     except Exception as e:
#         print("Error occurred:", e)
#         resp = jsonify({"status": 0, "success": False, "message": "An error occurred while updating batch"})

#     finally:
#         cursor.close()
#         conn.close()
#         return resp




# @course.route('/editBatch', methods=['POST'])
# @cross_origin()
# def editBatch():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.get_json()
#         print("data", data)
#         batchId = data.get('id')
#         batchName = data.get('batchName')
#         mentorDetails = data.get('mentorName')
        
        
#         # Extract mentor names and IDs from the array of objects
#         # mentorNames_list = []
#         # mentorIds_list = []
#         # for mentor in data.get('mentorName', []):
#         #     mentorNames_list.append(mentor['name'])
#         #     mentorIds_list.append(str(mentor['id']))  # Assuming mentor ID is stored as string in the DB

#         courseType = data.get('courseName')
#         activeFlag = data.get('activeFlag', '2')
#         updatedOn = datetime.now().strftime('%Y-%m-%d')
#         # print("The Time is ---> ", updatedOn)

#         # if not (batchId and batchName and mentorNames_list and courseType):
#         #     raise ValueError("BatchId, batchName, mentorNames, and courseType are required fields.")

#         # mentorName = '/'.join(mentorNames_list)
#         # mentorId = '/'.join(mentorIds_list)

#         sql = "UPDATE `batchMaster` SET batchName=%s, mentorDetails=JSON_OBJECT(%s), courseType=%s, activeFlag=%s, updatedOn=%s WHERE id=%s"
#         cursor.execute(sql, (batchName, json.dumps(mentorDetails),  courseType, activeFlag, updatedOn, batchId))
#         conn.commit()

#         if cursor.rowcount > 0:
#             resp = jsonify({
#                 "status": 1,
#                 "success": True,
#                 "message": "Batch updated successfully"
#             })
#         else:
#             resp = jsonify({
#                 "status": 0,
#                 "success": False,
#                 "message": "Batch with ID {} not found".format(batchId)
#             })

#     except ValueError as ve:
#         print("Validation Error:", ve)
#         resp = jsonify({"status": 0, "success": False, "message": str(ve)})

#     except Exception as e:
#         print("Error occurred:", e)
#         resp = jsonify({"status": 0, "success": False, "message": "An error occurred while updating batch"})

#     finally:
#         cursor.close()
#         conn.close()
#         return resp




@course.route('/editBatch', methods=['POST'])
@cross_origin()
def editBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.get_json()
        print("data", data)
        batchId = data.get('id')
        batchName = data.get('batchName')
        mentorDetails = data.get('mentorName')  # assuming this is the correct key
        print("------------------>",mentorDetails)

        # Constructing JSON object for mentorDetails
        mentorNameD = []
        mentor_details_json = []
        for mentor in mentorDetails:
            mentorNameD.append(mentor['name'])

            mentor_details_json.append({'id': mentor['id'], 'name': mentor['name']})
        mentorName = '/'.join(mentorNameD)
        mentor_details_json_str = json.dumps(mentor_details_json)

        courseType = data.get('courseName')
        activeFlag = data.get('activeFlag', '2')
        updatedOn = datetime.now().strftime('%Y-%m-%d')

        sql = "UPDATE `batchMaster` SET batchName=%s, mentorName= %s,mentorDetails=%s, courseType=%s, activeFlag=%s, updatedOn=%s WHERE id=%s"
        cursor.execute(sql, (batchName, mentorName, mentor_details_json_str, courseType, activeFlag, updatedOn, batchId))
        conn.commit()

        if cursor.rowcount > 0:
            resp = jsonify({
                "status": 1,
                "success": True,
                "message": "Batch updated successfully"
            })
        else:
            resp = jsonify({
                "status": 0,
                "success": False,
                "message": "Batch with ID {} not found".format(batchId)
            })

    except ValueError as ve:
        print("Validation Error:", ve)
        resp = jsonify({"status": 0, "success": False, "message": str(ve)})

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred while updating batch"})

    finally:
        cursor.close()
        conn.close()
        return resp


@course.route('/send-email', methods=['POST'])
def send_email():
    try:
        data = request.get_json()

        id = data.get('id')
        # id = '18'
        print(id)
        email_ids = data.get('email', [])
        print("email_ids", email_ids)

        if not id or not email_ids:
            return jsonify({"error": "Missing required parameters"}), 400

        # Connect to MySQL and fetch email addresses based on id
        conn = connect_mysql()
        cursor = conn.cursor()
        
        # Example SQL query assuming 'email' is the column name in 'users' table
        sql = "SELECT `courseName`, `code`, `description`, `folderName` FROM `courseMaster` WHERE id = %s"
        cursor.execute(sql, (id,))
        email_result = cursor.fetchone()
        print("Email Result", email_result)

        if not email_result:
            return jsonify({"error": f"No email found for id {id}"}), 404

        subject = "Course Details"
        text_body = f"Course Name: {email_result['courseName']}\nCode: {email_result['code']}\nDescription: {email_result['description']}"
        file_path = os.path.join(UPLOAD_COURSE_PATH, email_result['folderName'])
        # file_path = r'static/courseDetails/course/course.pdf'
        print(file_path)
        html_text = f'''<span>Hi,<br><br>
                        {email_result['courseName']}.<br>
                        Code: {email_result['code']}<br>
                        Description: {email_result['description']}<br><br>
                        Please find the attachment</a>.<br><br>
                        Best regards,<br>
                        Techno Struct Academy</span>'''

        # cursor.close()
        # conn.close()

        obj = sendMail(email_ids, text_body, subject, html_text, file_path)
        obj.sendmail()
        resp = jsonify({"status":1, "success":True,   "message": "Email Sent Successfully" })
        
            

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    finally:
        cursor.close()
        conn.close()
        return resp



#MentorCourseList section
@course.route('/assignCourseToMentor', methods=['POST'])
@cross_origin()
def assignCourseToMentor():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.get_json()
        mentorId = data.get('id')
        courses_list = data.get('courseList', [])
        print(courses_list)

        if not (mentorId and courses_list):
            raise ValueError("Mentor ID and courses list are required fields.")

        # Convert courses list to a string format for MySQL storage
        # courses_str = ','.join(courses_list)

        # updatedOn = datetime.now().strftime('%Y-%m-%d')

        sql = "UPDATE mentorMaster SET courses=%s WHERE id=%s"
        cursor.execute(sql, (str(courses_list), mentorId))
        conn.commit()

        if cursor.rowcount > 0:
            resp = jsonify({
                "status": 1,
                "success": True,
                "message": "Courses assigned to mentor successfully"
            })
        else:
            resp = jsonify({
                "status": 0,
                "success": False,
                "message": "Mentor with ID {} not found".format(mentorId)
            })

    except ValueError as ve:
        print("Validation Error:", ve)
        resp = jsonify({"status": 0, "success": False, "message": str(ve)})

    except Exception as e:
        print("Error occurred:", e)
        resp = jsonify({"status": 0, "success": False, "message": "An error occurred while assigning courses to mentor"})

    finally:
        cursor.close()
        conn.close()
        return resp





# @course.route('/courseXlslDocumentUpload', methods=['POST'])
# @cross_origin()
# def courseXlslDocumentUpload():
#     global courseFolderPath

#     try:
#         # Extract form data
#         courseName = request.form.get('courseName')
#         file = request.files.get('file')
        
#         if not file or not courseName:
#             return {"error": "Missing file or courseName"}, 400

#         contentExtension = file.filename.split('.')[-1].lower()
#         if contentExtension not in ['xlsx', 'xls']:
#             return {"error": "Invalid file format"}, 400

#         path = os.path.join(UPLOAD_COURSE_PATH, courseName)
#         timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

#         if not os.path.exists(path):
#             os.makedirs(path)

#         excel_file_name = f"{courseName}_{timestamp}.{contentExtension}"
#         filePath = os.path.join(path, excel_file_name)
#         file.save(filePath)

#         # Load the Excel file
#         df = pd.read_excel(filePath)

#         # Initialize the dictionary to hold all courses
#         courses = {}

#         # Variables to keep track of current course and module
#         current_course_id = None
#         current_module = None

#         for idx, row in df.iterrows():
#             course_id = row['id']
#             course_name = row['Course']
#             module = row['Module']
#             day = row['Days']
#             courseContent = row['Content']
#             duration = row['Duration(hrs)']
            
#             # Check if the course ID is not missing
#             if pd.notna(course_id):
#                 current_course_id = int(course_id)
#                 if current_course_id not in courses:
#                     courses[current_course_id] = {
#                         'CourseId': current_course_id,
#                         'CourseName': course_name,
#                         'Modules': []
#                     }
            
#             # Check if the module name is not missing
#             if pd.notna(module):
#                 current_module = {
#                     'Module': module,
#                     'Days': []
#                 }
#                 courses[current_course_id]['Modules'].append(current_module)
            
#             # Append day, content, and duration to the current module
#             if pd.notna(day):
#                 current_module['Days'].append({
#                     'Day': day,
#                     'Content': courseContent,
#                     'Duration': duration
#                 })

#         # Function to create PDF
#         def create_pdf(data, file_name):
#             c = canvas.Canvas(file_name, pagesize=letter)
#             width, height = letter
#             c.setFont("Helvetica", 12)
            
#             margin = 40
#             y = height - margin
            
#             for course_id, course_info in data.items():
#                 c.setFont("Helvetica-Bold", 14)
#                 c.drawString(margin, y, f"Course ID: {course_info['CourseId']}")
#                 y -= 20
#                 c.drawString(margin, y, f"Course Name: {course_info['CourseName']}")
#                 y -= 20
#                 c.setFont("Helvetica", 12)
                
#                 for module in course_info['Modules']:
#                     c.setFont("Helvetica-Bold", 12)
#                     c.drawString(margin + 20, y, f"Module: {module['Module']}")
#                     y -= 20
#                     c.setFont("Helvetica", 12)
#                     for day in module['Days']:
#                         c.drawString(margin + 40, y, f"Day: {day['Day']}, Content: {day['Content']}, Duration: {day['Duration']} hours")
#                         y -= 20
                        
#                         if y < margin:
#                             c.showPage()
#                             y = height - margin
#                             c.setFont("Helvetica", 12)
            
#             c.save()

#         # Create PDF
#         pdf_file_name = f"{courseName}_{timestamp}.pdf"
#         pdf_file_path = os.path.join(path, pdf_file_name)
#         create_pdf(courses, pdf_file_path)
#         pdfFilePath = f"{courseName}/{pdf_file_name}"
#         courseFolderPath= pdfFilePath


#         return {"success": "File processed and PDF created successfully", "pdf_path": pdf_file_path}, 200

#     except Exception as e:
#         print(e)
#         return {"error": "An error occurred during file upload or processing"}, 500
    




# @course.route('/courseModuleUpload', methods=['POST'])
# @cross_origin()
# def courseModuleUpload():
#     try:
#         data = request.get_json()
#         print("--->",data)
#         if not data:
#             return {"error": "No data provided"}, 400

#         conn = connect_mysql()
#         cursor = conn.cursor()

#         for course in data:
#             courseId = course['courseId']
#             courseName = course['courseName']
#             contentName = course['contentName']
#             contentDuration = int(course['contentDuration'])

#             for subModule in course['subModules']:
#                 subModuleNm = subModule['subModuleNm']
#                 contentDetails = subModule['contentDetails']
#                 moduleDuration = int(subModule['moduleDuration'])

#                 # Insert data into CourseData table
#                 cursor.execute(
#                     """
#                     INSERT INTO courseContentMaster (courseId, courseName, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration)
#                     VALUES (%s, %s, %s, %s, %s, %s, %s)
#                     """, (courseId, courseName, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration)
#                 )

#         conn.commit()
#         cursor.close()
#         conn.close()

#         return jsonify(data), 200

#     except Exception as e:
#         print(e)
#         return {"error": "An error occurred during data processing"}, 500


@course.route('/courseModuleUpload', methods=['POST'])
@cross_origin()
def courseModuleUpload():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)


         # Validate all incoming data before any database operations
        for course in data:
            contentName = course.get('contentName')
            contentId = course.get('contentId')
            if not contentName or not contentId:
                return {"error": "Blank contentName or contentId found, aborting operation."}, 400

            subModules = course.get('subModules')
            if not subModules:
                return {"error": "No subModules provided."}, 400
            
            for subModule in subModules:
                subModuleNm = subModule.get('subModuleNm')
                if not subModuleNm:
                    return {"error": "Blank subModuleNm found, aborting operation."}, 400


        

        if not isinstance(data, list):
            return {"error": "Invalid data format. Expected a list."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        for course in data:
            if not isinstance(course, dict):
                return {"error": "Invalid course format. Expected a dictionary."}, 400

            try:
                courseId = course['courseId']
                courseName = course['courseName']
                contentName = course['contentName']
                contentId = course['contentId']
                contentDuration = course['contentDuration']  # Keep as string

                if 'subModules' not in course or not isinstance(course['subModules'], list):
                    return {"error": "Invalid subModules format. Expected a list."}, 400

                for subModule in course['subModules']:
                    if not isinstance(subModule, dict):
                        return {"error": "Invalid subModule format. Expected a dictionary."}, 400

                    try:
                        subModuleNm = subModule['subModuleNm']
                        contentDetails = subModule['contentDetails']
                        moduleDuration = subModule['moduleDuration']  # Keep as string

                        # Insert data into CourseContentMaster table
                        cursor.execute(
                            """
                            INSERT INTO courseContentMaster (courseId, courseName, contentName, contentId, contentDuration, subModuleNm, contentDetails, moduleDuration)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """, (courseId, courseName, contentName, contentId,contentDuration, subModuleNm, contentDetails, moduleDuration)
                        )
                    except KeyError as e:
                        print(f"KeyError in subModule: {e}")
                        return {"error": f"Missing key in subModule: {e}"}, 400
                    except Exception as e:
                        print(f"Error in subModule processing: {e}")
                        return {"error": f"Error in subModule processing: {e}"}, 500
            except KeyError as e:
                print(f"KeyError in course: {e}")
                return {"error": f"Missing key in course: {e}"}, 400
            except Exception as e:
                print(f"Error in course processing: {e}")
                return {"error": f"Error in course processing: {e}"}, 500

        conn.commit()

        courseNamePDF = ''
        for course in data:
            courseNamePDF = course['courseName']
            
            
        pdfSqlQuery = """
            SELECT courseId, courseName, contentId, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration 
            FROM courseContentMaster
            WHERE courseName=%s
        """

        print("------------------>",courseNamePDF)
        cursor.execute(pdfSqlQuery,courseNamePDF)
        rows = cursor.fetchall()
        # print(rows)
        

        course_data = {}
        for row in rows:
            print(row)
            courseId = row['courseId']
            if courseId not in course_data:
                course_data[courseId] = {
                    "courseName": row['courseName'],
                    "contents": []
                }
            course_data[courseId]["contents"].append({
                "contentId": row['contentId'],
                "moduleDuration": row['moduleDuration'],                
                "contentDuration": row['contentDuration'],
                "subModuleNm": row['subModuleNm'],
                "contentDetails": row['contentDetails'],
                "contentName": row['contentName']
                
            })
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        # print(course_data)
        # Create PDF
        pdf_folder_path = os.path.join(UPLOAD_COURSE_PATH, courseName)
        if not os.path.exists(pdf_folder_path):
            os.makedirs(pdf_folder_path)
        pdf_file_name = f"{courseName}_{timestamp}.pdf"
        print('pdfFilePath---', pdf_file_name)
        pdfFile = f"{pdf_folder_path}/{pdf_file_name}"
        pdfCreation(course_data, pdfFile)
        pdfNameDB =f"{courseName}/{pdf_file_name}"

        print("bfnkd-----------", courseNamePDF)
        # Retrieve existing folderName from the database
        selectSql = "SELECT folderName FROM courseMaster WHERE courseName = %s"

        cursor.execute(selectSql,courseNamePDF)
        result = cursor.fetchone()

        print("45678987654",result)
        
        if result and result['folderName'] is not None:
            existingFolderName = result['folderName']
        else:
            existingFolderName = ''

        # Initialize courseFolderName
        if existingFolderName:
            courseFolderName = f"{existingFolderName},{pdfNameDB}"
        else:
            courseFolderName = pdfNameDB

        updateSql = "UPDATE courseMaster SET folderName = %s WHERE courseName = %s"

        cursor.execute(updateSql, (courseFolderName, courseNamePDF))
    # connection.commit()


        conn.commit()
        cursor.close()
        conn.close()
        

        return jsonify({"message": "Data inserted successfully"}), 200

    # except mysql.connector.Error as err:
    #     print(f"Database error: {err}")
    #     return {"error": "Database error"}, 500
    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data processing: {e}"}, 500
    


@course.route('/getCourseModuleList', methods=['POST'])
@cross_origin()
def getCourseModuleList():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)

        courseId = data.get('courseId')
        courseName = data.get('courseName')

        if not courseId or not courseName:
            return {"error": "Missing courseId or courseName in the request."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch data from CourseContentMaster table
        query = """
            SELECT DISTINCT(contentName), courseId, courseName FROM courseContentMaster
            WHERE courseId = %s AND courseName = %s
        """
        cursor.execute(query, (courseId, courseName))
        rows = cursor.fetchall()
        print(rows)

        if not rows:
            return {"error": "No data found for the given courseId and courseName."}, 404


        return jsonify(rows), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data retrieval: {e}"}, 500
    

#Get all Module Subcontent
# @course.route('/getCourseModuleData', methods=['POST'])
# @cross_origin()
# def getCourseModuleData():
#     try:
#         # Get JSON data from the request
#         data = request.get_json()
#         print("---> Received data:", data)

#         courseId = data.get('courseId')
#         courseName = data.get('courseName')

#         if not courseId or not courseName:
#             return {"error": "Missing courseId or courseName in the request."}, 400

#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # Fetch data from CourseContentMaster table
#         query = """
#             SELECT * FROM courseContentMaster
#             WHERE courseId = %s AND courseName = %s
#         """
#         cursor.execute(query, (courseId, courseName))
#         rows = cursor.fetchall()

#         if not rows:
#             return {"error": "No data found for the given courseId and courseName."}, 404

#         # Get column names from the cursor description
#         columns = [desc[0] for desc in cursor.description]
#         print("---> Columns from DB:", columns)  # Debug statement

#         # Convert rows to dictionaries
#         data_list = [dict(zip(columns, row)) for row in rows]
#         print("---> Fetched data from DB:", data_list)  # Debug statement

#         cursor.close()
#         conn.close()

#         # Structure the data
#         result = {}
#         for row in rows:
#             contentId = row['contentId']
#             if contentId not in result:
#                 result[contentId] = {
                    
#                     'courseId': row['courseId'],
#                     'courseName': row['courseName'],
#                     'contentName': row['contentName'],
#                     'contentId': contentId,
#                     'contentDuration': row['contentDuration'],
#                     'subModules': []
#                 }

#             subModule = {
#                 'subModuleId' : row['id'],
#                 'subModuleNm': row['subModuleNm'],
#                 'contentDetails': row['contentDetails'],
#                 'moduleDuration': row['moduleDuration']
#             }
#             result[contentId]['subModules'].append(subModule)

#         # Convert result to list
#         response_data = list(result.values())
#         print("---> Structured response data:", response_data)  # Debug statement

#         return jsonify(response_data), 200

#     except Exception as e:
#         print(f"An error occurred: {e}")
#         return {"error": f"An error occurred during data retrieval: {e}"}, 500


@course.route('/getCourseModuleData', methods=['POST'])
@cross_origin()
def getCourseModuleData():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)

        courseId = data.get('courseId')
        courseName = data.get('courseName')
        contentName = data.get('contentName')

        if not contentName:
            return {"error": "Missing courseId, courseName, or contentName in the request."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch data from CourseContentMaster table
        query = """
            SELECT * FROM courseContentMaster
            WHERE  contentName = %s AND courseName = %s AND courseId = %s
        """
        cursor.execute(query, (contentName,courseName,courseId))
        rows = cursor.fetchall()

        if not rows:
            return {"error": "No data found for the given courseId, courseName, and contentName."}, 404

        cursor.close()
        conn.close()

        # Structure the data
        result = {}
        for row in rows:
            contentId = row['contentId']
            if contentId not in result:
                result[contentId] = {
                    'courseId': row['courseId'],
                    'courseName': row['courseName'],
                    'contentName': row['contentName'],
                    'contentId': contentId,
                    'contentDuration': row['contentDuration'],
                    'subModules': []
                }

            subModule = {
                'subModuleId': row['id'],
                'subModuleNm': row['subModuleNm'],
                # 'contentDetails': row['contentDetails'],
                # 'moduleDuration': row['moduleDuration']
            }
            result[contentId]['subModules'].append(subModule)

        # Convert result to list
        response_data = list(result.values())
        print("---> Structured response data:", response_data)  # Debug statement

        return jsonify(response_data), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data retrieval: {e}"}, 500







# @course.route('/getCourseModuleDataBySubModuleName', methods=['POST'])
# @cross_origin()
# def getCourseModuleDataBySubModuleName():
#     try:
#         # Get JSON data from the request
#         data = request.get_json()
#         print("---> Received data:", data)


#         subModuleId = data.get('subModuleId')
#         subModuleNm = data.get('subModuleNm')


#         if not subModuleId:
#             return {"error": "Missing subModuleId in the request."}, 400

#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # Fetch data from CourseContentMaster table
#         query = """
#             SELECT * FROM courseContentMaster
#             WHERE  id = %s AND subModuleNm = %s 
#         """
#         cursor.execute(query, (subModuleId, subModuleNm))
#         rows = cursor.fetchall()

#         if not rows:
#             return {"error": "No data found for the given courseId, courseName"}, 404

#         cursor.close()
#         conn.close()

#         # Structure the data
#         result = {}
#         for row in rows:
#             contentId = row['contentId']
#             if contentId not in result:
#                 result[contentId] = {
#                     'courseId': row['courseId'],
#                     'courseName': row['courseName'],
#                     'contentName': row['contentName'],
#                     'contentId': contentId,
#                     'contentDuration': row['contentDuration'],
#                     'subModules': []
#                 }

#             subModule = {
#                 'subModuleId' : row['id'],
#                 'contentDetails': row['contentDetails'],
#                 'moduleDuration': row['moduleDuration']
#             }
#             result[contentId]['subModules'].append(subModule)

#         # Convert result to list
#         response_data = list(result.values())
#         print("---> Structured response data:", response_data)  # Debug statement

#         return jsonify(response_data), 200

#     except Exception as e:
#         print(f"An error occurred: {e}")
#         return {"error": f"An error occurred during data retrieval: {e}"}, 500

@course.route('/getCourseModuleDataBySubModuleName', methods=['POST'])
@cross_origin()
def getCourseModuleDataBySubModuleName():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)

        subModuleId = data.get('subModuleId')
        subModuleNm = data.get('subModuleNm')

        if not subModuleId:
            return {"error": "Missing subModuleId in the request."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch data from CourseContentMaster table
        query = """
            SELECT * FROM courseContentMaster
            WHERE id = %s AND subModuleNm = %s 
        """
        cursor.execute(query, (subModuleId, subModuleNm))
        rows = cursor.fetchall()

        if not rows:
            return {"error": "No data found for the given subModuleId and subModuleNm"}, 404

        cursor.close()
        conn.close()

        # Structure the data
        result = {}
        for row in rows:
            contentId = row['contentId']
            if contentId not in result:
                result[contentId] = {
                    'courseId': row['courseId'],
                    'courseName': row['courseName'],
                    'contentName': row['contentName'],
                    'contentId': contentId,
                    'contentDuration': row['contentDuration'],
                    'subModules': {}
                }

            subModule = {
                'subModuleId': row['id'],
                'subModuleNm': row['subModuleNm'],
                'contentDetails': row['contentDetails'],
                'moduleDuration': row['moduleDuration']
            }
            result[contentId]['subModules']= subModule

        # Convert result to list
        response_data = list(result.values())
        print("---> Structured response data:", response_data)  # Debug statement

        return jsonify(response_data), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data retrieval: {e}"}, 500


@course.route('/editCourseModuleData', methods=['POST'])
@cross_origin()
def editCourseModuleData():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)

        subModuleId = data.get('subModuleId')
        # subModuleNm = data.get('subModuleNm')
        updatedContentDetails = data.get('contentDetails')
        updatedModuleDuration = data.get('moduleDuration')

        if not subModuleId:
            return {"error": "Missing subModuleId or subModuleNm in the request."}, 400

        if updatedContentDetails is None or updatedModuleDuration is None:
            return {"error": "Missing data to update."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        # Update data in CourseContentMaster table
        query = """
            UPDATE courseContentMaster
            SET contentDetails = %s, moduleDuration = %s
            WHERE id = %s 
        """
        cursor.execute(query, (updatedContentDetails, updatedModuleDuration, subModuleId))
        conn.commit()

        if cursor.rowcount == 0:
            return {"error": "No record found to update for the given subModuleId and subModuleNm."}, 404
        

        getID = """
                SELECT courseName     
                FROM courseContentMaster
            WHERE id=%s
            """
        cursor.execute(getID,subModuleId)
        courseNameID = cursor.fetchall()   
        print("courseNameID", courseNameID)


        courseNamePDF = ''
        for courseContentName in courseNameID:
            courseNamePDF = courseContentName['courseName']
            
        pdfSqlQuery = """
            SELECT courseId, courseName, contentId, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration 
            FROM courseContentMaster
            WHERE courseName=%s
        """

        print("------------------>",courseNamePDF)
        cursor.execute(pdfSqlQuery,courseNamePDF)
        rows = cursor.fetchall()
        # print(rows)
        

        course_data = {}
        for row in rows:
            print(row)
            courseId = row['courseId']
            if courseId not in course_data:
                course_data[courseId] = {
                    "courseName": row['courseName'],
                    "contents": []
                }
            course_data[courseId]["contents"].append({
                "contentId": row['contentId'],
                "moduleDuration": row['moduleDuration'],                
                "contentDuration": row['contentDuration'],
                "subModuleNm": row['subModuleNm'],
                "contentDetails": row['contentDetails'],
                "contentName": row['contentName']
                
            })
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        # print(course_data)
        # Create PDF
        pdf_folder_path = os.path.join(UPLOAD_COURSE_PATH, courseNamePDF)
        if not os.path.exists(pdf_folder_path):
            os.makedirs(pdf_folder_path)
        pdf_file_name = f"{courseNamePDF}_{timestamp}.pdf"
        print('pdfFilePath---', pdf_file_name)
        pdfFile = f"{pdf_folder_path}/{pdf_file_name}"
        pdfCreation(course_data, pdfFile)
        pdfNameDB =f"{courseNamePDF}/{pdf_file_name}"

        print("bfnkd-----------", courseNamePDF)
        # Retrieve existing folderName from the database
        selectSql = "SELECT folderName FROM courseMaster WHERE courseName = %s"

        cursor.execute(selectSql,courseNamePDF)
        result = cursor.fetchone()

        print("45678987654",result)
        
        if result and result['folderName'] is not None:
            existingFolderName = result['folderName']
        else:
            existingFolderName = ''

        # Initialize courseFolderName
        if existingFolderName:
            courseFolderName = f"{existingFolderName},{pdfNameDB}"
        else:
            courseFolderName = pdfNameDB

        updateSql = "UPDATE courseMaster SET folderName = %s WHERE courseName = %s"

        cursor.execute(updateSql, (courseFolderName, courseNamePDF))
    # connection.commit()


        conn.commit()
        cursor.close()
        conn.close()
        

        return jsonify({"message": "Data inserted successfully"}), 200

    # except mysql.connector.Error as err:
    #     print(f"Database error: {err}")
    #     return {"error": "Database error"}, 500
    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data processing: {e}"}, 500
    

        cursor.close()
        conn.close()

        return jsonify({"message": "Record updated successfully"}), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data update: {e}"}, 500


@course.route('/uploadCourseSubModuleDocumnet', methods=['POST'])
@cross_origin()
def uploadCourseSubModuleDocumnet():
    try:
        conn = connect_mysql()
        print("Hellooo", conn)
        cursor = conn.cursor()
        print("-------->",request.form)
        # Extract form data
        contentFileId = request.form.get('contentId')
        subModuleNm = request.form.get('subModuleNm')
        
        print("---->",contentFileId)

        
        file = request.files.get('file')
        contentExtension = file.filename.split('.')[-1].lower()

        print("File---> ", file)
        if not file:
            return {"error": "No file uploaded"}, 400
        
#----------------------------------- [{'contentName': 'Artificial Inteligenge'}]

        getContentName = "SELECT contentName, courseName FROM courseContentMaster WHERE id=%s"
        cursor.execute(getContentName, contentFileId)
        rows = cursor.fetchall()
        print("-----------------------------------",rows)

        contentNm = ''
        courseName = ''
        for row in rows:
            contentNm = row['contentName']
            courseName = row['courseName']

        path = os.path.join(UPLOAD_SUBMODULE_CONTENT_PATH, courseName, contentNm)
        # Generate current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        if not os.path.exists(path):
            os.makedirs(path)
        
        fileName = f"{subModuleNm}.{contentExtension}"

        file.save(os.path.join(path,fileName))

        folderName = f"{courseName}/{contentNm}/{fileName}"

        print("folderName-->", folderName)


        # Update data in CourseContentMaster table
        query = """
            UPDATE courseContentMaster
            SET contentDocPath = %s
            WHERE id = %s 
        """
        cursor.execute(query, (folderName,contentFileId))
        conn.commit()

        # if cursor.rowcount == 0:
        #     return {"error": "No record found to update for the given subModuleId and subModuleNm."}, 404

        cursor.close()
        conn.close()

        return jsonify({"message": "Record updated and File uploaded successfully"}), 200


        
    except Exception as e:
        print(e)
        return {"error": "An error occurred during file upload"}, 500

        
@course.route('/getSubModuleContentFolderPath', methods=['POST'])
@cross_origin()
def getSubModuleContentFolderPath():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)

        subModuleId = data.get('subModuleId')
        subModuleNm = data.get('subModuleNm')

        if not subModuleId:
            return {"error": "Missing subModuleId in the request."}, 400

        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch data from CourseContentMaster table
        query = """
            SELECT * FROM courseContentMaster
            WHERE id = %s
        """
        cursor.execute(query, (subModuleId))
        rows = cursor.fetchall()

        if not rows:
            return {"error": "No data found for the given subModuleId and subModuleNm"}, 404

        cursor.close()
        conn.close()

        # Structure the data
        # result = {}
        # for row in rows:
        #     contentId = row['contentId']
        #     if contentId not in result:
        #         result[contentId] = {
        #             'courseId': row['courseId'],
        #             'courseName': row['courseName'],
        #             'contentName': row['contentName'],
        #             'contentId': contentId,
        #             'contentDuration': row['contentDuration'],
        #             'subModules': {}
        #         }

        #     subModule = {
        #         'subModuleId': row['id'],
        #         'subModuleNm': row['subModuleNm'],
        #         'contentDetails': row['contentDetails'],
        #         'moduleDuration': row['moduleDuration'],
        #         'moduleDuration': row['moduleDuration']
        #     }
        #     result[contentId]['subModules']= subModule
        folder_names = [course['contentDocPath'] for course in rows]
        print("courseList['folderName']", folder_names)

       
        for course in rows:
            print("--------------> contentPath", course)
            folder_name = course['contentDocPath']
            if folder_name:
                folders_and_files = folder_name.split(',')  # Split by comma to get individual entries
                folder_file_list = []
                for entry in folders_and_files:
                    parts = entry.split('/')  # Split each entry by slash to get folder and file
                    if len(parts) == 3:
                        folder_file_dict = {
                            'path': entry.strip(),
                            'file': parts[-1].strip()
                        }
                        folder_file_list.append(folder_file_dict)

                course['contentDocPath'] = folder_file_list

    
        return jsonify(rows), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data retrieval: {e}"}, 500




# @course.route('/editMultiCourseModuleData', methods=['POST'])
# @cross_origin()
# def editMultiCourseModuleData():
#     try:
#         # Get JSON data from the request
#         data = request.get_json()
#         print("---> Received data:", data)

#         # Connect to the database
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # Iterate through the array of objects from the frontend
#         for item in data:
#             subModuleId = item.get('subModuleId')
#             updatedContentDetails = item.get('contentDetails')
#             updatedModuleDuration = item.get('moduleDuration')

#             if updatedContentDetails is None or updatedModuleDuration is None:
#                 return {"error": "Missing data to update or insert."}, 400

#             if subModuleId:
#                 # Update data in CourseContentMaster table
#                 query = """
#                     UPDATE courseContentMaster
#                     SET contentDetails = %s, moduleDuration = %s
#                     WHERE id = %s
#                 """
#                 cursor.execute(query, (updatedContentDetails, updatedModuleDuration, subModuleId))
#                 # if cursor.rowcount == 0:
#                 #     return {"error": f"No record found to update for subModuleId {subModuleId}."}, 404
#             else:
#                 # Insert new data into CourseContentMaster table
#                 query = """
#                     INSERT INTO courseContentMaster (contentDetails, moduleDuration)
#                     VALUES (%s, %s)
#                 """
#                 cursor.execute(query, (updatedContentDetails, updatedModuleDuration))

#         conn.commit()

#         cursor.close()
#         conn.close()

#         return jsonify({"message": "Records processed successfully"}), 200

#     except Exception as e:
#         print(f"An error occurred: {e}")
#         return {"error": f"An error occurred during data processing: {e}"}, 500




@course.route('/editMultipleCourseModuleData', methods=['POST'])
@cross_origin()
def editMultipleCourseModuleData():
    try:
        # Get JSON data from the request
        data = request.get_json()
    

        # Validate all incoming data before any database operations
        for course in data:
            subModules = course.get('subModules')
            if not subModules:
                return {"error": "No subModules provided."}, 400
            
            for subModule in subModules:
                subModuleNm = subModule.get('subModuleNm')
                if not subModuleNm:
                    return {"error": "Blank subModuleNm found, aborting operation."}, 400

        # Connect to the database
        conn = connect_mysql()
        cursor = conn.cursor()

        # Iterate through the array of course objects from the frontend
        for course in data:
            contentDuration = course.get('contentDuration')
            contentId = course.get('contentId')
            contentName = course.get('contentName')
            courseId = course.get('courseId')
            courseName = course.get('courseName')
            subModules = course.get('subModules')

            if not subModules:
                return {"error": "No subModules provided."}, 400

            # Iterate through the array of submodule objects within each course
            for subModule in subModules:
                subModuleId = subModule.get('subModuleId')
                subModuleNm = subModule.get('subModuleNm')
                contentDetails = subModule.get('contentDetails')
                moduleDuration = subModule.get('moduleDuration')

                if contentDetails is None or moduleDuration is None:
                    return {"error": "Missing data to update or insert."}, 400

                if subModuleId:
                    # Update existing submodule data in CourseContentMaster table
                    query = """
                        UPDATE courseContentMaster
                        SET contentDetails = %s, moduleDuration = %s, contentName = %s, 
                            contentDuration = %s, courseId = %s, courseName = %s
                        WHERE id = %s
                    """
                    cursor.execute(query, (contentDetails, moduleDuration, contentName, 
                                           contentDuration, courseId, courseName, subModuleId))
                    # if cursor.rowcount == 0:
                    #     return {"error": f"No record found to update for subModuleId {subModuleId}."}, 404
                else:
                    # Insert new submodule data into CourseContentMaster table
                    query = """
                        INSERT INTO courseContentMaster (contentId, contentDetails, moduleDuration, contentName, 
                                                         contentDuration, courseId, courseName, subModuleNm)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(query, (contentId, contentDetails, moduleDuration, contentName, 
                                           contentDuration, courseId, courseName, subModuleNm))

        conn.commit()
        # Fetch data to create the PDF
        
        courseNamePDF = ''
        for course in data:
            courseNamePDF = course['courseName']
            
            
        pdfSqlQuery = """
            SELECT courseId, courseName, contentId, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration 
            FROM courseContentMaster
            WHERE courseName=%s
        """

        print("------------------>",courseNamePDF)
        cursor.execute(pdfSqlQuery,courseNamePDF)
        rows = cursor.fetchall()
        # print(rows)
        

        course_data = {}
        for row in rows:
            print(row)
            courseId = row['courseId']
            if courseId not in course_data:
                course_data[courseId] = {
                    "courseName": row['courseName'],
                    "contents": []
                }
            course_data[courseId]["contents"].append({
                "contentId": row['contentId'],
                "moduleDuration": row['moduleDuration'],                
                "contentDuration": row['contentDuration'],
                "subModuleNm": row['subModuleNm'],
                "contentDetails": row['contentDetails'],
                "contentName": row['contentName']
                
            })
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        # print(course_data)
        # Create PDF
        pdf_folder_path = os.path.join(UPLOAD_COURSE_PATH, courseName)
        if not os.path.exists(pdf_folder_path):
            os.makedirs(pdf_folder_path)
        pdf_file_name = f"{courseName}_{timestamp}.pdf"
        print('pdfFilePath---', pdf_file_name)
        pdfFile = f"{pdf_folder_path}/{pdf_file_name}"
        pdfCreation(course_data, pdfFile)
        pdfNameDB =f"{courseName}/{pdf_file_name}"

        print("bfnkd-----------", courseNamePDF)
        # Retrieve existing folderName from the database
        selectSql = "SELECT folderName FROM courseMaster WHERE courseName = %s"

        cursor.execute(selectSql,courseNamePDF)
        result = cursor.fetchone()

        print("45678987654",result)
        
        if result and result['folderName'] is not None:
            existingFolderName = result['folderName']
        else:
            existingFolderName = ''

        # Initialize courseFolderName
        if existingFolderName:
            courseFolderName = f"{existingFolderName},{pdfNameDB}"
        else:
            courseFolderName = pdfNameDB

        updateSql = "UPDATE courseMaster SET folderName = %s WHERE courseName = %s"

        cursor.execute(updateSql, (courseFolderName, courseNamePDF))
    # connection.commit()


        conn.commit()
        cursor.close()
        conn.close()
        

        return jsonify({"message": "Data inserted successfully"}), 200

    # except mysql.connector.Error as err:
    #     print(f"Database error: {err}")
    #     return {"error": "Database error"}, 500
    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data processing: {e}"}, 500
    


    


@course.route('/courseXlslDocumentUpload', methods=['POST'])
@cross_origin()
def courseXlslDocumentUpload():
    global courseFolderPath

    try:
        # Extract form data
        courseName = request.form.get('courseName')
        file = request.files.get('file')
        
        if not file or not courseName:
            return {"error": "Missing file or courseName"}, 400

        contentExtension = file.filename.split('.')[-1].lower()
        if contentExtension not in ['xlsx', 'xls']:
            return {"error": "Invalid file format"}, 400

        path = os.path.join(UPLOAD_COURSE_PATH, courseName)
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

        if not os.path.exists(path):
            os.makedirs(path)

        excel_file_name = f"{courseName}_{timestamp}.{contentExtension}"
        filePath = os.path.join(path, excel_file_name)
        file.save(filePath)

        # Load the Excel file
        df = pd.read_excel(filePath)

        # Initialize the dictionary to hold all courses
        courses = {}

        # Variables to keep track of current course and module
        current_course_id = None
        current_module = None

        # Open a database connection
        conn = connect_mysql()
        cursor = conn.cursor()

        for idx, row in df.iterrows():
            course_id = row['id']
            course_name = row['Course']
            module = row['Module']
            day = row['Days']
            courseContent = row['Content']
            duration = row['Duration(hrs)']
            
            # Convert NaN values to None
            course_id = None if pd.isna(course_id) else int(course_id)
            course_name = None if pd.isna(course_name) else course_name
            module = None if pd.isna(module) else module
            day = None if pd.isna(day) else day
            courseContent = None if pd.isna(courseContent) else courseContent
            duration = None if pd.isna(duration) else duration
            
            # Check if the course ID is not missing
            if course_id is not None:
                current_course_id = course_id
                if current_course_id not in courses:
                    courses[current_course_id] = {
                        'CourseId': current_course_id,
                        'CourseName': courseName,
                        'Modules': []
                    }
            
            # Check if the module name is not missing
            if module is not None:
                current_module = {
                    'Module': module,
                    'Days': []
                }
                courses[current_course_id]['Modules'].append(current_module)
            
            # Append day, content, and duration to the current module
            if day is not None:
                current_module['Days'].append({
                    'Day': day,
                    'Content': courseContent,
                    'Duration': duration
                })

                # Insert data into SQL table
                query = """
                    INSERT INTO courseContentMaster (contentId, contentDetails, moduleDuration, contentName, 
                                                     contentDuration, courseId, courseName, subModuleNm)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    current_module['Module'], 
                    day, 
                    duration, 
                    courseContent, 
                    duration, 
                    current_course_id, 
                    courseName, 
                    courseContent
                ))

        # Commit the transaction
        conn.commit()

        # Close the database connection
        cursor.close()
        conn.close()

        # Function to create PDF
        def create_pdf(data, file_name):
            c = canvas.Canvas(file_name, pagesize=letter)
            width, height = letter
            c.setFont("Helvetica", 12)
            
            margin = 40
            y = height - margin
            
            for course_id, course_info in data.items():
                c.setFont("Helvetica-Bold", 14)
                c.drawString(margin, y, f"Course ID: {course_info['CourseId']}")
                y -= 20
                c.drawString(margin, y, f"Course Name: {course_info['CourseName']}")
                y -= 20
                c.setFont("Helvetica", 12)
                
                for module in course_info['Modules']:
                    c.setFont("Helvetica-Bold", 12)
                    c.drawString(margin + 20, y, f"Module: {module['Module']}")
                    y -= 20
                    c.setFont("Helvetica", 12)
                    for day in module['Days']:
                        c.drawString(margin + 40, y, f"Day: {day['Day']}, Content: {day['Content']}, Duration: {day['Duration']} hours")
                        y -= 20
                        
                        if y < margin:
                            c.showPage()
                            y = height - margin
                            c.setFont("Helvetica", 12)
            
            c.save()

        # Create PDF
        pdf_file_name = f"{courseName}_{timestamp}.pdf"
        pdf_file_path = os.path.join(path, pdf_file_name)
        create_pdf(courses, pdf_file_path)
        pdfFilePath = f"{courseName}/{pdf_file_name}"
        courseFolderPath = pdfFilePath

        return {"success": "File processed and PDF created successfully", "pdf_path": pdf_file_path}, 200

    except Exception as e:
        print(e)
        return {"error": "An error occurred during file upload or processing"}, 500
    

def pdfCreation(data, file_name):
    c = canvas.Canvas(file_name, pagesize=letter)
    width, height = letter
    c.setFont("Helvetica", 12)

    margin = 40
    y = height - margin

    line_height = 14

    for course_id, course_info in data.items():
        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin, y, f"Course ID: {course_id}")
        y -= line_height
        c.drawString(margin, y, f"Course Name: {course_info['courseName']}")
        y -= line_height
        c.setFont("Helvetica", 12)

        for content in course_info['contents']:
            c.drawString(margin + 20, y, f"Content ID: {content['contentId']}")
            y -= line_height
            c.drawString(margin + 20, y, f"Content Name: {content['contentName']}")
            y -= line_height
            c.drawString(margin + 20, y, f"Module Duration: {content['moduleDuration']} hours")
            y -= line_height
            
            c.drawString(margin + 20, y, f"Sub Module Name: {content['subModuleNm']}")
            y -= line_height
            
            content_details = content['contentDetails']
            wrapped_text = simpleSplit(content_details, c._fontname, c._fontsize, width - 2 * margin)
            
            for line in wrapped_text:
                if y < margin:
                    c.showPage()
                    y = height - margin
                    c.setFont("Helvetica", 12)
                
                c.drawString(margin + 20, y, line)
                y -= line_height

            c.drawString(margin + 20, y, f"Content Duration: {content['contentDuration']} hours")
            y -= line_height

            if y < margin:
                c.showPage()
                y = height - margin
                c.setFont("Helvetica", 12)

    c.save()




@course.route('/editMultiCourseModuleData', methods=['POST'])
@cross_origin()
def editMultiCourseModuleData():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)
        courseName = data.get("courseName")
        # Connect to the database
        conn = connect_mysql()
        cursor = conn.cursor()

        # Iterate through the array of objects from the frontend
        for item in data:
            subModuleId = item.get('subModuleId')
            updatedContentDetails = item.get('contentDetails')
            updatedModuleDuration = item.get('moduleDuration')

            if updatedContentDetails is None or updatedModuleDuration is None:
                return {"error": "Missing data to update or insert."}, 400

            if subModuleId:
                # Update data in CourseContentMaster table
                query = """
                    UPDATE courseContentMaster
                    SET contentDetails = %s, moduleDuration = %s
                    WHERE id = %s
                """
                cursor.execute(query, (updatedContentDetails, updatedModuleDuration, subModuleId))
            else:
                # Insert new data into CourseContentMaster table
                query = """
                    INSERT INTO courseContentMaster (contentDetails, moduleDuration)
                    VALUES (%s, %s)
                """
                cursor.execute(query, (updatedContentDetails, updatedModuleDuration))

        conn.commit()

        # Fetch data to create the PDF
        pdfSqlQuery = """
            SELECT courseId, courseName, contentId, contentName, contentDuration, subModuleNm, contentDetails, moduleDuration 
            FROM courseContentMaster
            WHERE courseName=%s
        """
        cursor.execute(pdfSqlQuery,courseName)
        rows = cursor.fetchall()
        
        # Format data for the PDF
        course_data = {}
        for row in rows:
            course_id, course_name, content_id, content_name, content_duration, sub_module_nm, content_details, module_duration = row
            if course_id not in course_data:
                course_data[course_id] = {
                    "courseName": course_name,
                    "contents": []
                }
            course_data[course_id]["contents"].append({
                "contentId": content_id,
                "contentName": content_name,
                "contentDuration": content_duration,
                "subModuleNm": sub_module_nm,
                "contentDetails": content_details,
                "moduleDuration": module_duration
            })

        # Create PDF
        pdf_file_path = os.path.join(UPLOAD_SUBMODULE_CONTENT_PATH, courseName)
        pdf_file_name = f"{pdf_file_path}.{courseName}.pdf"
        print('pdfFilePath---', pdf_file_name)
        pdfCreation(course_data, pdf_file_name)

        cursor.close()
        conn.close()

        return jsonify({"message": "Records processed successfully, PDF created"}), 200
    except Exception as e:
        print(f"An error occurred: {e}")
        return {"error": f"An error occurred during data processing: {e}"}, 500
