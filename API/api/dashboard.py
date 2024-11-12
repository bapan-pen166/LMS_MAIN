from flask import Blueprint
from flask.json import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
import math, random
import os
import requests
from datetime import timedelta
from api.sendMail import *

engine = create_engine('mysql+pymysql://root:%s@localhost/lms' % quote_plus('password'))

dashboard = Blueprint('dashboard', __name__)


@dashboard.route('/getStudentOnBoardedGraph', methods=['POST'])
@cross_origin()
def getStudentOnBoardedGraph():
    try:

        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        day_range = int(req['day_range'])
        dateList = []
        today = datetime.today()
        for i in range(day_range - 1, -1, -1):
            yesterday = today - timedelta(days=i)
            dateList.append(str(yesterday.date()))
        print(dateList)
        newDf = pd.DataFrame()
        # print(newDf)
        newDf['date'] = dateList
        # print(newDf)

        sql = "SELECT date(updatedAt) as date,COUNT(*)as count FROM `converted_student_data` WHERE status = 2 AND date(updatedAt) IN " + str(
            tuple(dateList)) + " GROUP BY date(updatedAt) ORDER BY `converted_student_data`.`status`  DESC;"

        df = pd.read_sql(sql, con=engine)
        # print(df)

        df['date'] = df[['date']].astype(str)
        df1 = pd.merge(newDf, df, left_on=['date'], right_on=['date'], how='left')
        # print(df1)
        df1 = df1.fillna(0)

        dateList = df1['date'].to_list()
        countList = df1['count'].to_list()

        # cursor.execute(sql)

        # print("33333",sql)
        resp = jsonify({"status": 1, "success": True, "message": " Successfully Fetched", 'dateList': dateList,
                        "countList": countList})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        return resp


# re = getStudentOnBoardedGraph(7)

# today = datetime.today()
# print("Today is: ", today)

# # Yesterday date
# yesterday = today - timedelta(days = 1)
# print("Yesterday was: ", yesterday)


@dashboard.route('/getThisWeekOnboardedStudents', methods=['GET'])
@cross_origin()
def getThisWeekOnboardedStudents():
    try:

        conn = connect_mysql()
        cursor = conn.cursor()
        dateListThisWeek = []
        today = datetime.today()
        for i in range(6, -1, -1):
            yesterday = today - timedelta(days=i)
            dateListThisWeek.append(str(yesterday.date()))
        print(dateListThisWeek)

        prevWeek = []
        for i in range(13, 6, -1):
            # print(i)

            yesterday = today - timedelta(days=i)
            prevWeek.append(str(yesterday.date()))
        print(prevWeek)

        thisweekCount_sql = "SELECT COUNT(*)as count FROM `converted_student_data` WHERE status = 2 AND date(updatedAt) IN " + str(
            tuple(dateListThisWeek)) + "  ORDER BY `converted_student_data`.`status`  DESC;"
        cursor.execute(thisweekCount_sql)
        weekStudentsOnboarded = cursor.fetchone()['count']

        prevWeekCount_sql = "SELECT COUNT(*)as count FROM `converted_student_data` WHERE status = 2 AND date(updatedAt) IN " + str(
            tuple(prevWeek)) + "  ORDER BY `converted_student_data`.`status`  DESC;"
        cursor.execute(prevWeekCount_sql)
        prevWeekCount_res = cursor.fetchone()['count']
        # print("prevWeekCount_res---------------------",prevWeekCount_res)
        if prevWeekCount_res != 0:
            weekGrowthDegrowth = ((weekStudentsOnboarded - prevWeekCount_res) / prevWeekCount_res) * 100
        else:

            weekGrowthDegrowth = 100
        resp = jsonify({"status": 1, "success": True, "message": " Successfully Fetched",
                        'weekStudentsOnboarded': weekStudentsOnboarded,
                        "weekGrowthDegrowth": round(weekGrowthDegrowth, 2)})

        # resp =True



    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        return resp


# getThisWeekOnboardedStudents()


# getThisMonthOnboardedStudents


@dashboard.route('/getThisMonthOnboardedStudents', methods=['GET'])
@cross_origin()
def getThisMonthOnboardedStudents():
    try:

        conn = connect_mysql()
        cursor = conn.cursor()
        dateListThisWeek = []
        today = datetime.today()
        thisMonth1stDate = datetime.today().replace(day=1)
        print(thisMonth1stDate)
        mon = datetime.today().month
        if mon == 1:

            preMonth1stDate = thisMonth1stDate.replace(month=12, year=datetime.today().year - 1)
        else:
            preMonth1stDate = thisMonth1stDate.replace(month=mon - 1)
        prevLastDate = thisMonth1stDate - timedelta(days=1)
        # print("preMonth1stDate----",preMonth1stDate)
        # print("prevLastDate-----",prevLastDate)

        thisweekCount_sql = "SELECT COUNT(*) as count FROM `converted_student_data` WHERE status = 2 AND date(updatedAt) BETWEEN '" + str(
            thisMonth1stDate) + "' AND '" + str(today) + "' "

        print("thisweekCount_sql>>>", thisweekCount_sql)
        cursor.execute(thisweekCount_sql)
        monthStudentsOnboarded = cursor.fetchone()['count']

        prevWeekCount_sql = "SELECT COUNT(*) as count FROM `converted_student_data` WHERE status = 2 AND date(updatedAt) BETWEEN '" + str(
            preMonth1stDate) + "' AND '" + str(prevLastDate) + "' "
        print("prevWeekCount_sql>>>", prevWeekCount_sql)

        cursor.execute(prevWeekCount_sql)
        prevMonthCount_res = cursor.fetchone()['count']
        if prevMonthCount_res != 0:
            monthGrowthDegrowth = ((monthStudentsOnboarded - prevMonthCount_res) / prevMonthCount_res) * 100
        else:
            monthGrowthDegrowth = 100

        resp = jsonify({"status": 1, "success": True, "message": " Successfully Fetched",
                        'monthGrowthDegrowth': round(monthGrowthDegrowth, 2),
                        "monthStudentsOnboarded": monthStudentsOnboarded})

        # resp =True



    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        return resp


# getThisMonthOnboardedStudents()


# @dashboard.route('/getTopCourse', methods=['GET'])
# @cross_origin()
# def getTopCourse():
#     try:

#         conn = connect_mysql()
#         cursor = conn.cursor()

#         sql = " SELECT mx_Course as Course,COUNT(*) as studentCount FROM `converted_student_data` where mx_Course is not Null GROUP BY mx_Course ORDER BY studentCount DESC limit 3"
#         cursor.execute(sql)
#         res = cursor.fetchall()

#         resp = jsonify({"status": 1, "success": True, "message": " Successfully Fetched", "topCourse": res})





#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0})
#     finally:
#         cursor.close()
#         return resp



#RISHABH MUDGAL
@dashboard.route('/getTopCourse', methods=['GET'])
@cross_origin()
def getTopCourse():
    try:

        conn = connect_mysql()
        cursor = conn.cursor()

        sql = "SELECT mx_Course AS Course, COUNT(*) AS studentCount FROM converted_student_data WHERE mx_Course IS NOT NULL GROUP BY mx_Course ORDER BY studentCount DESC LIMIT 3"
        cursor.execute(sql)
        res = cursor.fetchall()

        resp = jsonify({"status": 1, "success": True, "message": " Successfully Fetched", "topCourse": res})
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        return resp


@dashboard.route('/getTotalOverallStudentsCount', methods=['GET'])
@cross_origin()
def getTotalOverallStudentsCount():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT COUNT(*) as Count FROM converted_student_data""")
            # """SELECT COUNT(*) as Count FROM converted_student_data WHERE EmailAddress NOT IN (SELECT conv.EmailAddress FROM converted_student_data AS conv JOIN placementStudentMaster AS psm ON conv.EmailAddress = psm.email)""")
        count = cursor.fetchall()
        return jsonify({"status": 1, "success": True, "count": count[0]['Count']})
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
        return resp


# @dashboard.route('/getCourseWiseStudents', methods=['GET'])
# @cross_origin()
# def getCourseWiseStudents():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         cursor.execute(
#             """SELECT mx_Course, Count(*) as Count FROM converted_student_data WHERE mx_Course IS NOT NULL GROUP BY mx_Course ORDER BY Count DESC""")
#         count = cursor.fetchall()
#         return jsonify({"status": 1, "success": True, "courseWiseStudents": count})
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0})
#         return resp

# @dashboard.route('/getCourseWiseStudents', methods=['POST'])
# @cross_origin()
# def getCourseWiseStudents():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         startDate = data.get('startDate', '').strip()
#         endDate = data.get('endDate', '').strip()

#         if not (startDate and endDate):
#             return jsonify({"status": 0, "success": False, "message": "Both startDate and endDate are required."}), 400

#         cursor.execute("SELECT DISTINCT mx_Course FROM converted_student_data WHERE mx_Course IS NOT NULL")
#         all_courses = cursor.fetchall()

#         course_counts = {course['mx_Course']: 0 for course in all_courses}

#         cursor.execute(
#             """
#             SELECT mx_Course, COUNT(*) as Count 
#             FROM converted_student_data 
#             WHERE mx_Course IS NOT NULL AND DATE(CreatedOn) BETWEEN %s AND %s 
#             GROUP BY mx_Course 
#             ORDER BY Count DESC
#             """, (startDate, endDate)
#         )
#         results = cursor.fetchall()

#         for result in results:
#             course_counts[result['mx_Course']] = result['Count']

#         course_counts_list = [{"mx_Course": course, "Count": count} for course, count in course_counts.items()]

#         return jsonify({"status": 1, "success": True, "courseWiseStudents": course_counts_list})

#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0, "success": False, "message": "Error fetching course-wise students."})
#         return resp

#     finally:
#         cursor.close()
#         conn.close()

@dashboard.route('/getCourseWiseStudents', methods=['POST'])
@cross_origin()
def getCourseWiseStudents():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        startDate = data.get('startDate', '').strip()
        endDate = data.get('endDate', '').strip()

        if startDate and endDate:
            cursor.execute(
                """
                SELECT DISTINCT mx_Course FROM converted_student_data WHERE mx_Course IS NOT NULL
                """
            )
            all_courses = cursor.fetchall()

            course_counts = {course['mx_Course']: 0 for course in all_courses}

            cursor.execute(
                """
                SELECT mx_Course, COUNT(*) as Count 
                FROM converted_student_data 
                WHERE mx_Course IS NOT NULL AND DATE(CreatedOn) BETWEEN %s AND %s 
                GROUP BY mx_Course 
                ORDER BY Count DESC
                """, (startDate, endDate)
            )
            results = cursor.fetchall()

            for result in results:
                course_counts[result['mx_Course']] = result['Count']

            course_counts_list = [{"mx_Course": course, "Count": count} for course, count in course_counts.items()]

            return jsonify({"status": 1, "success": True, "courseWiseStudents": course_counts_list})

        else:
            cursor.execute(
                """
                SELECT mx_Course, COUNT(*) as Count 
                FROM converted_student_data 
                WHERE mx_Course IS NOT NULL 
                GROUP BY mx_Course 
                ORDER BY Count DESC
                """
            )
            count = cursor.fetchall()
            return jsonify({"status": 1, "success": True, "courseWiseStudents": count})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "success": False, "message": "Error fetching course-wise students."})
        return resp

    finally:
        cursor.close()
        conn.close()



# @dashboard.route('/getBatchWiseStudents', methods=['GET'])
# @cross_origin()
# def getBatchWiseStudents():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         cursor.execute(
#             """SELECT batch, Count(*) as Count FROM converted_student_data WHERE batch IS NOT NULL GROUP BY batch ORDER BY Count DESC""")
#         count = cursor.fetchall()
#         return jsonify({"status": 1, "success": True, "courseWiseStudents": count})
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0})
#         return resp

@dashboard.route('/getBatchWiseStudents', methods=['POST'])
@cross_origin()
def getBatchWiseStudents():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        
        data = request.json
        startDate = data.get('startDate', '').strip()
        endDate = data.get('endDate', '').strip()

        
        cursor.execute(
            """
            SELECT DISTINCT batch 
            FROM converted_student_data 
            WHERE batch IS NOT NULL
            """
        )
        all_batches = cursor.fetchall()

        
        batch_counts = {batch['batch']: 0 for batch in all_batches}

        if startDate and endDate:
            
            cursor.execute(
                """
                SELECT batch, COUNT(*) as Count 
                FROM converted_student_data 
                WHERE batch IS NOT NULL AND DATE(CreatedOn) BETWEEN %s AND %s 
                GROUP BY batch 
                ORDER BY Count DESC
                """, (startDate, endDate)
            )
            results = cursor.fetchall()

            
            for result in results:
                batch_counts[result['batch']] = result['Count']

        else:
            
            cursor.execute(
                """
                SELECT batch, COUNT(*) as Count 
                FROM converted_student_data 
                WHERE batch IS NOT NULL 
                GROUP BY batch 
                ORDER BY Count DESC
                """
            )
            results = cursor.fetchall()

            for result in results:
                batch_counts[result['batch']] = result['Count']

        batch_counts_list = [{"batch": batch, "Count": count} for batch, count in batch_counts.items()]

        return jsonify({"status": 1, "success": True, "batchWiseStudents": batch_counts_list})

    except Exception as e:
        print(f"Error: {e}")
        resp = jsonify({"status": 0, "success": False, "message": "Error fetching batch-wise students."})
        return resp

    finally:
        cursor.close()
        conn.close()




@dashboard.route('/getBatchWiseAssignmentData', methods=['POST'])
def getBatchWiseAssignmentData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()  

        data = request.json
        mentorEmail = data['mentorEmail']

        batchListSql = """
            SELECT batch FROM assignmentMaster WHERE mentorEmail=%s
        """
        cursor.execute(batchListSql, (mentorEmail,))
        batchList = cursor.fetchall()

        unique_batches = {}
        for row in batchList:
            batch_str = row['batch']
            try:
                batch_data = json.loads(batch_str)
                for batch in batch_data:
                    key = batch['batchName']
                    if key not in unique_batches:
                        unique_batches[key] = batch
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON: {e}")
                continue

        unique_batch_list = list(unique_batches.values())

        final_result = []

        for each_batch in unique_batch_list:
            studentListInBatchSql = """
                SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
            """
            cursor.execute(studentListInBatchSql, (each_batch['batchName'],))
            studentListInBatch = cursor.fetchall()

            
            each_batch['assignments'] = []

            assignmentListSql = """
                SELECT * FROM assignmentMaster 
                WHERE mentorEmail=%s AND JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
            """
            cursor.execute(assignmentListSql, (mentorEmail, each_batch['batchName']))
            assignmentList = cursor.fetchall()

            total_batch_obtained_marks = 0
            total_batch_possible_marks = 0

            for assignment in assignmentList:
                assignment['students'] = []

                total_assignment_obtained_marks = 0
                total_assignment_possible_marks = int(assignment['totalMarks']) if assignment['totalMarks'] else 0

                for student in studentListInBatch:
                    marksSql = """
                        SELECT marks FROM assignmentStudentMaster 
                        WHERE assignmentId=%s AND email=%s
                    """
                    cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
                    marks = cursor.fetchone()

                    student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
                    total_assignment_obtained_marks += student_marks

                    student_data = {
                        'studentId': student['studentId'],
                        'EmailAddress': student['EmailAddress'],
                        'marks': student_marks
                    }

                    assignment['students'].append(student_data)

                total_batch_obtained_marks += total_assignment_obtained_marks
                total_batch_possible_marks += total_assignment_possible_marks * len(studentListInBatch)

                if total_assignment_possible_marks > 0 and len(studentListInBatch) > 0:
                    assignment['averageMarksPercentage'] = (total_assignment_obtained_marks / (total_assignment_possible_marks * len(studentListInBatch))) * 100
                else:
                    assignment['averageMarksPercentage'] = 0

                each_batch['assignments'].append(assignment)
                each_batch['students'] = assignment['students']
            each_batch['totalMarksObtained'] = total_batch_obtained_marks
            each_batch['totalMarksPossible'] = total_batch_possible_marks

            if each_batch['totalMarksPossible'] > 0 and len(studentListInBatch) > 0:
                    each_batch['totalAverageBatchAssignmentPerformance'] = round((total_batch_obtained_marks / total_batch_possible_marks) * 100, 2)
            else:
                each_batch['totalAverageBatchAssignmentPerformance'] = 0
        
            final_result.append(each_batch)

        return jsonify(final_result)

    finally:
        cursor.close()
        conn.close()



 
@dashboard.route('/getBatchWiseAttendanceData', methods=['POST'])
def getBatchWiseAttendanceData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
 
        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']
 
        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
            AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        mentorEmailLike = '%"{}"%'.format(mentorEmail)
        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
        meetingIds = cursor.fetchall()
        
        meetings = []
        
        for meet in meetingIds:
            meetings.append(meet['meetingId'])
        
        
        totalStudentInBatchSql = """
         SELECT COUNT(id) as NoOfStudent FROM `converted_student_data` WHERE batch=%s;
        """
        cursor.execute(totalStudentInBatchSql, (batch,))
        totalStudentInBatch = cursor.fetchone()
        totalStudentInBatch = totalStudentInBatch['NoOfStudent'] if totalStudentInBatch else 0
        
        
        totalStudentGivenFeedback = 0
        totalFeedback = 0
        attendeeCount = 0
        
        for row in meetings:
            sql = """
                SELECT COUNT(id) as totalAttendee,
                       SUM(IFNULL(Feedback, 0)) as feedback,
                       COUNT(Feedback) as feedbackGivenBy
                FROM attendanceFeedback
                WHERE meetingID=%s
            """
            cursor.execute(sql, (row,))
            result = cursor.fetchone()
 
            if result:
                attendeeCount += result['totalAttendee'] or 0
                totalFeedback += result['feedback'] or 0
                totalStudentGivenFeedback += result['feedbackGivenBy'] or 0
        print('totalFeedback',totalFeedback,'feedbackGivenBy',totalStudentGivenFeedback)
        possibleAttendanceCount = totalStudentInBatch * len(meetings)
 
        averageAttendance = (attendeeCount / possibleAttendanceCount) * 100 if possibleAttendanceCount > 0 else 0
        
        averageFeedback = totalFeedback / totalStudentGivenFeedback if totalStudentGivenFeedback > 0 else 0
        
        return jsonify({
            "success": True,
            'batchAverageAttendance': round(averageAttendance, 2),
            'batchAverageFeedback': round(averageFeedback, 2),
            'message': "attendance and rating fetched!",
            'meetings':meetings,
            'possibleAttendanceCount':possibleAttendanceCount
        })
    
 
    except Exception as e:
        return jsonify({'error': str(e),
                        "success":False,
                        'message': "Some error occurs!"})
    finally:
        cursor.close()
        conn.close()

@dashboard.route('/getBatchWiseTestData', methods=['POST'])
def getBatchWiseTestData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']

        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()
        TestListSql = """
            SELECT * FROM tests 
            WHERE mentorEmail=%s AND JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (mentorEmail, batch))
        testList = cursor.fetchall()
        print("studentListInBatch=====", testList)



        studentEmails = [student['EmailAddress'] for student in studentListInBatch]
        print("studentEmails=====", studentEmails)


        studentMarksMap = {}

        if studentEmails:
            marksSql = """
                SELECT testId, email, marksObtained FROM studentTestAnswers 
                WHERE testId IN (%s) AND email IN (%s)
            """ % (
                ",".join([str(test['id']) for test in testList]),
                ",".join(["%s"] * len(studentEmails))
            )
            cursor.execute(marksSql, studentEmails)
            marksList = cursor.fetchall()

            for marks in marksList:
                testId = marks['testId']
                email = marks['email']
                studentMarksMap[(testId, email)] = marks['marksObtained']

        total_batch_obtained_marks = 0
        total_batch_possible_marks = 0
        testAll = []

        for test in testList:
            test['students'] = []
            total_test_obtained_marks = 0
            total_test_possible_marks = int(test['totalMarks']) if test['totalMarks'] else 0

            for student in studentListInBatch:
                studentEmail = student['EmailAddress']
                student_marks = int(studentMarksMap.get((test['id'], studentEmail), 0))
                total_test_obtained_marks += student_marks

                student_data = {
                    'studentId': student['studentId'],
                    'EmailAddress': studentEmail,
                    'marks': student_marks
                }
                test['students'].append(student_data)

            total_batch_obtained_marks += total_test_obtained_marks
            total_batch_possible_marks += total_test_possible_marks * len(studentListInBatch)

            if total_test_possible_marks > 0 and len(studentListInBatch) > 0:
                test['averageMarksPercentage'] = round(
                    (total_test_obtained_marks / (total_test_possible_marks * len(studentListInBatch))) * 100, 2
                )
            else:
                test['averageMarksPercentage'] = 0

            testAll.append(test)

        totalMarksObtained = total_batch_obtained_marks
        totalMarksPossible = total_batch_possible_marks

        if totalMarksPossible > 0:
            totalAverageBatchPerformance = round((totalMarksObtained / totalMarksPossible) * 100, 2)
        else:
            totalAverageBatchPerformance = 0

        final_result = {
            'tests': testAll,
            'totalMarksObtained': totalMarksObtained,
            'totalMarksPossible': totalMarksPossible,
            'totalAverageBatchTestPerformance': totalAverageBatchPerformance
        }

        return jsonify(final_result)

    finally:
        cursor.close()
        conn.close()


@dashboard.route('/getBatchOverallPerformance', methods=['POST'])
def getBatchOverallPerformance():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']

        # Step 1: Get batch-wise assignment performance
        # ============================================
        assignmentListSql = """
            SELECT * FROM assignmentMaster 
            WHERE mentorEmail=%s AND JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(assignmentListSql, (mentorEmail, batch))
        assignmentList = cursor.fetchall()

        total_assignment_obtained_marks = 0
        total_assignment_possible_marks = 0

        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()

        for assignment in assignmentList:
            total_assignment_possible_marks += int(assignment['totalMarks']) * len(studentListInBatch) if assignment['totalMarks'] else 0

            for student in studentListInBatch:
                marksSql = """
                    SELECT marks FROM assignmentStudentMaster 
                    WHERE assignmentId=%s AND email=%s
                """
                cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
                marks = cursor.fetchone()

                student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
                total_assignment_obtained_marks += student_marks

        if total_assignment_possible_marks > 0:
            totalAverageBatchAssignmentPerformance = round(
                (total_assignment_obtained_marks / total_assignment_possible_marks) * 100, 2
            )
        else:
            totalAverageBatchAssignmentPerformance = 0

        # Step 2: Get batch-wise test performance
        # =======================================
        testListSql = """
            SELECT * FROM tests 
            WHERE mentorEmail=%s AND JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(testListSql, (mentorEmail, batch))
        testList = cursor.fetchall()

        total_test_obtained_marks = 0
        total_test_possible_marks = 0

        studentMarksMap = {}
        if studentListInBatch:
            studentEmails = [student['EmailAddress'] for student in studentListInBatch]
            marksSql = """
                SELECT testId, email, marksObtained FROM studentTestAnswers 
                WHERE testId IN (%s) AND email IN (%s)
            """ % (
                ",".join([str(test['id']) for test in testList]),
                ",".join(["%s"] * len(studentEmails))
            )
            cursor.execute(marksSql, studentEmails)
            marksList = cursor.fetchall()

            for marks in marksList:
                testId = marks['testId']
                email = marks['email']
                studentMarksMap[(testId, email)] = marks['marksObtained']

        for test in testList:
            total_test_possible_marks += int(test['totalMarks']) * len(studentListInBatch) if test['totalMarks'] else 0

            for student in studentListInBatch:
                studentEmail = student['EmailAddress']
                student_marks = int(studentMarksMap.get((test['id'], studentEmail), 0))
                total_test_obtained_marks += student_marks

        if total_test_possible_marks > 0:
            totalAverageBatchTestPerformance = round(
                (total_test_obtained_marks / total_test_possible_marks) * 100, 2
            )
        else:
            totalAverageBatchTestPerformance = 0

        # Step 3: Get batch-wise attendance and feedback performance
        # =========================================================
        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
            AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        mentorEmailLike = '%"{}"%'.format(mentorEmail)
        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
        meetingIds = cursor.fetchall()

        meetings = [meet['meetingId'] for meet in meetingIds]
        
        totalStudentInBatchSql = """
            SELECT COUNT(id) as NoOfStudent FROM `converted_student_data` WHERE batch=%s;
        """
        cursor.execute(totalStudentInBatchSql, (batch,))
        totalStudentInBatch = cursor.fetchone()
        totalStudentInBatch = totalStudentInBatch['NoOfStudent'] if totalStudentInBatch else 0
        
        totalStudentGivenFeedback = 0
        totalFeedback = 0
        attendeeCount = 0

        for meetingId in meetings:
            attendanceFeedbackSql = """
                SELECT COUNT(id) as totalAttendee,
                       SUM(IFNULL(Feedback, 0)) as feedback,
                       COUNT(Feedback) as feedbackGivenBy
                FROM attendanceFeedback
                WHERE meetingID=%s
            """
            cursor.execute(attendanceFeedbackSql, (meetingId,))
            result = cursor.fetchone()

            if result:
                attendeeCount += result['totalAttendee'] or 0
                totalFeedback += result['feedback'] or 0
                totalStudentGivenFeedback += result['feedbackGivenBy'] or 0

        possibleAttendanceCount = totalStudentInBatch * len(meetings)

        averageAttendance = (attendeeCount / possibleAttendanceCount) * 100 if possibleAttendanceCount > 0 else 0
        averageFeedback = totalFeedback / totalStudentGivenFeedback if totalStudentGivenFeedback > 0 else 0

        # Step 4: Calculate overall performance (assignment, test, attendance)
        # ====================================================================
        overallPerformance = round(
            (totalAverageBatchAssignmentPerformance + totalAverageBatchTestPerformance + averageAttendance) / 3, 2
        )

        final_result = {
            'batch': batch,
            'totalAverageBatchAssignmentPerformance': totalAverageBatchAssignmentPerformance,
            'totalAverageBatchTestPerformance': totalAverageBatchTestPerformance,
            'batchAverageAttendance': round(averageAttendance, 2),
            'batchAverageFeedback': round(averageFeedback, 2),
            'overallPerformance': overallPerformance
        }

        return jsonify(final_result)

    except Exception as e:
        return jsonify({'error': str(e), 'success': False, 'message': 'Error fetching overall batch performance!'})

    finally:
        cursor.close()
        conn.close()
 

@dashboard.route('/getStudentsRatingPerformances', methods=['POST'])
def getStudentsRatingPerformances():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']

        
        studentAssignmentMarks = getStudentAssignmentMarks(mentorEmail, batch)
        studentTestMarks = getStudentTestMarks(mentorEmail, batch)
        studentAttendanceAvg, _ = getStudentAttendanceAverage(mentorEmail, batch)

    
        combined_data = {}
        for student in studentAssignmentMarks:
            email = student['EmailAddress']
            combined_data[email] = {
                'studentId': student['studentId'],
                'userName': student['userName'] if 'userName' in student else "", 
                'averageAssignmentMarks': student['averageAssignmentMarks'],
                'averageTestMarks': 0, 
                'attendanceAverage': 0 
            }

        for student in studentTestMarks:
            email = student['EmailAddress']
            if email in combined_data:
                combined_data[email]['averageTestMarks'] = student['averageTestMarks']

        for email, attendance_info in studentAttendanceAvg.items():
            if email in combined_data:
                combined_data[email]['attendanceAverage'] = attendance_info['attendancePercentage']
                if not combined_data[email]['userName']:
                    combined_data[email]['userName'] = attendance_info.get('firstName', "")

        students_list = []
        for email, details in combined_data.items():
            overall_score = (
                details['averageAssignmentMarks'] +
                details['averageTestMarks'] +
                details['attendanceAverage']
            ) / 3  
            details['overallScore'] = round(overall_score, 2)
            details['EmailAddress'] = email
            students_list.append(details)

        students_list.sort(key=lambda x: x['overallScore'], reverse=True)

        top_3_students = students_list[:3]
        bottom_3_students = students_list[-3:]

        return jsonify({
            'top_3_students': top_3_students,
            'bottom_3_students': bottom_3_students
        })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False, 'message': 'Error fetching batch assignment data!'})

    finally:
        cursor.close()
        conn.close()

# @dashboard.route('/getStudentBatchName', methods=['POST'])
# def getStudentBatchName():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.json
#         studentEmail = data['studentEmail']
        
#         totalStudentInBatchSql = """
#                                 SELECT batch
#                                 FROM converted_student_data
#                                 WHERE EmailAddress=%s
#                             """
                            
#         cursor.execute(totalStudentInBatchSql, studentEmail)
#         batchName = cursor.fetchone()

 
#         return jsonify(batchName)

#     except Exception as e:
#         return jsonify({'error': str(e)}), 500

#     finally:
#         cursor.close()
#         conn.close()

@dashboard.route('/getStudentBatchName', methods=['POST'])
def getStudentBatchName():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        studentEmail = data['studentEmail']

        # Query to get the batch name from converted_student_data table
        getBatchNameSql = """
            SELECT batch
            FROM converted_student_data
            WHERE EmailAddress = %s
        """
        cursor.execute(getBatchNameSql, (studentEmail,))
        batchNameResult = cursor.fetchone()
        print("---",batchNameResult)
        if not batchNameResult:
            return jsonify({'error': 'Batch not found for the given student email'}), 404
        
        batchName = batchNameResult['batch']
        print(batchName)
        # Query to get the course name from batchMaster table using the batch name
        getCourseNameSql = """
            SELECT courseType
            FROM batchMaster
            WHERE batchName = %s
        """
        cursor.execute(getCourseNameSql, (batchName,))
        courseNameResult = cursor.fetchone()
        print("course",courseNameResult)
        if not courseNameResult:
            return jsonify({'error': 'Course not found for the given batch name'}), 404

        courseName = courseNameResult['courseType']

        # Returning both batch name and course name
        return jsonify({'batchName': batchName, 'courseName': courseName})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()




@dashboard.route('/getStudentTodayClass', methods=['POST'])
def getStudentTodayClass():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        batchName = data['batchName']
        data = []

        today_date = datetime.now().date()

        meetingIdSql = """
            SELECT meetingId, meetingLink, startTime, endTime, topic
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
            AND DATE(startDate) = %s
        """

        batchLike = '%"{}"%'.format(batchName)

        cursor.execute(meetingIdSql, (batchLike, today_date))
        meetings = cursor.fetchall()
        
        print(meetings)

        if not meetings:
            data.append({
                'batch': batchName,
                'class': 'No class today',
                'message': 'No meetings found for today'
            })
        else:
            for meeting in meetings:
                data.append({
                    'batch': batchName,
                    'meetingId': meeting['meetingId'],
                    'meetingLink': meeting['meetingLink'],
                    'startTime': meeting['startTime'],
                    'endTime': meeting['endTime'],
                    'topic': meeting['topic']
                })

        return jsonify(data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()     


@dashboard.route('/getStudentUpcomingClasses', methods=['POST'])
def getStudentUpcomingClasses():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        batchName = data['batchName']
        response_data = []

        current_time = datetime.now()

        meetingIdSql = """
            SELECT meetingId, meetingLink, startTime, endTime, topic, startDate
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
            AND startDate > %s
            ORDER BY startDate ASC
        """

        batchLike = '%"{}"%'.format(batchName)

        cursor.execute(meetingIdSql, (batchLike, current_time))
        meetings = cursor.fetchall()
        
        if not meetings:
            response_data.append({
                'batch': batchName,
                'class': 'No upcoming classes',
                'message': 'No upcoming meetings found'
            })
        else:
            for meeting in meetings:
                response_data.append({
                    'batch': batchName,
                    'meetingId': meeting['meetingId'],
                    'meetingLink': meeting['meetingLink'],
                    'startTime': meeting['startTime'],
                    'endTime': meeting['endTime'],
                    'startDate' : meeting['startDate'],
                    'topic': meeting['topic']
                })

        return jsonify(response_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()        



@dashboard.route('/getIndividualStudentAttendance', methods=['POST'])
def getIndividualStudentAttendance():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json
        
        batchName = data['batchName']
        studentEmail = data['studentEmail']
        
        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        batchLike = '%"{}"%'.format(batchName)
        cursor.execute(meetingIdSql, (batchLike,))
        meetingIds = cursor.fetchall()
        
        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given batch'}), 404
        
        meetingIds = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]

        attendanceData = {'present': 0, 'absent': 0}

        for meetingId in meetingIds:
            attendanceSql = """
                SELECT u.email
                FROM attendanceFeedback a
                JOIN user u ON u.id = SUBSTRING_INDEX(a.StudentName, '_', -1)
                WHERE a.meetingId = %s AND u.email = %s
            """
            cursor.execute(attendanceSql, (meetingId, studentEmail))
            presentStudent = cursor.fetchone()

            if presentStudent:
                attendanceData['present'] += 1
            else:
                attendanceData['absent'] += 1

        totalClasses = attendanceData['present'] + attendanceData['absent']
        attendancePercentage = (attendanceData['present'] / totalClasses) * 100 if totalClasses > 0 else 0

        attendanceResponse = {
            'email': studentEmail,
            'attendancePercentage': round(attendancePercentage, 2)
        }

        return jsonify(attendanceResponse)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()    



@dashboard.route('/getIndividualStudentTestResult', methods=['POST'])
def getIndividualStudentTestResult():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json

        studentEmail = data['studentEmail']
        batch = data['batchName']
        
        TestListSql = """
            SELECT * FROM tests 
            WHERE JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (batch,))
        testList = cursor.fetchall()

        studentMarksMap = {}

        for test in testList:
            marksSql = """
                SELECT marksObtained FROM studentTestAnswers 
                WHERE testId=%s AND email=%s
            """
            cursor.execute(marksSql, (test['id'], studentEmail))
            marks = cursor.fetchone()

            if marks:
                studentMarksMap[test['id']] = marks['marksObtained']
            else:
                studentMarksMap[test['id']] = 0

        total_test_obtained_marks = 0
        total_test_possible_marks = 0

        for test in testList:

            total_test_obtained_marks += int(studentMarksMap.get(test['id'], 0))
            total_test_possible_marks += int(test.get('totalMarks', 0)) if test.get('totalMarks') else 0

        if total_test_possible_marks > 0:
            average_marks_percentage = round((total_test_obtained_marks / total_test_possible_marks) * 100, 2)
        else:
            average_marks_percentage = 0

        final_result = {
            'studentEmail': studentEmail,
            'batch': batch,
            'averageTestMarks': average_marks_percentage,
            'total_test_possible_marks' : total_test_possible_marks,
            'total_test_obtained_marks' : total_test_obtained_marks
        }

        return final_result

    finally:
        cursor.close()
        conn.close()               


@dashboard.route('/getIndividualStudentAssignmentResult', methods=['POST'])
def getIndividualStudentAssignmentResult():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json
        email = data['studentEmail']
        
        batch = data['batchName']


        data = {} 
        sql = """
        SELECT am.totalMarks, am.assignmentName, am.Id, asm.marks
        FROM assignmentMaster am
        JOIN assignmentStudentMaster asm ON am.Id = asm.assignmentId
        WHERE asm.email = %s AND JSON_CONTAINS(am.batch, %s, '$')
        """
        cursor.execute(sql, (email, '{"batchName": "' + batch + '"}'))
        assignments = cursor.fetchall()
        print(assignments)
        
        sum_total_marks = 0
        sum_obtained_marks = 0

        assignment_data = []
        for assignment in assignments:
            print(assignment)
            total_marks = int(assignment['totalMarks'])  
            obtained_marks = int(assignment['marks']) if assignment['marks'] is not None else 0  
            
            sum_total_marks += total_marks
            sum_obtained_marks += obtained_marks
            
            assignment_data.append({
                "totalMarks": total_marks,
                "assignmentName": assignment['assignmentName'],
                "Id": assignment['Id'],
                "obtainedMarks": obtained_marks
            })
        
        
        assignmentResultPercentage = (sum_obtained_marks / sum_total_marks) * 100 if sum_total_marks > 0 else 0
        data = {
            'sum_total_marks' : round(sum_total_marks,2),
            'sum_obtained_marks' : round(sum_obtained_marks,2),
            'assignmentResultPercentage' : round(assignmentResultPercentage,2)
        }
        return jsonify({"assignmentResult" : data})
    
    finally:
        cursor.close()
        conn.close()


@dashboard.route('/getIndividualCourseProgressBar', methods=['POST'])
def getIndividualCourseProgressBar():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        batch = data['batchName']

        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """

        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (batchLike))
        meetingIds = cursor.fetchall()

        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given mentor and batch'}), 404

        meetings = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]
        total_meetings = len(meetings)

        if total_meetings == 0:
            return jsonify({'batch': batch, 'completion_percentage': 0.0}), 200

        attendanceSql = """
            SELECT meetingId
            FROM attendanceFeedback
            WHERE meetingId IN ({})
        """.format(','.join(['%s'] * total_meetings))

        cursor.execute(attendanceSql, tuple(meetings))
        attended_meetings = cursor.fetchall()

        attended_meeting_ids = set(meet['meetingId'] for meet in attended_meetings if meet['meetingId'] is not None)
        attended_count = len(attended_meeting_ids)

        completion_percentage = (attended_count / total_meetings) * 100

        result = {
            'batch': batch,
            'completion_percentage': round(completion_percentage, 2)
        }

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@dashboard.route('/getIndividualStudentOverallPerformanceResult', methods=['POST'])
def getIndividualStudentOverallPerformanceResult():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json
        studentEmail = data['studentEmail']
        
        batch = data['batchName']
        
        
    #Get Attendance
    #--------------------------------------------------
         
        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        batchLike = '%"{}"%'.format(batch)
        cursor.execute(meetingIdSql, (batchLike,))
        meetingIds = cursor.fetchall()
        
        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given batch'}), 404
        
        meetingIds = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]

        attendanceData = {'present': 0, 'absent': 0}

        for meetingId in meetingIds:
            attendanceSql = """
                SELECT u.email
                FROM attendanceFeedback a
                JOIN user u ON u.id = SUBSTRING_INDEX(a.StudentName, '_', -1)
                WHERE a.meetingId = %s AND u.email = %s
            """
            cursor.execute(attendanceSql, (meetingId, studentEmail))
            presentStudent = cursor.fetchone()

            if presentStudent:
                attendanceData['present'] += 1
            else:
                attendanceData['absent'] += 1

        totalClasses = attendanceData['present'] + attendanceData['absent']
        
        attendancePercentage = (attendanceData['present'] / totalClasses) * 100 if totalClasses > 0 else 0

    #Get Test Result
    #-------------------------------------------------
    
        TestListSql = """
            SELECT * FROM tests 
            WHERE JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (batch,))
        testList = cursor.fetchall()

        studentMarksMap = {}

        for test in testList:
            marksSql = """
                SELECT marksObtained FROM studentTestAnswers 
                WHERE testId=%s AND email=%s
            """
            cursor.execute(marksSql, (test['id'], studentEmail))
            marks = cursor.fetchone()

            if marks:
                studentMarksMap[test['id']] = marks['marksObtained']
            else:
                studentMarksMap[test['id']] = 0

        total_test_obtained_marks = 0
        total_test_possible_marks = 0

        for test in testList:

            total_test_obtained_marks += int(studentMarksMap.get(test['id'], 0))
            total_test_possible_marks += int(test.get('totalMarks', 0)) if test.get('totalMarks') else 0

        if total_test_possible_marks > 0:
            average_test_marks_percentage = round((total_test_obtained_marks / total_test_possible_marks) * 100, 2)
        else:
            average_test_marks_percentage = 0

    #Assignment Result
    #---------------------------------------------

        sql = """
        SELECT am.totalMarks, am.assignmentName, am.Id, asm.marks
        FROM assignmentMaster am
        JOIN assignmentStudentMaster asm ON am.Id = asm.assignmentId
        WHERE asm.email = %s AND JSON_CONTAINS(am.batch, %s, '$')
        """
        cursor.execute(sql, (studentEmail, '{"batchName": "' + batch + '"}'))
        assignments = cursor.fetchall()
        print(assignments)
        
        sum_total_marks = 0
        sum_obtained_marks = 0

        assignment_data = []
        for assignment in assignments:
            print(assignment)
            total_marks = int(assignment['totalMarks'])  
            obtained_marks = int(assignment['marks']) if assignment['marks'] is not None else 0  
            
            sum_total_marks += total_marks
            sum_obtained_marks += obtained_marks
            
            assignment_data.append({
                "totalMarks": total_marks,
                "assignmentName": assignment['assignmentName'],
                "Id": assignment['Id'],
                "obtainedMarks": obtained_marks
            })
        
        
        assignmentResultPercentage = (sum_obtained_marks / sum_total_marks) * 100 if sum_total_marks > 0 else 0
        
        
        overallPerformance = round(
            (assignmentResultPercentage + average_test_marks_percentage + attendancePercentage) / 3, 2
        )
    
        return jsonify({"overallPerformance":overallPerformance})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()      


@dashboard.route('/getIndividualStudentGrades', methods=['POST'])
def getIndividualStudentGrades():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json
        studentEmail = data['studentEmail']
        
        batch = data['batchName']
        
         #Get Attendance
    #--------------------------------------------------
         
        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        batchLike = '%"{}"%'.format(batch)
        cursor.execute(meetingIdSql, (batchLike,))
        meetingIds = cursor.fetchall()
        
        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given batch'}), 404
        
        meetingIds = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]

        attendanceData = {'present': 0, 'absent': 0}

        for meetingId in meetingIds:
            attendanceSql = """
                SELECT u.email
                FROM attendanceFeedback a
                JOIN user u ON u.id = SUBSTRING_INDEX(a.StudentName, '_', -1)
                WHERE a.meetingId = %s AND u.email = %s
            """
            cursor.execute(attendanceSql, (meetingId, studentEmail))
            presentStudent = cursor.fetchone()

            if presentStudent:
                attendanceData['present'] += 1
            else:
                attendanceData['absent'] += 1

        totalClasses = attendanceData['present'] + attendanceData['absent']
        
        attendancePercentage = (attendanceData['present'] / totalClasses) * 100 if totalClasses > 0 else 0
  

    #Get Test Result
    #-------------------------------------------------
    
        TestListSql = """
            SELECT * FROM tests 
            WHERE JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (batch,))
        testList = cursor.fetchall()

        studentMarksMap = {}

        for test in testList:
            marksSql = """
                SELECT marksObtained FROM studentTestAnswers 
                WHERE testId=%s AND email=%s
            """
            cursor.execute(marksSql, (test['id'], studentEmail))
            marks = cursor.fetchone()

            if marks:
                studentMarksMap[test['id']] = marks['marksObtained']
            else:
                studentMarksMap[test['id']] = 0

        total_test_obtained_marks = 0
        total_test_possible_marks = 0

        for test in testList:

            total_test_obtained_marks += int(studentMarksMap.get(test['id'], 0))
            total_test_possible_marks += int(test.get('totalMarks', 0)) if test.get('totalMarks') else 0

        if total_test_possible_marks > 0:
            average_test_marks_percentage = round((total_test_obtained_marks / total_test_possible_marks) * 100, 2)
        else:
            average_test_marks_percentage = 0

    #Assignment Result
    #---------------------------------------------

        sql = """
        SELECT am.totalMarks, am.assignmentName, am.Id, asm.marks
        FROM assignmentMaster am
        JOIN assignmentStudentMaster asm ON am.Id = asm.assignmentId
        WHERE asm.email = %s AND JSON_CONTAINS(am.batch, %s, '$')
        """
        cursor.execute(sql, (studentEmail, '{"batchName": "' + batch + '"}'))
        assignments = cursor.fetchall()
        print(assignments)
        
        sum_total_marks = 0
        sum_obtained_marks = 0

        assignment_data = []
        for assignment in assignments:
            print(assignment)
            total_marks = int(assignment['totalMarks'])  
            obtained_marks = int(assignment['marks']) if assignment['marks'] is not None else 0  
            
            sum_total_marks += total_marks
            sum_obtained_marks += obtained_marks
            
            assignment_data.append({
                "totalMarks": total_marks,
                "assignmentName": assignment['assignmentName'],
                "Id": assignment['Id'],
                "obtainedMarks": obtained_marks
            })
        
        
        assignmentResultPercentage = (sum_obtained_marks / sum_total_marks) * 100 if sum_total_marks > 0 else 0
        
        
        overallPerformance = round(
            (assignmentResultPercentage + average_test_marks_percentage + attendancePercentage) / 3, 2
        )

        
        grade = ''
      
        if overallPerformance <= 35:
            grade =  "E"
        elif overallPerformance <= 45:
            grade = "D"
        elif overallPerformance <= 55:
            grade = "C"
        elif overallPerformance <= 65:
            grade = "B"
        elif overallPerformance < 80:
            grade = "A"
        else:
            grade = "A+"

        certificateCheck = """
        SELECT activeFlag
        FROM certificateMaster
        WHERE studentEmail=%s AND batch=%s
        """
        cursor.execute(certificateCheck, (studentEmail,batch))
        certificateCheck = cursor.fetchone()
        
        qualifyForPlacement = 1
        if certificateCheck:
            
            if certificateCheck['activeFlag'] == '0' and grade == 'E':
                qualifyForPlacement = 0
            else:
                qualifyForPlacement = 1

        return jsonify({"finalGrade":grade, 'qualifyForPlacement': qualifyForPlacement})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

 

@dashboard.route('/getIndividualStudentFiftyPercentAttendanceTarget', methods=['POST'])
def getIndividualStudentFiftyPercentAttendanceTarget():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        data = request.json
        studentEmail = data['studentEmail']
        batch = data['batchName']
                
        meetingIdSql = """
            SELECT meetingId, startDate
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        batchLike = '%"{}"%'.format(batch)
        cursor.execute(meetingIdSql, (batchLike,))
        meetings = cursor.fetchall()
        
        if not meetings:
            return jsonify({'error': 'No meetings found for the given batch'}), 404
        
        totalClasses = len(meetings)
        remainingClasses  = 0
        attendanceData = {'present': 0, 'absent': 0}

        for meeting in meetings:
            meetingId = meeting['meetingId']
            startDate_str = meeting['startDate']  
            startDate = datetime.strptime(startDate_str, '%Y-%m-%d')  

            if startDate < datetime.now():
                attendanceSql = """
                    SELECT u.email
                    FROM attendanceFeedback a
                    JOIN user u ON u.id = SUBSTRING_INDEX(a.StudentName, '_', -1)
                    WHERE a.meetingId = %s AND u.email = %s
                """
                cursor.execute(attendanceSql, (meetingId, studentEmail))
                presentStudent = cursor.fetchone()
                print(presentStudent)
                
                if presentStudent:
                    attendanceData['present'] += 1
                else:
                    attendanceData['absent'] += 1
            
            count = 0 
            if startDate > datetime.now():
                count += 1
                remainingClasses +=  count
                

        attendancePercentage = (attendanceData['present'] / (attendanceData['present'] + attendanceData['absent'])) * 100 if (attendanceData['present'] + attendanceData['absent']) > 0 else 0


        targetAttendance = 0.5 * totalClasses
        
        if attendanceData['present'] < targetAttendance:
            min_class_to_attend_50 = targetAttendance - attendanceData['present']
            
        else:
            min_class_to_attend_50 = 0
            
        possiblity = False   
        if min_class_to_attend_50 < remainingClasses:
            possiblity = True
        else:
            possiblity = False


        response = {
            'email': studentEmail,
            'currentAttendancePercentage': round(attendancePercentage,2),
            'attendedClasses' : attendanceData['present'],
            'totalClasses': totalClasses,
            'remainingClasses' : remainingClasses,
            'minClassToReach50': round(min_class_to_attend_50),
            'possiblity' : possiblity
        }

        return jsonify(response)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()



@dashboard.route('/getIndividualStudentAllAssignmentTracker', methods=['POST'])
def getIndividualStudentAllAssignmentTracker():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        studentEmail = data['studentEmail']
        batch_name = data['batchName']

        assignmentStatusList = []

        assignmentListSql = """
            SELECT id, assignmentName 
            FROM assignmentMaster 
            WHERE JSON_CONTAINS(CAST(batch AS JSON), JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(assignmentListSql, (batch_name,))
        assignmentList = cursor.fetchall()

        if not assignmentList:
            return jsonify({
                "success": False,
                'message': "No assignments found for the provided batch",
                'assignments': []
            })

        for assignment in assignmentList:
            assignmentStatus = 0  
            completeAssignmentTrackerSql = """
                SELECT COUNT(ID) as comCount FROM assignmentStudentMaster 
                WHERE email=%s AND batch=%s AND assignmentId=%s AND activeFlag=1
            """
            cursor.execute(completeAssignmentTrackerSql, (studentEmail, batch_name, assignment['id']))
            completedTotalCount = cursor.fetchone()

            inProgressAssignmentTrackerSql = """
                SELECT COUNT(ID) as inProgressCount FROM assignmentStudentMaster 
                WHERE email=%s AND batch=%s AND assignmentId=%s AND activeFlag=2
            """
            cursor.execute(inProgressAssignmentTrackerSql, (studentEmail, batch_name, assignment['id']))
            inProgressTotalCount = cursor.fetchone()

            if completedTotalCount and completedTotalCount['comCount'] > 0:
                assignmentStatus = 1  
            elif inProgressTotalCount and inProgressTotalCount['inProgressCount'] > 0:
                assignmentStatus = 2 
                
            assignmentStatusList.append({
                assignment['assignmentName']: assignmentStatus
            })

        return jsonify({
            "success": True,
            "assignments": assignmentStatusList
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        })













def getStudentAssignmentMarks(mentorEmail, batch):
        conn = connect_mysql()
        cursor = conn.cursor()
        
        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()

        assignmentListSql = """
            SELECT * FROM assignmentMaster 
            WHERE mentorEmail=%s AND JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(assignmentListSql, (mentorEmail, batch))
        assignmentList = cursor.fetchall()

        final_result = {
            'batch': batch,
            'assignments': [],
            'students': [],
            'totalMarksObtained': 0,
            'totalMarksPossible': 0,
            'totalAverageBatchAssignmentPerformance': 0
        }

        total_batch_obtained_marks = 0
        total_batch_possible_marks = 0

        student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'assignment_count': 0} for student in studentListInBatch}

        for assignment in assignmentList:
            assignment['students'] = []

            total_assignment_obtained_marks = 0
            total_assignment_possible_marks = int(assignment['totalMarks']) if assignment['totalMarks'] else 0

            for student in studentListInBatch:
                marksSql = """
                    SELECT marks FROM assignmentStudentMaster 
                    WHERE assignmentId=%s AND email=%s
                """
                cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
                marks = cursor.fetchone()

                student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
                total_assignment_obtained_marks += student_marks

                if student['EmailAddress'] in student_marks_data:
                    student_marks_data[student['EmailAddress']]['total_marks'] += student_marks
                    student_marks_data[student['EmailAddress']]['assignment_count'] += 1

                student_data = {
                    'studentId': student['studentId'],
                    'EmailAddress': student['EmailAddress'],
                    'marks': student_marks
                }

                assignment['students'].append(student_data)

            total_batch_obtained_marks += total_assignment_obtained_marks
            total_batch_possible_marks += total_assignment_possible_marks * len(studentListInBatch)

            if total_assignment_possible_marks > 0 and len(studentListInBatch) > 0:
                assignment['averageMarksPercentage'] = (total_assignment_obtained_marks / (total_assignment_possible_marks * len(studentListInBatch))) * 100
            else:
                assignment['averageMarksPercentage'] = 0

            final_result['assignments'].append(assignment)

        final_result['totalMarksObtained'] = total_batch_obtained_marks
        final_result['totalMarksPossible'] = total_batch_possible_marks

        if final_result['totalMarksPossible'] > 0:
            final_result['totalAverageBatchAssignmentPerformance'] = round((total_batch_obtained_marks / total_batch_possible_marks) * 100, 2)
        else:
            final_result['totalAverageBatchAssignmentPerformance'] = 0

        for student in studentListInBatch:
            email = student['EmailAddress']
            if email in student_marks_data:
                total_marks = student_marks_data[email]['total_marks']
                assignment_count = student_marks_data[email]['assignment_count']
                average_marks = total_marks / assignment_count if assignment_count > 0 else 0
                student_data = {
                    'studentId': student['studentId'],
                    'EmailAddress': email,
                    'averageAssignmentMarks': round(average_marks, 2)
                }
                final_result['students'].append(student_data)

        return final_result['students']


def getStudentTestMarks(mentorEmail, batch):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()

        TestListSql = """
            SELECT * FROM tests 
            WHERE mentorEmail=%s AND JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (mentorEmail, batch))
        testList = cursor.fetchall()

        studentEmails = [student['EmailAddress'] for student in studentListInBatch]
        studentMarksMap = {}

        if studentEmails:
            marksSql = """
                SELECT testId, email, marksObtained FROM studentTestAnswers 
                WHERE testId IN (%s) AND email IN (%s)
            """ % (
                ",".join([str(test['id']) for test in testList]),
                ",".join(["%s"] * len(studentEmails))
            )
            cursor.execute(marksSql, studentEmails)
            marksList = cursor.fetchall()

            for marks in marksList:
                testId = marks['testId']
                email = marks['email']
                studentMarksMap[(testId, email)] = marks['marksObtained']

        total_batch_obtained_marks = 0
        total_batch_possible_marks = 0
        testAll = []

        student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'test_count': 0} for student in studentListInBatch}
        totalMarksInAllTest = 0
        for test in testList:
            test['students'] = []
            total_test_obtained_marks = 0
            total_test_possible_marks = int(test['totalMarks']) if test['totalMarks'] else 0
            totalMarksInAllTest += int(test['totalMarks']) if test['totalMarks'] else 0

            for student in studentListInBatch:
                studentEmail = student['EmailAddress']
                student_marks = int(studentMarksMap.get((test['id'], studentEmail), 0))
                total_test_obtained_marks += student_marks
           
                if studentEmail in student_marks_data:
                    student_marks_data[studentEmail]['total_marks'] += student_marks
                    student_marks_data[studentEmail]['test_count'] += 1

                student_data = {
                    'studentId': student['studentId'],
                    'EmailAddress': studentEmail,
                    'marks': student_marks
                }
                test['students'].append(student_data)

            total_batch_obtained_marks += total_test_obtained_marks
            total_batch_possible_marks += total_test_possible_marks * len(studentListInBatch)

            if total_test_possible_marks > 0 and len(studentListInBatch) > 0:
                test['averageMarksPercentage'] = round(
                    (total_test_obtained_marks / (total_test_possible_marks * len(studentListInBatch))) * 100, 2
                )
            else:
                test['averageMarksPercentage'] = 0

            testAll.append(test)

        totalMarksObtained = total_batch_obtained_marks
        totalMarksPossible = total_batch_possible_marks

        if totalMarksPossible > 0:
            totalAverageBatchPerformance = round((totalMarksObtained / totalMarksPossible) * 100, 2)
        else:
            totalAverageBatchPerformance = 0

        studentAverageMarksList = []
        for student in studentListInBatch:
            email = student['EmailAddress']
            total_marks = student_marks_data[email]['total_marks']
            test_count = student_marks_data[email]['test_count']
            average_marks = round(total_marks / totalMarksInAllTest * 100, 2) 

            studentAverageMarksList.append({
                'studentId': student['studentId'],
                'EmailAddress': email,
                'averageTestMarks': average_marks
            })

        final_result = {
            'tests': testAll,
            'totalMarksObtained': totalMarksObtained,
            'totalMarksPossible': totalMarksPossible,
            'totalAverageBatchTestPerformance': totalAverageBatchPerformance,
            'students': studentAverageMarksList  
        }

        return final_result['students']

    finally:
        cursor.close()
        conn.close()
        
def getStudentAttendanceAverage(mentorEmail, batch):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
            AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        mentorEmailLike = '%"{}"%'.format(mentorEmail)
        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
        meetingIds = cursor.fetchall()

        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given mentor and batch'}), 404

        meetings = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]

        totalStudentInBatchSql = """
            SELECT EmailAddress, FirstName
            FROM converted_student_data
            WHERE batch = %s;
        """
        cursor.execute(totalStudentInBatchSql, (batch,))
        totalStudentData = cursor.fetchall()

        if not totalStudentData:
            return jsonify({'error': 'No students found for the given batch'}), 404

        studentInfo = {student['EmailAddress']: student['FirstName'] for student in totalStudentData}
        totalStudentInBatch = len(totalStudentData)

        attendanceData = {email: {'present': 0, 'absent': 0, 'firstName': firstName} for email, firstName in studentInfo.items()}

        for meetingId in meetings:
            attendanceSql = """
                SELECT StudentName
                FROM attendanceFeedback
                WHERE meetingId = %s
            """
            cursor.execute(attendanceSql, (meetingId,))
            presentStudentsData = cursor.fetchall()

            presentStudents = []
            for res in presentStudentsData:
                studentName = res['StudentName']
                name_parts = studentName.split('_')

                if len(name_parts) > 1:
                    userId = name_parts[-1]

                    userSql = """
                        SELECT email
                        FROM user
                        WHERE id = %s
                    """
                    cursor.execute(userSql, (userId,))
                    user_result = cursor.fetchone()

                    if user_result and user_result['email']:
                        presentStudents.append(user_result['email'])

            for email in studentInfo:
                if email in presentStudents:
                    attendanceData[email]['present'] += 1
                else:
                    attendanceData[email]['absent'] += 1

        attendanceAverage = {
            email: {
                'firstName': data['firstName'],
                'email': email,
                'attendancePercentage': (data['present'] / (data['present'] + data['absent'])) * 100 if (data['present'] + data['absent']) > 0 else 0
            }
            for email, data in attendanceData.items()
        }

        result = {
            'totalMeetings': meetings,
            'totalStudents': totalStudentInBatch,
            'studentEmails': list(studentInfo.keys()),
            'attendanceData': attendanceData,
            'attendanceAverage': attendanceAverage
        }
        print(result['attendanceAverage'])
        return result['attendanceAverage'], 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@dashboard.route('/getCertificateEligiblityData', methods=['POST'])
def getCertificateEligiblityData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']

        studentAssignmentMarks = getStudentAssignmentMarks(mentorEmail, batch)
        studentTestMarks = getStudentTestMarks(mentorEmail, batch)
        studentAttendanceAvg, _ = getStudentAttendanceAverage(mentorEmail, batch)  # Unpack if necessary
        checkResumeUpload = checkResumeUploaded(batch)

        combined_data = {}

        for student in studentAssignmentMarks:
            email = student['EmailAddress']
            combined_data[email] = {
                'assignmentMarks': student['averageAssignmentMarks'],
                'testMarks': 0,  
                'attendancePercentage': 0,  
                'resumeUploaded': False  
            }

        for student in studentTestMarks:
            email = student['EmailAddress']
            if email in combined_data:
                combined_data[email]['testMarks'] = student['averageTestMarks']

        for email, details in studentAttendanceAvg.items():  
            if email in combined_data:
                combined_data[email]['attendancePercentage'] = details['attendancePercentage']

        for email, details in checkResumeUpload.items():  
            if email in combined_data:
                combined_data[email]['resumeUploaded'] = details['resumeUploaded']

        eligible_count = 0
        not_eligible_count = 0

        for email, details in combined_data.items():
            eligible = (details['attendancePercentage'] >= 50 and
                        details['testMarks'] >= 35 and
                        details['assignmentMarks'] >= 35 and
                        details['resumeUploaded'])
            if eligible:
                eligible_count += 1
            else:
                not_eligible_count += 1

        total_students = eligible_count + not_eligible_count
        percentage_eligible = (eligible_count / total_students * 100) if total_students > 0 else 0
        percentage_not_eligible = (not_eligible_count / total_students * 100) if total_students > 0 else 0

        return jsonify({
            'eligible_count': eligible_count,
            'not_eligible_count': not_eligible_count,
            'eligible': round(percentage_eligible, 2),
            'notEligible': round(percentage_not_eligible, 2)
        })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False, 'message': 'Error fetching batch assignment data!'})

    finally:
        cursor.close()
        conn.close()

      
@dashboard.route('/checkResumeUploaded', methods=['POST'])
def checkResumeUploaded(batch):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        totalStudentInBatchSql = """
            SELECT EmailAddress, FirstName
            FROM converted_student_data
            WHERE batch = %s;
        """
        cursor.execute(totalStudentInBatchSql, (batch,))
        totalStudentData = cursor.fetchall()

        if not totalStudentData:
            return jsonify({'error': 'No students found for the given batch'}), 404

        studentInfo = {student['EmailAddress']: student['FirstName'] for student in totalStudentData}

        attendanceData = {email: {'firstName': firstName, 'resumeUploaded': False}
                          for email, firstName in studentInfo.items()}

        resumeCheckSql = """
            SELECT email, resumePath
            FROM studentRegistrationDetails
            WHERE email IN (%s);
        """ % ','.join(['%s'] * len(studentInfo))  
        
        cursor.execute(resumeCheckSql, list(studentInfo.keys()))
        resumeData = cursor.fetchall()

        for student in resumeData:
            email = student['email']
            resumePath = student['resumePath']
            if resumePath:  
                attendanceData[email]['resumeUploaded'] = True

        result = {email: data for email, data in attendanceData.items()}

        return result

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

# @dashboard.route('/getBatchAssignmentReport', methods=['POST'])
# def getBatchAssignmentReport():
#     conn = connect_mysql()
#     cursor = conn.cursor()
    
#     data = request.json
#     mentorEmail = data['mentorEmail']
#     batch = data['batch']
    
#     studentListInBatchSql = """
#         SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
#     """
#     cursor.execute(studentListInBatchSql, (batch,))
#     studentListInBatch = cursor.fetchall()

#     assignmentListSql = """
#         SELECT * FROM assignmentMaster 
#         WHERE mentorEmail=%s AND JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
#     """
#     cursor.execute(assignmentListSql, (mentorEmail, batch))
#     assignmentList = cursor.fetchall()

#     total_batch_obtained_marks = 0
#     total_batch_possible_marks = 0

#     student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'assignment_count': 0} for student in studentListInBatch}

#     for assignment in assignmentList:
#         total_assignment_obtained_marks = 0
#         total_assignment_possible_marks = int(assignment['totalMarks']) if assignment['totalMarks'] else 0

#         for student in studentListInBatch:
#             marksSql = """
#                 SELECT marks FROM assignmentStudentMaster 
#                 WHERE assignmentId=%s AND email=%s
#             """
#             cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
#             marks = cursor.fetchone()

#             student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
#             total_assignment_obtained_marks += student_marks

#             if student['EmailAddress'] in student_marks_data:
#                 student_marks_data[student['EmailAddress']]['total_marks'] += student_marks
#                 student_marks_data[student['EmailAddress']]['assignment_count'] += 1

#         total_batch_obtained_marks += total_assignment_obtained_marks
#         total_batch_possible_marks += total_assignment_possible_marks * len(studentListInBatch)

#     total_students = len(studentListInBatch)
#     pass_count = 0
#     fail_count = 0

#     for student in studentListInBatch:
#         email = student['EmailAddress']
#         if email in student_marks_data:
#             total_marks = student_marks_data[email]['total_marks']
#             assignment_count = student_marks_data[email]['assignment_count']
#             average_marks = total_marks / assignment_count if assignment_count > 0 else 0
#             average_percentage = (average_marks / (total_batch_possible_marks / total_students)) * 100 if total_batch_possible_marks > 0 else 0

#             if average_percentage >= 35:
#                 pass_count += 1
#             else:
#                 fail_count += 1

#     percentage_pass = (pass_count / total_students * 100) if total_students > 0 else 0
#     percentage_fail = (fail_count / total_students * 100) if total_students > 0 else 0

#     final_result = {
#         'batch': batch,
#         'fail_count': fail_count,
#         'pass_count': pass_count,
#         'percentage_fail': round(percentage_fail, 2),
#         'percentage_pass': round(percentage_pass, 2)
#     }

#     return jsonify(final_result)

@dashboard.route('/getBatchAssignmentReport', methods=['POST'])
def getBatchAssignmentReport():
    conn = connect_mysql()
    cursor = conn.cursor()
    
    data = request.json
    mentorEmail = data['mentorEmail']
    batch = data['batch']
    
    studentListInBatchSql = """
        SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
    """
    cursor.execute(studentListInBatchSql, (batch,))
    studentListInBatch = cursor.fetchall()

    assignmentListSql = """
        SELECT * FROM assignmentMaster 
        WHERE mentorEmail=%s AND JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
    """
    cursor.execute(assignmentListSql, (mentorEmail, batch))
    assignmentList = cursor.fetchall()
    print("assignmentList------>",assignmentList)
    total_batch_obtained_marks = 0
    total_batch_possible_marks = 0

    for assignment in assignmentList:
        total_batch_possible_marks += int(assignment['totalMarks']) if assignment['totalMarks'] else 0

    student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'assignment_count': 0} for student in studentListInBatch}

    for assignment in assignmentList:
        for student in studentListInBatch:
            marksSql = """
                SELECT marks FROM assignmentStudentMaster 
                WHERE assignmentId=%s AND email=%s
            """
            cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
            marks = cursor.fetchone()

            student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
            total_batch_obtained_marks += student_marks

            if student['EmailAddress'] in student_marks_data:
                student_marks_data[student['EmailAddress']]['total_marks'] += student_marks
                student_marks_data[student['EmailAddress']]['assignment_count'] += 1

    pass_count = 0
    fail_count = 0

    for student in studentListInBatch:
        email = student['EmailAddress']
        if email in student_marks_data:
            total_marks = student_marks_data[email]['total_marks']
            if total_batch_possible_marks > 0:
                percentage = (total_marks / total_batch_possible_marks) * 100
            else:
                percentage = 0
            print(percentage)
            if percentage >= 35:
                pass_count += 1
            else:
                fail_count += 1

    percentage_pass = (pass_count / len(studentListInBatch) * 100) if len(studentListInBatch) > 0 else 0
    percentage_fail = (fail_count / len(studentListInBatch) * 100) if len(studentListInBatch) > 0 else 0

    final_result = {
        'batch': batch,
        'fail_count': fail_count,
        'pass_count': pass_count,
        'percentage_fail': round(percentage_fail, 2),
        'percentage_pass': round(percentage_pass, 2)
    }

    return jsonify(final_result)

@dashboard.route('/getCourseCompletion', methods=['POST'])
def getCourseCompletion():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']

        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
            AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        mentorEmailLike = '%"{}"%'.format(mentorEmail)
        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
        meetingIds = cursor.fetchall()

        if not meetingIds:
            return jsonify({'error': 'No meetings found for the given mentor and batch'}), 404

        meetings = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]
        total_meetings = len(meetings)

        if total_meetings == 0:
            return jsonify({'batch': batch, 'completion_percentage': 0.0}), 200

        attendanceSql = """
            SELECT meetingId
            FROM attendanceFeedback
            WHERE meetingId IN ({})
        """.format(','.join(['%s'] * total_meetings))

        cursor.execute(attendanceSql, tuple(meetings))
        attended_meetings = cursor.fetchall()

        attended_meeting_ids = set(meet['meetingId'] for meet in attended_meetings if meet['meetingId'] is not None)
        attended_count = len(attended_meeting_ids)

        completion_percentage = (attended_count / total_meetings) * 100

        result = {
            'batch': batch,
            'completion_percentage': round(completion_percentage, 2)
        }

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@dashboard.route('/getAssignmentReport', methods=['POST'])
def getAssignmentReport():
    conn = connect_mysql()
    cursor = conn.cursor()

    data = request.json
    mentorEmail = data['mentorEmail']
  

    batchListSql = """
        SELECT DISTINCT JSON_UNQUOTE(JSON_EXTRACT(CAST(batch AS JSON), '$[0].batchName')) as batchName 
        FROM assignmentMaster 
        WHERE mentorEmail=%s
    """
    cursor.execute(batchListSql, (mentorEmail,))
    batchList = cursor.fetchall()

    print(f"Fetched batchList: {batchList}")

    if not batchList:
        return jsonify({
            "message": "No batches found for the provided mentorEmail",
            "status": "error"
        }), 404

    final_result = {}

    passing_percentage_threshold = 35

    for batch_row in batchList:
        batch = batch_row['batchName']

     
        if not batch:
            print("No batch name found, skipping...")
            continue

        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress 
            FROM converted_student_data 
            WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()

        print(f"Student list for batch {batch}: {studentListInBatch}")

        if not studentListInBatch:
            print(f"No students found in batch {batch}, skipping...")
            continue

        assignmentListSql = """
            SELECT id, assignmentName, totalMarks 
            FROM assignmentMaster 
            WHERE mentorEmail=%s AND JSON_CONTAINS(CAST(batch AS JSON), JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(assignmentListSql, (mentorEmail, batch))
        assignmentList = cursor.fetchall()

        for assignment in assignmentList:
            assignment_id = assignment['id']
            assignment_name = assignment['assignmentName']
            total_assignment_possible_marks = int(assignment['totalMarks']) if assignment['totalMarks'] else 0

            pass_count = 0
            fail_count = 0

            for student in studentListInBatch:
                marksSql = """
                    SELECT marks FROM assignmentStudentMaster 
                    WHERE assignmentId=%s AND email=%s
                """
                cursor.execute(marksSql, (assignment_id, student['EmailAddress']))
                marks = cursor.fetchone()

                student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0

                if total_assignment_possible_marks > 0:
                    student_percentage = (student_marks / total_assignment_possible_marks) * 100
                    if student_percentage >= passing_percentage_threshold:
                        pass_count += 1
                    else:
                        fail_count += 1

            if pass_count > 0 or fail_count > 0:
                final_result[assignment_name] = {
                    'pass': pass_count,
                    'fail': fail_count
                }

    return jsonify(final_result)


@dashboard.route('/getBatchWiseAttendanceReport', methods=['POST'])
def getBatchWiseAttendanceReports():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']

        mentorIdSql = """
            SELECT id
            FROM mentorMaster
            WHERE emailId = %s
        """
        cursor.execute(mentorIdSql, (mentorEmail,))
        mentorIdResult = cursor.fetchone()

        if not mentorIdResult:
            return jsonify({
                "success": False,
                'message': "Mentor not found",
                'batchAverageAttendance': 0,
                'batchAverageFeedback': 0,
                'possibleAttendanceCount': 0
            })

        mentorId = mentorIdResult['id']

        batchSql = """
            SELECT id, batchName, mentorDetails
            FROM batchMaster
        """
        cursor.execute(batchSql)
        mentor_batches = []

        for batches in cursor.fetchall():
            mentor_details = batches['mentorDetails']

            if mentor_details and batches['batchName']:
                try:
                    if isinstance(mentor_details, str):
                        mentor_details = mentor_details.replace("'", '"')  
                    mentor_details = json.loads(mentor_details)

                    for mentors in mentor_details:
                        if mentorId == mentors['id']:
                            mentor_batches.append({
                                'batchId': batches['id'],
                                'batchName': batches['batchName']
                            })
                except json.JSONDecodeError:
                    return jsonify({
                        "success": False,
                        "message": "Error parsing JSON data"
                    })

        if not mentor_batches:
            return jsonify({
                "success": False,
                'message': "No batches found for the provided mentor",
                'batchAverageAttendance': 0,
                'batchAverageFeedback': 0,
                'possibleAttendanceCount': 0
            })

        final_results = []  

        for batch in mentor_batches:
            batchId = batch['batchId']
            batchName = batch['batchName']

            studentCountSql = """
                SELECT COUNT(id) as totalStudents
                FROM converted_student_data
                WHERE batch = %s
            """
            cursor.execute(studentCountSql, (batchName,))
            totalStudentResult = cursor.fetchone()
            totalStudentInBatch = totalStudentResult['totalStudents'] if totalStudentResult else 0

            meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
            AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
            mentorEmailLike = '%"{}"%'.format(mentorEmail)
            batchLike = '%"{}"%'.format(batchName)
            # print(mentorEmailLike)
            # print(batchLike)

            cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
            
            meetings = cursor.fetchall()
            print("----------->", meetings)
            if not meetings:
                final_results.append({
                    'name': batchName,
                    'y': 0
                })
                continue
            print("------>",meetings)
            batch_attendee_count = 0
            possibleAttendanceCount = totalStudentInBatch * len(meetings)

            for meeting in meetings:
                meetingId = meeting['meetingId']

                attendanceFeedbackSql = """
                    SELECT COUNT(id) as totalAttendee
                    FROM attendanceFeedback
                    WHERE meetingID = %s
                """
                cursor.execute(attendanceFeedbackSql, (meetingId,))
                result = cursor.fetchone()
                batch_attendee_count += result['totalAttendee'] if result else 0

            averageAttendance = (batch_attendee_count / possibleAttendanceCount) * 100 if possibleAttendanceCount > 0 else 0

            final_results.append({
                'name': batchName,
                'y': round(averageAttendance, 2)
            })

        return jsonify({
            "success": True,
            "result": final_results  
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        })

    finally:
        cursor.close()
        conn.close()

@dashboard.route('/getBatchWiseAllAssignmentTracker', methods=['POST'])
def getBatchWiseAllAssignmentTracker():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        
        dataObj = {}
        
        mentorIdSql = """
            SELECT id
            FROM mentorMaster
            WHERE emailId = %s
        """
        cursor.execute(mentorIdSql, (mentorEmail,))
        mentorIdResult = cursor.fetchone()

        if not mentorIdResult:
            return jsonify({
                "success": False,
                'message': "Mentor not found",
                'batchAverageAttendance': 0,
                'batchAverageFeedback': 0,
                'possibleAttendanceCount': 0
            })

        mentorId = mentorIdResult['id']

        batchSql = """
            SELECT id, batchName, mentorDetails
            FROM batchMaster
        """
        cursor.execute(batchSql)
        batchesList = cursor.fetchall()  

        mentor_batches = []

        for batch in batchesList:
            mentor_details = batch['mentorDetails']

            if mentor_details and batch['batchName']:
                try:
                    if isinstance(mentor_details, str):
                        mentor_details = mentor_details.replace("'", '"')  
                    mentor_details = json.loads(mentor_details)
                    
                    for mentors in mentor_details:
                        if mentorId == mentors['id']:
                            batch_name = batch['batchName']
                            
                            totalStudentInBatchSql = """
                                SELECT COUNT(id) as NoOfStudent
                                FROM converted_student_data
                                WHERE batch=%s
                            """
                            cursor.execute(totalStudentInBatchSql, (batch_name,))
                            totalStudentInBatchResult = cursor.fetchone()
                            
                            totalStudentInBatch = totalStudentInBatchResult['NoOfStudent'] if totalStudentInBatchResult else 0
                            
                            mentor_batches.append({
                                'batchId': batch['id'],
                                'batchName': batch['batchName'],
                                'NoOfStudent': totalStudentInBatch
                            })
            
                except json.JSONDecodeError:
                    return jsonify({
                        "success": False,
                        "message": "Error parsing JSON data"
                    })

        if not mentor_batches:
            return jsonify({
                "success": False,
                'message': "No batches found for the provided mentor",
                'batchAverageAttendance': 0,
                'batchAverageFeedback': 0,
                'possibleAttendanceCount': 0
            })
        
        for mentor in mentor_batches:
            batch_name = mentor['batchName']
            dataObj[batch_name] = {}

        for mentor in mentor_batches:
            batch_name = mentor['batchName']

            assignmentListSql = """
                SELECT id, assignmentName 
                FROM assignmentMaster 
                WHERE mentorEmail=%s AND JSON_CONTAINS(CAST(batch AS JSON), JSON_OBJECT('batchName', %s), '$')
            """
            cursor.execute(assignmentListSql, (mentorEmail, batch_name))
            assignmentList = cursor.fetchall()
            
            for assignment in assignmentList:
                completeAssignmetTrackerSql = """
                    SELECT COUNT(ID) as comCount FROM assignmentStudentMaster WHERE batch=%s AND assignmentId=%s AND activeFlag=1
                """
                cursor.execute(completeAssignmetTrackerSql, (batch_name, assignment['id']))
                completedTotalCount = cursor.fetchone()
                
                inProgressAssignmetTrackerSql = """
                    SELECT COUNT(ID) as inProgressCount FROM assignmentStudentMaster WHERE batch=%s AND assignmentId=%s AND activeFlag=2
                """
                cursor.execute(inProgressAssignmetTrackerSql, (batch_name, assignment['id']))
                inProgressTotalCount = cursor.fetchone()
                
                studentCount = mentor['NoOfStudent']
                completedCount = completedTotalCount['comCount'] if completedTotalCount else 0
                inProgressCount = inProgressTotalCount['inProgressCount'] if inProgressTotalCount else 0
                inProgressCounts = inProgressTotalCount['inProgressCount'] if inProgressTotalCount and inProgressTotalCount['inProgressCount'] != 0 else None
                notStartedStudents = studentCount - (completedCount + inProgressCount)
                
                dataObj[batch_name][assignment['assignmentName']] = {
                    'Completed': completedCount,
                    'In progress': inProgressCounts,
                    'Not Started': notStartedStudents
                }
        
        return jsonify(dataObj)
    
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        })

@dashboard.route('/getBatchwiseAssignmentEachStudent', methods=['POST'])
@cross_origin()
def getBatchwiseAssignmentEachStudent():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
 
        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
        mentorId = cursor.fetchone()['id']
        mentor_batches = []
 
        batch_student_data = []
        
        cursor.execute("""SELECT id, batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append({
                            'batchId': batches['id'],
                            'batchName': batches['batchName']
                        })
                        
        print('=======>>>>>>>>>>>>',mentor_batches)
        for batch in mentor_batches:
            batchId = batch['batchId']
            batchName = batch['batchName']
           

            assignmentListSql = """
                SELECT COUNT(id) as NoOfAssignment, assignmentName 
                FROM assignmentMaster 
                WHERE mentorEmail=%s AND JSON_CONTAINS(CAST(batch AS JSON), JSON_OBJECT('batchName', %s), '$')
                GROUP BY assignmentName
            """
            cursor.execute(assignmentListSql, (email, batchName))
            assignmentList = cursor.fetchall()
            
            for row in assignmentList:
                NoOfAssignment = row['NoOfAssignment']
                batch_student_data.append({
                    'id' : batchId,
                    'name': batchName,
                    'y': NoOfAssignment
                })
                    
        return jsonify({'success': True, 'message' : "no. of student fetched!", 'result': batch_student_data})
        
    except Exception as e:
        print(str(e))
        return jsonify({'status': 0})


@dashboard.route('/getEachStudentAssignmentData', methods=['POST'])
def getEachStudentAssignmentData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        batch = data['batch']
        response_data = []    

        assignmentListSql = """
            SELECT id, assignmentName, endDate, totalMarks 
            FROM assignmentMaster 
            WHERE mentorEmail=%s AND JSON_CONTAINS(CAST(batch AS JSON), JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(assignmentListSql, (mentorEmail, batch))
        assignmentList = cursor.fetchall()
        
        for assignment in assignmentList:
            assignmentId = assignment['id']
            assignmentName = assignment['assignmentName']
            endDate = assignment['endDate']
            totalMarks = assignment['totalMarks']

            studentAssignmentsDataSql = """
                SELECT email, assignmentName, submittedDate, batch, marks  
                FROM assignmentStudentMaster 
                WHERE batch=%s AND assignmentId=%s AND activeFlag=1
            """
            cursor.execute(studentAssignmentsDataSql, (batch, assignmentId))
            studentAssignments = cursor.fetchall()

            for studentAssignment in studentAssignments:
                email = studentAssignment['email']
                submittedDate = studentAssignment['submittedDate']
                marks = studentAssignment['marks']

                getStudentIdSql = """
                    SELECT id, FirstName, LastName
                    FROM converted_student_data 
                    WHERE EmailAddress=%s
                """
                cursor.execute(getStudentIdSql, (email,))
                studentIdResult = cursor.fetchone()

                studentId = studentIdResult['id'] if studentIdResult else None
                lastName = studentIdResult['LastName'] if studentIdResult else None
                studentName = f"{studentIdResult['FirstName']} {lastName}"

                response_data.append({
                    'studentEmail': email,  
                    'studentId': studentId,
                    'studentName' : studentName,
                    'assignmentName': assignmentName,
                    'submittedDate': submittedDate,
                    'lastDate': endDate,
                    'totalMarks': totalMarks,
                    'marks': marks
                })

        return jsonify(response_data)

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        })

@dashboard.route('/getAllCourseCompletionReport', methods=['POST'])
def getAllCourseCompletionReport():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']

        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (mentorEmail,))
        mentorId = cursor.fetchone()['id']
        mentor_batches = []

        cursor.execute("""SELECT id, batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append({
                            'batchId': batches['id'],
                            'batchName': batches['batchName']
                        })
        
        print(mentor_batches)
        data = []

        for batch in mentor_batches:
            batchName = batch['batchName']
            print(batchName)

            meetingIdSql = """
                SELECT meetingId
                FROM meetings
                WHERE REPLACE(REPLACE(mentorList, '[', ''), ']', '') LIKE %s
                AND REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
            """

            mentorEmailLike = '%"{}"%'.format(mentorEmail)
            batchLike = '%"{}"%'.format(batchName)
            
            cursor.execute(meetingIdSql, (mentorEmailLike, batchLike))
            meetingIds = cursor.fetchall()

            if not meetingIds:
                data.append({
                    'name': batchName,
                    'y': 0.0,
                    'message': 'No meetings found'
                })
                continue

            meetings = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]
            total_meetings = len(meetings)

            if total_meetings == 0:
                data.append({
                    'name': batchName,
                    'y': 0.0,
                    'message': 'No valid meetings found'
                })
                continue

            attendanceSql = """
                SELECT meetingId
                FROM attendanceFeedback
                WHERE meetingId IN ({})
            """.format(','.join(['%s'] * total_meetings))

            cursor.execute(attendanceSql, tuple(meetings))
            attended_meetings = cursor.fetchall()

            attended_meeting_ids = set(meet['meetingId'] for meet in attended_meetings if meet['meetingId'] is not None)
            attended_count = len(attended_meeting_ids)

            completion_percentage = (attended_count / total_meetings) * 100

            data.append({
                'name': batchName,
                'y': round(completion_percentage, 2)
            })

        return jsonify(data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()



####################################################
@dashboard.route('/getTopBottomFiveStudentsAdmin', methods=['POST'])
def getTopBottomFiveStudentsAdmin():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        course_type = data.get('courseName')
        batch_name = data.get('batchName')

        combined_data = {}

        if batch_name:
            # If batchName is provided, use it directly
            batches = [batch_name]
        else:
            # If no specific batchName is provided, fetch all batches for the courseType
            searchSql = """
                SELECT DISTINCT(batchName) 
                FROM batchMaster 
                WHERE courseType = %s
            """
            cursor.execute(searchSql, (course_type,))
            batches = [batch['batchName'] for batch in cursor.fetchall()]

        for batch in batches:
            studentAssignmentMarks = getStudentAssignmentMarksAdmin(batch)
            studentTestMarks = getStudentTestMarksAdmin(batch)
            studentAttendanceAvg, _ = getStudentAttendanceAverageAdmin(batch)

            for student in studentAssignmentMarks:
                email = student['EmailAddress']

                nameSql = """
                    SELECT FirstName, LastName 
                    FROM converted_student_data 
                    WHERE EmailAddress = %s
                """
                cursor.execute(nameSql, (email,))
                nameResult = cursor.fetchone()

                if nameResult:
                    first_name = nameResult['FirstName']
                    last_name = nameResult['LastName']

                    first_name = first_name if first_name and first_name != "NA" else ''
                    last_name = last_name if last_name and last_name != "NA" else ''

                    user_name = (first_name + ' ' + last_name).strip()
                else:
                    user_name = ""  

                if email not in combined_data:
                    combined_data[email] = {
                        'studentId': student['studentId'],
                        'userName': user_name,  
                        'averageAssignmentMarks': student['averageAssignmentMarks'],
                        'averageTestMarks': 0, 
                        'attendanceAverage': 0,
                        'batchName': batch  
                    }

            for student in studentTestMarks:
                email = student['EmailAddress']
                if email in combined_data:
                    combined_data[email]['averageTestMarks'] = student['averageTestMarks']

            for email, attendance_info in studentAttendanceAvg.items():
                if email in combined_data:
                    combined_data[email]['attendanceAverage'] = attendance_info['attendancePercentage']
                    if not combined_data[email]['userName']:
                        combined_data[email]['userName'] = attendance_info.get('firstName', "")

        students_list = []
        for email, details in combined_data.items():
            overall_score = (
                details['averageAssignmentMarks'] +
                details['averageTestMarks'] +
                details['attendanceAverage']
            ) / 3  
            details['overallScore'] = round(overall_score, 2)
            details['EmailAddress'] = email
            students_list.append(details)

        students_list.sort(key=lambda x: x['overallScore'], reverse=True)

        top_5_students = []
        for i, student in enumerate(students_list[:5], 1):  
            student['serialNumber'] = i
            top_5_students.append(student)

        students_list.sort(key=lambda x: x['overallScore'])

        bottom_5_students = []
        num_students_in_bottom = len(students_list[:5])  
        for i, student in enumerate(students_list[:5], 1):  
            student['serialNumber'] = num_students_in_bottom - i + 1 
            bottom_5_students.append(student)

        return jsonify({
            'topFiveStudents': top_5_students,
            'bottomFiveStudents': bottom_5_students
        })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False, 'message': 'Error fetching student performance data!'})

    finally:
        cursor.close()
        conn.close()
 

def getStudentAssignmentMarksAdmin(batch):
    conn = connect_mysql()
    cursor = conn.cursor()
    
    studentListInBatchSql = """
        SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
    """
    cursor.execute(studentListInBatchSql, (batch,))
    studentListInBatch = cursor.fetchall()

    assignmentListSql = """
        SELECT * FROM assignmentMaster 
        WHERE JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s), '$')
    """
    cursor.execute(assignmentListSql, (batch,))
    assignmentList = cursor.fetchall()

    final_result = {
        'batch': batch,
        'assignments': [],
        'students': [],
        'totalMarksObtained': 0,
        'totalMarksPossible': 0,
        'totalAverageBatchAssignmentPerformance': 0
    }

    total_batch_obtained_marks = 0
    total_batch_possible_marks = 0

    student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'assignment_count': 0} for student in studentListInBatch}

    for assignment in assignmentList:
        assignment['students'] = []

        total_assignment_obtained_marks = 0
        total_assignment_possible_marks = int(assignment['totalMarks']) if assignment['totalMarks'] else 0

        for student in studentListInBatch:
            marksSql = """
                SELECT marks FROM assignmentStudentMaster 
                WHERE assignmentId=%s AND email=%s
            """
            cursor.execute(marksSql, (assignment['id'], student['EmailAddress']))
            marks = cursor.fetchone()

            student_marks = int(marks['marks']) if marks and marks['marks'] is not None else 0
            total_assignment_obtained_marks += student_marks

            if student['EmailAddress'] in student_marks_data:
                student_marks_data[student['EmailAddress']]['total_marks'] += student_marks
                student_marks_data[student['EmailAddress']]['assignment_count'] += 1

            student_data = {
                'studentId': student['studentId'],
                'EmailAddress': student['EmailAddress'],
                'marks': student_marks
            }

            assignment['students'].append(student_data)

        total_batch_obtained_marks += total_assignment_obtained_marks
        total_batch_possible_marks += total_assignment_possible_marks * len(studentListInBatch)

        if total_assignment_possible_marks > 0 and len(studentListInBatch) > 0:
            assignment['averageMarksPercentage'] = (total_assignment_obtained_marks / (total_assignment_possible_marks * len(studentListInBatch))) * 100
        else:
            assignment['averageMarksPercentage'] = 0

        final_result['assignments'].append(assignment)

    final_result['totalMarksObtained'] = total_batch_obtained_marks
    final_result['totalMarksPossible'] = total_batch_possible_marks

    if final_result['totalMarksPossible'] > 0:
        final_result['totalAverageBatchAssignmentPerformance'] = round((total_batch_obtained_marks / total_batch_possible_marks) * 100, 2)
    else:
        final_result['totalAverageBatchAssignmentPerformance'] = 0

    for student in studentListInBatch:
        email = student['EmailAddress']
        if email in student_marks_data:
            total_marks = student_marks_data[email]['total_marks']
            assignment_count = student_marks_data[email]['assignment_count']
            average_marks = total_marks / assignment_count if assignment_count > 0 else 0
            student_data = {
                'studentId': student['studentId'],
                'EmailAddress': email,
                'averageAssignmentMarks': round(average_marks, 2)
            }
            final_result['students'].append(student_data)

    cursor.close()
    conn.close()

    return final_result['students']


def getStudentTestMarksAdmin(batch):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch students in the batch
        studentListInBatchSql = """
            SELECT id as studentId, EmailAddress FROM converted_student_data WHERE batch=%s
        """
        cursor.execute(studentListInBatchSql, (batch,))
        studentListInBatch = cursor.fetchall()

        # Fetch tests for the batch
        TestListSql = """
            SELECT * FROM tests 
            WHERE JSON_CONTAINS(batchName, JSON_OBJECT('batchName', %s), '$')
        """
        cursor.execute(TestListSql, (batch,))
        testList = cursor.fetchall()

        studentEmails = [student['EmailAddress'] for student in studentListInBatch]

        studentMarksMap = {}

        if testList and studentEmails:
            # Only execute the marks query if testList and studentEmails are not empty
            testIds = ",".join([str(test['id']) for test in testList])
            emailPlaceholders = ",".join(["%s"] * len(studentEmails))

            marksSql = f"""
                SELECT testId, email, marksObtained 
                FROM studentTestAnswers 
                WHERE testId IN ({testIds}) 
                AND email IN ({emailPlaceholders})
            """
            cursor.execute(marksSql, studentEmails)
            marksList = cursor.fetchall()

            for marks in marksList:
                testId = marks['testId']
                email = marks['email']
                studentMarksMap[(testId, email)] = marks['marksObtained']

        total_batch_obtained_marks = 0
        total_batch_possible_marks = 0
        testAll = []

        student_marks_data = {student['EmailAddress']: {'total_marks': 0, 'test_count': 0} for student in studentListInBatch}
        totalMarksInAllTest = 0

        for test in testList:
            test['students'] = []
            total_test_obtained_marks = 0
            total_test_possible_marks = int(test['totalMarks']) if test['totalMarks'] else 0
            totalMarksInAllTest += int(test['totalMarks']) if test['totalMarks'] else 0

            for student in studentListInBatch:
                studentEmail = student['EmailAddress']
                student_marks = int(studentMarksMap.get((test['id'], studentEmail), 0))
                total_test_obtained_marks += student_marks
               
                if studentEmail in student_marks_data:
                    student_marks_data[studentEmail]['total_marks'] += student_marks
                    student_marks_data[studentEmail]['test_count'] += 1

                student_data = {
                    'studentId': student['studentId'],
                    'EmailAddress': studentEmail,
                    'marks': student_marks
                }
                test['students'].append(student_data)

            total_batch_obtained_marks += total_test_obtained_marks
            total_batch_possible_marks += total_test_possible_marks * len(studentListInBatch)

            if total_test_possible_marks > 0 and len(studentListInBatch) > 0:
                test['averageMarksPercentage'] = round(
                    (total_test_obtained_marks / (total_test_possible_marks * len(studentListInBatch))) * 100, 2
                )
            else:
                test['averageMarksPercentage'] = 0

            testAll.append(test)

        totalMarksObtained = total_batch_obtained_marks
        totalMarksPossible = total_batch_possible_marks

        if totalMarksPossible > 0:
            totalAverageBatchPerformance = round((totalMarksObtained / totalMarksPossible) * 100, 2)
        else:
            totalAverageBatchPerformance = 0

        studentAverageMarksList = []
        for student in studentListInBatch:
            email = student['EmailAddress']
            total_marks = student_marks_data[email]['total_marks']
            test_count = student_marks_data[email]['test_count']
            average_marks = round(total_marks / totalMarksInAllTest * 100, 2) if totalMarksInAllTest > 0 else 0

            studentAverageMarksList.append({
                'studentId': student['studentId'],
                'EmailAddress': email,
                'averageTestMarks': average_marks
            })

        final_result = {
            'tests': testAll,
            'totalMarksObtained': totalMarksObtained,
            'totalMarksPossible': totalMarksPossible,
            'totalAverageBatchTestPerformance': totalAverageBatchPerformance,
            'students': studentAverageMarksList  
        }

        return final_result['students']

    finally:
        cursor.close()
        conn.close()


def getStudentAttendanceAverageAdmin(batch):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        meetingIdSql = """
            SELECT meetingId
            FROM meetings
            WHERE REPLACE(REPLACE(batch, '[', ''), ']', '') LIKE %s
        """
        
        batchLike = '%"{}"%'.format(batch)
        
        cursor.execute(meetingIdSql, (batchLike,))
        meetingIds = cursor.fetchall()

        if not meetingIds:
            return {}, 404  # Returning empty data if no meetings are found

        meetings = [meet['meetingId'] for meet in meetingIds if meet['meetingId'] is not None]

        totalStudentInBatchSql = """
            SELECT EmailAddress, FirstName
            FROM converted_student_data
            WHERE batch = %s;
        """
        cursor.execute(totalStudentInBatchSql, (batch,))
        totalStudentData = cursor.fetchall()

        if not totalStudentData:
            return {}, 404  # Returning empty data if no students are found

        studentInfo = {student['EmailAddress']: student['FirstName'] for student in totalStudentData}

        attendanceData = {email: {'present': 0, 'absent': 0, 'firstName': firstName} for email, firstName in studentInfo.items()}

        for meetingId in meetings:
            attendanceSql = """
                SELECT StudentName
                FROM attendanceFeedback
                WHERE meetingId = %s
            """
            cursor.execute(attendanceSql, (meetingId,))
            presentStudentsData = cursor.fetchall()

            presentStudents = []
            for res in presentStudentsData:
                studentName = res['StudentName']
                name_parts = studentName.split('_')

                if len(name_parts) > 1:
                    userId = name_parts[-1]

                    userSql = """
                        SELECT email
                        FROM user
                        WHERE id = %s
                    """
                    cursor.execute(userSql, (userId,))
                    user_result = cursor.fetchone()

                    if user_result and user_result['email']:
                        presentStudents.append(user_result['email'])

            for email in studentInfo:
                if email in presentStudents:
                    attendanceData[email]['present'] += 1
                else:
                    attendanceData[email]['absent'] += 1

        attendanceAverage = {
            email: {
                'firstName': data['firstName'],
                'email': email,
                'attendancePercentage': round((data['present'] / (data['present'] + data['absent'])) * 100 ,2) if (data['present'] + data['absent'])> 0 else 0
            }
            for email, data in attendanceData.items()
        }

        return attendanceAverage, 200

    except Exception as e:
        return {}, 500

    finally:
        cursor.close()
        conn.close()

##############################################################################################################################

#need to add certification :::::::::

@dashboard.route('/getGraduatedStudentsCount', methods=['GET'])
@cross_origin()
def getGraduatedStudentsCount():
    conn = connect_mysql()  
    cursor = conn.cursor() 
    current_date = datetime.now().strftime('%Y-%m-%d')  
    
    try:
        cursor.execute(
            """
            SELECT COUNT(*) as Count
            FROM converted_student_data csd
            JOIN batchMaster bm ON csd.batch = bm.batchName
            WHERE bm.batchTentiveEndingDate < %s
            """, (current_date,)
        )
        
        count = cursor.fetchall()

        return jsonify({"status": 1, "success": True, "graduated_count": count[0]['Count']})
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": str(e)})
        return resp
    finally:
        cursor.close() 
        conn.close()



@dashboard.route('/getOngoingBatchStudentsCount', methods=['GET'])
@cross_origin()
def getOngoingBatchStudentsCount():
    conn = connect_mysql()  
    cursor = conn.cursor() 
    current_date = datetime.now().strftime('%Y-%m-%d')  
    
    try:
        cursor.execute(
            """
            SELECT bm.batchName, COUNT(*) as Count
            FROM converted_student_data csd
            JOIN batchMaster bm ON csd.batch = bm.batchName
            WHERE bm.startingDate <= %s AND bm.batchTentiveEndingDate >= %s
            GROUP BY bm.batchName
            """, (current_date, current_date)
        )
        
        result = cursor.fetchall()
        ongoing_data = [{"batchName": row['batchName'], "ongoing_count": row['Count']} for row in result]

        total_ongoing_count = sum(row['Count'] for row in result)

        return jsonify({"status": 1, "success": True, "ongoing_data": ongoing_data, "totalCount": total_ongoing_count})
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": str(e)})
        return resp
    finally:
        cursor.close() 
        conn.close()


@dashboard.route('/getYetToStartBatchCount', methods=['GET'])
@cross_origin()
def getYetToStartBatchCount():
    conn = connect_mysql()  
    cursor = conn.cursor() 
    current_date = datetime.now().strftime('%Y-%m-%d')  
    
    try:
        cursor.execute(
            """
            SELECT bm.batchName, 
                   IFNULL(COUNT(csd.EmailAddress), 0) as Count
            FROM batchMaster bm
            LEFT JOIN converted_student_data csd ON bm.batchName = csd.batch
            WHERE bm.startingDate > %s
            GROUP BY bm.batchName
            """, (current_date,)
        )
        
        result = cursor.fetchall()
        yet_to_start_data = [{"batchName": row['batchName'], "yet_to_start_count": row['Count']} for row in result]

        total_yet_to_start_count = sum(row['Count'] for row in result)

        return jsonify({"status": 1, "success": True, "yet_to_start_data": yet_to_start_data, "yetToStartTotalCount": total_yet_to_start_count})
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": str(e)})
        return resp
    finally:
        cursor.close() 
        conn.close()






#NOV 11


@dashboard.route('/getTopBottomFiveStudentsAmongAllBatches', methods=['POST'])
def getTopBottomFiveStudentsAmongAllBatches():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        # Fetch all batch names from batchMaster
        searchSql = "SELECT DISTINCT(batchName) FROM batchMaster"
        cursor.execute(searchSql)
        batches = [batch['batchName'] for batch in cursor.fetchall()]

        combined_data = {}

        for batch in batches:
            studentAssignmentMarks = getStudentAssignmentMarksAdmin(batch)
            studentTestMarks = getStudentTestMarksAdmin(batch)
            studentAttendanceAvg, _ = getStudentAttendanceAverageAdmin(batch)

            for student in studentAssignmentMarks:
                email = student['EmailAddress']

                nameSql = """
                    SELECT FirstName, LastName 
                    FROM converted_student_data 
                    WHERE EmailAddress = %s
                """
                cursor.execute(nameSql, (email,))
                nameResult = cursor.fetchone()

                if nameResult:
                    first_name = nameResult['FirstName']
                    last_name = nameResult['LastName']

                    first_name = first_name if first_name and first_name != "NA" else ''
                    last_name = last_name if last_name and last_name != "NA" else ''

                    user_name = (first_name + ' ' + last_name).strip()
                else:
                    user_name = ""  

                if email not in combined_data:
                    combined_data[email] = {
                        'studentId': student['studentId'],
                        'userName': user_name,  
                        'averageAssignmentMarks': student['averageAssignmentMarks'],
                        'averageTestMarks': 0, 
                        'attendanceAverage': 0,
                        'batchName': batch  
                    }

            for student in studentTestMarks:
                email = student['EmailAddress']
                if email in combined_data:
                    combined_data[email]['averageTestMarks'] = student['averageTestMarks']

            for email, attendance_info in studentAttendanceAvg.items():
                if email in combined_data:
                    combined_data[email]['attendanceAverage'] = attendance_info['attendancePercentage']
                    if not combined_data[email]['userName']:
                        combined_data[email]['userName'] = attendance_info.get('firstName', "")

        students_list = []
        for email, details in combined_data.items():
            overall_score = (
                details['averageAssignmentMarks'] +
                details['averageTestMarks'] +
                details['attendanceAverage']
            ) / 3  
            details['overallScore'] = round(overall_score, 2)
            details['EmailAddress'] = email
            students_list.append(details)

        students_list.sort(key=lambda x: x['overallScore'], reverse=True)

        # Get Top 5 students
        top_5_students = []
        for i, student in enumerate(students_list[:5], 1):  
            student['serialNumber'] = i
            top_5_students.append(student)

        # Get Bottom 5 students
        students_list.sort(key=lambda x: x['overallScore'])
        bottom_5_students = []
        num_students_in_bottom = len(students_list[:5])  
        for i, student in enumerate(students_list[:5], 1):  
            student['serialNumber'] = num_students_in_bottom - i + 1 
            bottom_5_students.append(student)

        return jsonify({
            'topFiveStudents': top_5_students,
            'bottomFiveStudents': bottom_5_students
        })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False, 'message': 'Error fetching student performance data!'})

    finally:
        cursor.close()
        conn.close()



