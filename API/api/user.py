from flask import Blueprint
from flask.json import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
# from sendmail import sendMail
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
from linkedin_api import Linkedin
from urllib.parse import urlparse
from .send_mail_outlook import send_email
import string
import random
from api.sendMail import * 

# from docxtpl import DocxTemplate
# from docxtpl import InlineImage
# from docx.shared import Mm



engine = create_engine('mysql+pymysql://root:%s@localhost/lms' % quote_plus('password'))


user = Blueprint('user', __name__)




@user.route('/',methods=['GET'])
def test():
    return jsonify({"status":"OK"})


'''---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------'''

@user.route('/login',methods=['POST'])
@cross_origin()
def login():
    try:
        req = request.json
        username = req['username']
        passKey = req['passKey']
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()


        # renewalcursor = connect_mysql_renewal().cursor()

        

        query  = "SELECT * FROM user where (userName = %s or email = %s or mobileNumber = %s) AND activeFlag = 1"

        
        data = (username,username,username)
        cursor.execute(query,data)
        user = cursor.fetchone()
        print(user)
        if user is not None :
            print(user['password'])

            # bytes = passKey.encode('utf-8')            

            # storedHash = user['password']

            # hash = storedHash.encode('utf-8')
            


            # if bcrypt.checkpw(bytes, hash):
            #     print("match")

            #     currentTime = str(datetime.now())
            #     del user['password']
            if user['password'] == passKey:

              

                conn.close()
                return jsonify({"status":1, "success":True, "userDetails":user, "message": "Logged in successfully", })
           
             
            else:
                print("does not match")
                conn.close()
                return jsonify({"status":0, "success":False,"message":"Invalid Password", "userDetails":None})
            
            
        else:
            conn.close()
            return jsonify({"status":0, "success":False, "userDetails":None, "message":"Connect with Admin"})
    except Exception as e:
        print(e)
        return jsonify({"status":0})
    


@user.route('/accessRight',methods=['POST'])
@cross_origin()
def accessRight():
    try:
        req = request.json
        userId = req['userId']
        # userType = 'Student'

        conn = connect_mysql()
        cursor = conn.cursor()

        userTypeSql = " SELECT * FROM user WHERE id = "+str(userId)
        cursor.execute(userTypeSql)
        userType = cursor.fetchone()['userType']

        accessSql = " SELECT DISTINCT(userType),menuMaster.menuId,menuMaster.menuName,menuMaster.icon,menuMaster.menuLink,accessRight.seq FROM `accessRight` LEFT JOIN menuMaster ON accessRight.menuId = menuMaster.menuId WHERE userType = '"+str(userType)+"' AND menuMaster.active =1 ORDER BY `accessRight`.`seq` ASC "
        # print(accessSql)
        cursor.execute(accessSql)
        res = cursor.fetchall()
        accesslist = [] 
        # list1 ={}
        for each  in res:
            list1= {}
            # print("each-----------------",each)
            list1['userType'] = each['userType']
            list1['menu'] = each['menuName']
            list1['menuId'] = each['menuId']
            list1['menuLink'] = each['menuLink']
            list1['icon'] = each['icon']
            list1['submenu'] = []

            subMenuSql = " SELECT accessRight.subMenuId,subMenuMaster.subMenuName,subMenuMaster.subMenuLink,subMenuMaster.subMenuIcon FROM `accessRight` LEFT JOIN subMenuMaster ON accessRight.subMenuId = subMenuMaster.subMenuId WHERE userType = '"+str(userType)+"' AND subMenuMaster.active =1  AND subMenuMaster.menuId = "+str(each['menuId'])
            # print(subMenuSql)
            cursor.execute(subMenuSql)
            subMenuRes = cursor.fetchall()
            list1['submenu'] = subMenuRes
            accesslist.append(list1)
    
        # print(accesslist)
        return jsonify({"status":1, "success":True, "menuList":accesslist, "message": "successfully", })
    except Exception as e:
        print(e)
        return jsonify({"status":0})

    finally :
        cursor.close()
        
        
@user.route('/accessRight_handset',methods=['POST'])
@cross_origin()
def accessRight_handset():
    try:
        # req = request.json
        # userId = req['userId']
        userId = 1

        conn = connect_mysql()
        cursor = conn.cursor()

        userTypeSql = " SELECT * FROM user WHERE id = "+str(userId)
        cursor.execute(userTypeSql)
        userType = cursor.fetchone()['userType']

        accessSql = " SELECT DISTINCT(userType),menuMaster.menuId,menuMaster.menuName,menuMaster.icon,menuMaster.menuLink,accessRight.seq FROM `accessRight` LEFT JOIN menuMaster ON accessRight.menuId = menuMaster.menuId WHERE userType = '"+str(userType)+"' AND menuMaster.active =1 ORDER BY `accessRight`.`seq` ASC "
        # print(accessSql)
        cursor.execute(accessSql)
        res = cursor.fetchall()
        accesslist = [] 
        # list1 ={}
        for each  in res:
            
            # print("each-----------------",each)
            

            subMenuSql = " SELECT accessRight.subMenuId,subMenuMaster.subMenuName,subMenuMaster.subMenuLink,subMenuMaster.subMenuIcon FROM `accessRight` LEFT JOIN subMenuMaster ON accessRight.subMenuId = subMenuMaster.subMenuId WHERE userType = '"+str(userType)+"' AND subMenuMaster.active =1  AND subMenuMaster.menuId = "+str(each['menuId'])
            # print(subMenuSql)
            cursor.execute(subMenuSql)
            subMenuRes = cursor.fetchall()
            # print(subMenuRes)
            if subMenuRes is not None and len(subMenuRes) != 0 :
                for eachSubMenu in subMenuRes:
                    list1= {}
                    list1['userType'] = each['userType']
                    list1['menu'] = eachSubMenu['subMenuName']
                    list1['menuId'] = eachSubMenu['subMenuId']
                    list1['menuLink'] = eachSubMenu['subMenuLink']
                    list1['icon'] = eachSubMenu['subMenuIcon']


                    accesslist.append(list1)
            else:
                # print(each)
                list1= {}
                list1['userType'] = each['userType']
                list1['menu'] = each['menuName']
                list1['menuId'] = each['menuId']
                list1['menuLink'] = each['menuLink']
                list1['icon'] = each['icon']
                accesslist.append(list1)

    
        print(accesslist)
        return jsonify({"status":1, "success":True, "menuList":accesslist, "message": "successfully", })
    except Exception as e:
        print(e)
        return jsonify({"status":0})

    finally :
        cursor.close()





@user.route('/linkedinVerify', methods=['POST'])
def linkedin_profile():
    req = request.json  # Use request.args for GET parameters
    username = req['url']   
    print(username)
 
    if not username:
        return jsonify({"error": "URL parameter is required"}), 400
 
    # Extract the username from the LinkedIn URL
    parsed_url = urlparse(username)
    path_segments = parsed_url.path.split('/')
    if len(path_segments) < 3 or path_segments[1] != 'in':
        return jsonify({"error": "Invalid LinkedIn profile URL", "status":0, "success":False}), 400
    profile_username = path_segments[2]
    print(profile_username)
    # Initialize LinkedIn API client securely
    linkedin_username = 'Rishu.Raj@pentationanalytics.com'
    linkedin_password = 'rbo661'
    api = Linkedin(linkedin_username, linkedin_password)
 
    try:
        # Attempt to fetch the profile details
        profile_details = api.get_profile(profile_username)
        if profile_details:
            print("Profile Details:", profile_details)  # Debug output
            return jsonify({"status":1, "success":True, "response": profile_details}), 200
        else:
            return jsonify({"error": "Profile not found", "status":0, "success":False}), 404
    except Exception as e:
        # Log the exception details
        print(f"Exception occurred: {e}")
        return jsonify({"error": "An error occurred while fetching profile details",  "status":0, "success":False}), 500
 
        
# accessRight_handset()
# ()


#rishu Aug23,2024

@user.route('/addUser', methods=['POST'])
@cross_origin()
def addUser(res, role):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
 
        
        if role == 'Student':
            print('Student')
            firstName = res['FirstName'].strip()
            middleName = ""  
            lastName = res['LastName'].strip()
            role = role
            userId = "0"
            email = res['EmailAddress']
            active_flag = 1
            phone_number = res.get('Phone')
            random_suffix = ''.join(random.choices(string.ascii_letters + string.digits, k=6))
            userName = f"{firstName}{random_suffix}"
 
            
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=5))
                
            findEmail = """SELECT email FROM user WHERE email=%s"""
            cursor.execute(findEmail, email)
            emailId = cursor.fetchone()
            conn.commit()


            emailSendingList = []
            emailSendingList.append(email)
 
            print(email)
            print(emailId)
            
            if email == emailId['email']:
                print('hgyterwfdcgvnbjh')
                updateSql = """
                UPDATE user
                SET
                    firstName = %s,middleName = %s,lastName = %s,userType = %s,userName = %s,
                    password = %s, userId = %s,activeFlag = %s,mobileNumber = %s
                WHERE
                    email = %s
            """
 
                cursor.execute(updateSql, (firstName, middleName, lastName, role, userName, password, userId, active_flag, phone_number, email))
                conn.commit()
                textBody = ""
 
                subject = "TechnoStruct Login Credentials"
                htmlText = '''<span>Hi '''+ firstName +''',<br> Please find the login credentials below: . <br> username : ''' + userName + ''' <br> Password: '''+ password +'''<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
                
                obj = sendMail(emailSendingList,textBody,subject,htmlText,'')
                obj.sendmail()
                resp = jsonify({"status": 1, "success": True, "message": "User updated successfully"})
 
            else:
                sql = """
                INSERT INTO user (firstName, middleName, lastName, userType,userName,password, email, activeFlag, mobileNumber, userId, otp)
                VALUES (%s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(sql, (firstName, middleName, lastName, role,userName, password, email, active_flag, phone_number, userId, NULL))
                conn.commit()
                textBody = ""
 
                subject = "TechnoStruct Login Credentials"
                htmlText = '''<span>Hi '''+ firstName +''',<br> Please find the login credentials below: . <br> username : ''' + userName + ''' <br> Password: '''+ password +'''<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
                
                obj = sendMail(emailSendingList,textBody,subject,htmlText,'')
                obj.sendmail()
                resp = jsonify({"status": 1, "success": True, "message": "User added successfully"})
 
            
        elif role == 'Mentor':
            print('Mentor')
            name = res['name'].strip()
            
 
            name_parts = name.split(' ', 1)  
            firstName = name_parts[0]
            lastName = name_parts[1] if len(name_parts) > 1 else ''
            middleName = ''
        
 
            role = role
            userId = "0"
            email = res['email']
            emailSendingList = []
            emailSendingList.append(email)
            active_flag = 1
            phone_number = res.get('phoneNo')
            
            random_suffix = ''.join(random.choices(string.ascii_letters + string.digits, k=3))
            userName = f"{firstName}{random_suffix}"
 
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=5))
                
            findEmail = """SELECT email FROM user WHERE email=%s"""
            cursor.execute(findEmail, email)
            emailId = cursor.fetchone()
            conn.commit()
            
            
            if emailId is not None and email == emailId['email']:
                resp = jsonify({"status": 0, "success": False, "message": "Email Already Exists!"})
                    
            else:
                sql = """
                INSERT INTO user (firstName, middleName, lastName, userType,userName,password, email, activeFlag, mobileNumber, userId, otp)
                VALUES (%s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(sql, (firstName, middleName, lastName, role,userName, password, email, active_flag, phone_number, userId, NULL))
                conn.commit()
                textBody = ""
 
                subject = "TechnoStruct Login Credentials"
                htmlText = '''<span>Hi '''+ firstName +''',<br> Please find the login credentials below: . <br> username : ''' + userName + ''' <br> Password: '''+ password +'''<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
                
                obj = sendMail(emailSendingList,textBody,subject,htmlText,'')
                obj.sendmail()
        
                resp = jsonify({"status": 1, "success": True, "message": "User added successfully"})
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Failed to add user"})
    finally:
        cursor.close()
        conn.close()
        return resp
 
@user.route('/addAllUser', methods=['POST'])
@cross_origin()
def addAllUser():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        print(req)
           
        name = req['name'].strip()
        
        role = req['role']
        email = req['email']
        active_flag = req.get('activeFlag', '')
        phone_number = req.get('phoneNumber')
        emailSendingList = []
        emailSendingList.append(email)
        userId = "0"
        
        
        name_parts = name.split(' ', 1)  
        firstName = name_parts[0]
        lastName = name_parts[1] if len(name_parts) > 1 else ''
        middleName = ''
        random_suffix = ''.join(random.choices(string.ascii_letters + string.digits, k=3))
        userName = f"{firstName}{random_suffix}"

        password = ''.join(random.choices(string.ascii_letters + string.digits, k=5))
            
        
        findEmail = """SELECT email FROM user WHERE email=%s"""
        cursor.execute(findEmail, email)
        emailId = cursor.fetchone()
        print("---------->",emailId)
        conn.commit()
       

        # print("---------------->",emailId['email'])
        
         
        if emailId is not None:
            if email == emailId['email']:
                resp = jsonify({"status": 0, "success": False, "message": "Email Already Exists!"})
                        
        else:
            sql = """
            INSERT INTO user (firstName, middleName, lastName, userType,userName,password, email, activeFlag, mobileNumber, userId, otp)
            VALUES (%s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (firstName, middleName, lastName, role,userName, password, email, active_flag, phone_number, userId, NULL))
            conn.commit()
            textBody = ""
 
            subject = "TechnoStruct Login Credentials"
            htmlText = '''<span>Hi '''+ firstName +''',<br> Please find the login credentials below: . <br> username : ''' + userName + ''' <br> Password: '''+ password +'''<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            
            obj = sendMail(emailSendingList,textBody,subject,htmlText,'')
            obj.sendmail()
            resp= jsonify({"status": 1, "success": True, "message": "User added successfully"})
            
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Failed to add user"})
    finally:
        cursor.close()
        conn.close()
        return resp        
 
 
 
@user.route('/getAllUserDetails', methods=['POST'])
@cross_origin()
def getAllUserDetails():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        
        sql = """
        SELECT * FROM user where 1 ORDER BY id DESC
        """
        cursor.execute(sql)
        userList = cursor.fetchall()
        
        formattedUserList = []
        for user in userList:
            name = f"{user['firstName']}"
            if user['middleName']:
                name += f" {user['middleName']}"
            name += f" {user['lastName']}"
            
            formattedUser = {
                "id": user["id"],
                "firstName": user["firstName"],
                "lastName": user["lastName"],
                "middleName": user["middleName"],
                "email": user["email"],
                "userType": user["userType"],
                "activeFlag": user["activeFlag"],
                "name": name  
            }
            formattedUserList.append(formattedUser)
        
        resp = jsonify({"status": 1, "success": True, "data": formattedUserList, "message": "User fetched successfully"})
        
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Failed to get user list"})
        
    finally:
        cursor.close()
        conn.close()
        return resp
 
 
 
@user.route('/editUser', methods=['POST'])
@cross_origin()
def editUser():
    try:
        data = request.json
        user_id = data.get("id")
        active_flag = data.get("activeFlag")
 
        if not user_id or active_flag is None:
            return jsonify({"status": 0, "success": False, "message": "User ID and activeFlag are required"}), 400
 
        conn = connect_mysql()
        cursor = conn.cursor()
 
        sql = """
        UPDATE user
        SET activeFlag = %s
        WHERE id = %s
        """
        cursor.execute(sql, (active_flag, user_id))
        conn.commit()
 
        if cursor.rowcount == 0:
            resp = jsonify({"status": 0, "success": False, "message": "User not found"})
        else:
            resp = jsonify({"status": 1, "success": True, "message": "User active flag updated successfully"})
 
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Failed to update user active flag"})
 
    finally:
        cursor.close()
        conn.close()
        return resp
 
 
 
@user.route('/deleteUser', methods=['POST'])
@cross_origin()
def deleteUser():
    try:
        data = request.json
        user_id = data.get("id")
 
        if not user_id:
            return jsonify({"status": 0, "success": False, "message": "User ID is required"}), 400
 
        conn = connect_mysql()
        cursor = conn.cursor()
 
        sql = """
        DELETE FROM user
        WHERE id = %s
        """
        cursor.execute(sql, (user_id,))
        conn.commit()
 
        if cursor.rowcount == 0:
            resp = jsonify({"status": 0, "success": False, "message": "User not found"})
        else:
            resp = jsonify({"status": 1, "success": True, "message": "User deleted successfully"})
 
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Failed to delete user"})
 
    finally:
        cursor.close()
        conn.close()
        return resp
    



# RISHABH MUDGAL
@user.route('/sendOTP', methods=['POST'])
@cross_origin()
def sendOTP():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        otp = random.randint(100000, 999999)
        if validate_email(email):
            cursor.execute("""UPDATE user SET otp=%s WHERE email=%s""", (str(otp), email))
            conn.commit()

            body = f"""
                <pre>
                Your OTP for Resetting Password is: {otp}
                
                Thank You
                
                Pentation Analytics
                </pre>
            """

            send_email('OTP for Password Reset', body, email)

            return jsonify({"status": 1, "success": True})
        else:
            return jsonify({"status": 0, "success": False})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})


@user.route('/verifyOTP', methods=['POST'])
@cross_origin()
def verifyOTP():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        rcv_otp = data['otp']

        if validate_email(email) and rcv_otp:

            cursor.execute("""SELECT otp FROM user WHERE email=%s""", (email,))
            sent_otp = cursor.fetchone()['otp']

            if str(rcv_otp) != sent_otp:
                return jsonify({"status": 1, "success": False, "message": "OTP Verification Failed"})
            else:
                return jsonify({"status": 1, "success": True, "message": "OTP Verification Successful"})
        else:
            return jsonify({"status": 0, "success": False})
    except Exception as e:
        print(e)
        return jsonify({'status': 0})


@user.route('/changePassword', methods=['POST'])
@cross_origin()
def changePassword():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        password = data['password']
        if validate_email(email) and password:
            cursor.execute("""UPDATE user SET password=%s WHERE email=%s""", (password, email))
            conn.commit()

            return jsonify({"status": 1, "success": True, "message": "Password Changed"})
        else:
            return jsonify({"status": 0, "success": False})
    except Exception as e:
        print(e)
        return jsonify({'status': 0})


def validate_email(email):
    if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        return True
    else:
        return False
 

 #RishuRaj
@user.route('/changeCurrentPwd', methods=['POST'])
@cross_origin()
def changeCurrentPwd():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        currentPassword = data['currentPassword']
        newPassword = data['newPassword']

        pwdSql =  """
        SELECT password FROM user WHERE email=%s
        """
        cursor.execute(pwdSql,(email))
        pwd = cursor.fetchone()
        pwd = pwd['password']
        print("--->",pwd)
        
        if pwd == currentPassword:
            
            cursor.execute("""UPDATE user SET password=%s WHERE email=%s""", (newPassword, email))
            conn.commit()
            
            resp = jsonify({"status": 1, "success": True,"message": "Password changed successfully!"})
            
        else:
            resp = jsonify({"status": 0, "success": False,"message": "Current password is not matching, Check again!"})
        
    
    except Exception as e:
        print(e)
        resp = jsonify({'status': 0, "success" : False, "message" : "Some Error Occured!"})
    
    finally:
        cursor.close()
        conn.close()
        return resp