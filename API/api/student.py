# from flask import Blueprint
# from flask.json import jsonify
from flask import Blueprint, send_from_directory
from flask.json import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
import math, random
import os
import requests
import re
from api.sendMail import * 
from datetime import datetime
import traceback
from transformers import pipeline

student = Blueprint('student', __name__)



@student.route('/getStudentList',methods=['POST'])
@cross_origin()
def getStudentList():
    try:
        
        
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        course = req['course']
        status = req['status']
        
            
        studentListSql = " SELECT (CASE WHEN LastName is Null THEN FirstName ELSE concat(FirstName,' ',LastName) END) as Name,EmailAddress as emailID,mx_Course as Course,Phone as contactNo,status,id,batch,updatedAt  FROM `converted_student_data` Where 1 "
        if course != []:

            course = req['course'][0]['Course']
            studentListSql += " And mx_Course like '" +str(course)+ "'"
        
        if status != []:
            status = req['status'][0]['status']
            
            studentListSql += " And status like '" +str(status)+ "'"
        
        studentListSql += "ORDER BY `converted_student_data`.`id` DESC "
        cursor.execute(studentListSql)
        studentResult = cursor.fetchall()


        resp =  jsonify({"status":1, "success":True, "result": studentResult, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp

@student.route('/getCourseList',methods=['GET'])
@cross_origin()
def getCourseList():
    try:
        
        
        conn = connect_mysql()
        cursor = conn.cursor()
        searchSql = "SELECT DISTINCT(mx_Course) as Course FROM `converted_student_data` where mx_Course is not Null;"
        cursor.execute(searchSql)
        
        result = cursor.fetchall()

        resp =  jsonify({"status":1, "success":True, "courseList": result, "message": " Fetched successfully", })
    

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    

@student.route('/getBatchList',methods=['GET'])
@cross_origin()
def getBatchList():
    try:
        
        
        conn = connect_mysql()
        cursor = conn.cursor()
        searchSql = "SELECT DISTINCT(batchName) FROM `batchMaster` "
        cursor.execute(searchSql)
        
        result = cursor.fetchall()

        resp =  jsonify({"status":1, "success":True, "batchList": result, "message": " Fetched successfully", })
    

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    


@student.route('/sendRegistrationLink',methods = ['POST'])
@cross_origin()
def sendRegistrationLink():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        email = req['email']
        emailIdList =[]
        emailSearchSql = "SELECT * FROM `converted_student_data`  WHERE EmailAddress like '"+str(email)+"' "
        cursor.execute(emailSearchSql)
        res = cursor.fetchone()
        emailIdList =[]
        if res != None and len(res) != 0:

            textBody = ""
            subject = "Registration Link For "+str(res['mx_Course'])
            htmlText = '''<span>Hi '''+str(res['FirstName'])+''',<br> Thanks for your interest in ''' +str(res['mx_Course'])+'''. <br> Please complete your Registration Procedure. Your Registration Link is <a href= 'https://lms.technostructacademy.com/registration/'>Registration Link</a>.<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            # htmlText = '''<span>Hi '''+str(res['FirstName'])+''',<br> Thanks for your interest in ''' +str(res['mx_Course'])+'''. <br> Please complete your Registration Procedure. Your Registration Link is <a href= 'http://192.168.1.9:3001/lms/lms_reg/'>Registration Link</a>.<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            # htmlText = '''<span>Hi '''+str(res['FirstName'])+''',<br> Thanks for your interest in ''' +str(res['mx_Course'])+'''. <br> Please complete your Registration Procedure. Your Registration Link is <a href= 'http://192.168.29.171:3001/lms/lms_reg/'>Registration Link</a>.<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            

            
            # print(htmlText)
            emailIdList.append(email)
            obj = sendMail(emailIdList,textBody,subject,htmlText,'')
            obj.sendmail()


        resp = jsonify({"status":1, "success":True,  "message": " Mail Sent successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp


# sendRegistrationLink('arpand@pentationanalytics.com')

@student.route('/getCourseDoc', methods=['POST'])
@cross_origin()
def getCourseDoc():
    try:
        # Get JSON data from the request
        data = request.get_json()
        print("---> Received data:", data)
        courseName = data.get("email")
        # Connect to the database
        conn = connect_mysql()
        cursor = conn.cursor()


        # Fetch data to create the PDF
        pdfSqlQuery = """
            SELECT Course
            FROM StudentMaster
            WHERE emailID=%s
        """
        cursor.execute(pdfSqlQuery,courseName)
        rows = cursor.fetchall()
        
        print("data in the rows-->",rows)
        courseName=''
        for row in rows:
            courseName = row['Course']

        courseSql = " SELECT courseName , `folderName` ,activeFlag FROM `courseMaster` WHERE courseName = %s"
        print(courseSql)
        cursor.execute(courseSql,courseName)
        courseList = cursor.fetchall()
        print("\n\n",courseList)

        folder_names = [course['folderName'] for course in courseList]
        print("courseList['folderName']", folder_names)

       
        result_list = []

        for course in courseList:
            folder_name = course['folderName']
            latest_file_path = None
            latest_timestamp = None

            if folder_name:
                folders_and_files = folder_name.split(',')
                for entry in folders_and_files:
                    parts = entry.split('/')
                    if len(parts) == 2:
                        file_name = parts[1].strip()
                        timestamp_match = re.search(r'(\d{4}-\d{2}-\d{2}-\d{2}-\d{2}-\d{2})', file_name)
                        if timestamp_match:
                            timestamp = timestamp_match.group(1)
                            if not latest_timestamp or timestamp > latest_timestamp:
                                latest_timestamp = timestamp
                                latest_file_path = entry.strip()

            if latest_file_path:
                result_list.append({
                    "activeFlag": course["activeFlag"],
                    "courseName": course["courseName"],
                    "folderName": [
                        {
                            "file": latest_file_path.split('/')[-1],
                            "path": latest_file_path
                        }
                    ]
                })

        if result_list:
            resp = jsonify({"status": 1, "success": True, "result": result_list, "message": "Fetched successfully"})
        else:
            resp = jsonify({"status": 0, "message": "No valid files found in folderName"})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": str(e)})
    finally:
        cursor.close()
        conn.close()
        return resp


       
def get_data( api):
        print(api)
        response = requests.post(f"{api}")
        if response.status_code == 200:
            print("sucessfully fetched the data")
            print(response)
#             self.formatted_print(response.json())
        else:
            print(f"Hello person, there's a {response.status_code} error with your request")

api = "https://r-sg.leadsquared.com/rdPage.aspx?rdReport=Dashlets.Leads.Drill.LeadList&Field1=ProspectStage&Value1=Converted%20%E2%80%93%20100%25&Mode=OneField&AuthToken=c8c204f92390754ae8b4dc080ac71ab9c9195c1f10263c92251f669778d6f9fe&ReportTitle=List+of+Leads+where+Lead+Stage+is+Converted%20%E2%80%93%20100%25" 

# get_data( api)



@student.route('/updateStudentList', methods=['POST'])
@cross_origin()
def updateStudentList():
    conn = connect_mysql()
    cursor = conn.cursor()

    try:
        req = request.json
        mail = req['mail']
        batch = req['batchname'][0]['batchname']

        stmt = "UPDATE studentRegistrationDetails SET batch = %s WHERE email = %s"
        cursor.execute(stmt, (batch, mail))
        stmt = "UPDATE converted_student_data SET batch = %s WHERE emailAddress = %s"
        cursor.execute(stmt, (batch, mail))

        return jsonify({"status": 1, "success": True, "result": 'Updated successfully' })

    except Exception as e:
        print(e)

        return jsonify({"status": 0})
    finally:
        conn.close()



# RISHABH MUDGAL
# @student.route('/insert_placement', methods=['POST'])
# @cross_origin()
# def insert_placement():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json

#         company_name = data['companyName'] if data['companyName'] != '' else None
#         # course = str(data['course'])  if data['course'] != '' else None
#         industry_type = data['industryType'] if data['industryType'] != '' else None
#         employment_type = data['employementType'] if data['employementType'] != '' else None
#         graduation_year = data['graduationYear'] if data['graduationYear'] != '' else None
#         location = data['jobLocation'] if data['jobLocation'] != '' else None
#         designation = data['designation'] if data['designation'] != '' else None
#         date_of_arrival = data['dateOfArrival'] if data['dateOfArrival'] != '' else None
#         relevantExperience = data['relevantExperience'] if data['relevantExperience'] != '' else None
#         totalExperience = data['totalExperience'] if data['totalExperience'] != '' else None
#         salary_range = data['salaryRange'] if data['salaryRange'] != '' else None
#         placement_coordinator = data['placementCoOrdinator'] if data['placementCoOrdinator'] != '' else None
#         status =  None
#         batch =str(data['batch'])  if data['batch'] != '' else None
#         required_skills = data['requiredSkills'] if data['requiredSkills'] != '' else None
#         job_description = data['jobDescription'] if data['jobDescription'] != '' else None
#         total_rounds = data['totalRounds'] if data['totalRounds'] != '' else None
#         # print(type(graduation_year),graduation_year)
#         STMT = """INSERT INTO placement (company_name, industry_type, employment_type, graduation_year, location, designation, date_of_arrival, relevant_experience,total_experience, salary_range, placement_coordinator, status, batch, required_skills, job_description, total_rounds ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)"""

#         cursor.execute(STMT, (
#             company_name, industry_type, employment_type, graduation_year, location, designation,
#             date_of_arrival,
#             relevantExperience, totalExperience, salary_range, placement_coordinator, status, batch, required_skills, job_description,
#             total_rounds))
#         conn.commit() 
#         return jsonify({"status": 1, "success": True, "message": " Inserted successfully"})

#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})
#     finally:
#         conn.close()


@student.route('/insert_placement', methods=['POST'])
@cross_origin()
def insert_placement():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json

        company_name = data['companyName'] if data['companyName'] != '' else None
        # course = str(data['course'])  if data['course'] != '' else None
        industry_type = data['industryType'] if data['industryType'] != '' else None
        employment_type = data['employementType'] if data['employementType'] != '' else None
        graduation_year = data['graduationYear'] if data['graduationYear'] != '' else None
        location = data['jobLocation'] if data['jobLocation'] != '' else None
        designation = data['designation'] if data['designation'] != '' else None
        date_of_arrival = data['dateOfArrival'] if data['dateOfArrival'] != '' else None
        relevantExperience = data['relevantExperience'] if data['relevantExperience'] != '' else None
        totalExperience = data['totalExperience'] if data['totalExperience'] != '' else None
        salary_range = data['salaryRange'] if data['salaryRange'] != '' else None
        placement_coordinator = data['placementCoOrdinator'] if data['placementCoOrdinator'] != '' else None
        status =  None
        batch =str(data['batch'])  if data['batch'] != '' else None
        studentList = str(data['studentList']) if data['studentList'] != '' else None        #Rishu
        required_skills = data['requiredSkills'] if data['requiredSkills'] != '' else None
        job_description = data['jobDescription'] if data['jobDescription'] != '' else None
        total_rounds = data['totalRounds'] if data['totalRounds'] != '' else None
        companyEmail = data['companyEmail'] if data['companyEmail'] != '' else None


        # print(type(graduation_year),graduation_year)
        STMT = """INSERT INTO placement (company_name, industry_type, employment_type, graduation_year, location, designation, date_of_arrival, relevant_experience,total_experience, salary_range, placement_coordinator, status, batch, required_skills, job_description, total_rounds, studentList, companyEmail) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s)"""

        cursor.execute(STMT, (
            company_name, industry_type, employment_type, graduation_year, location, designation,
            date_of_arrival,
            relevantExperience, totalExperience, salary_range, placement_coordinator, status, batch, required_skills, job_description,
            total_rounds, studentList, companyEmail))
        conn.commit() 
        return jsonify({"status": 1, "success": True, "message": " Inserted successfully"})

    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        conn.close()

@student.route('/getPlacementById', methods=['POST'])
@cross_origin()
def getPlacementById():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        id = data['id']
        print(id)
        STMT = """SELECT * FROM placement WHERE id=%s"""

        cursor.execute(STMT,id)
        placementDetails = cursor.fetchall()

        data = []
        for row in placementDetails:
            data.append({
                'id': row['id'],
                'companyName': row['company_name'],
                # 'course': eval(row['course']),
                'dateOfArrival': row['date_of_arrival'],
                'salaryRange': row['salary_range'],
                'industryType': row['industry_type'],
                'placementCoOrdinator': row['placement_coordinator'],
                'jobLocation': row['location'],
                'employementType': row['employment_type'],
                'graduationYear': row['graduation_year'],
                'totalExperience': row['total_experience'],
                'relevantExperience': row['relevant_experience'],
                'designation': row['designation'],
                'totalRounds': row['total_rounds'],
                'batch': eval(row['batch']),
                'studentList':eval(row['studentList']),
                'requiredSkills': row['required_skills'],
                'jobDescription': row['job_description']
            })




        print("\n\n",placementDetails)
        return jsonify({"status": 1, "success": True, "message": " Updated successfully", "data": data})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})

    finally:
        conn.close()


@student.route('/update_placement', methods=['POST'])
@cross_origin()
def update_placement():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        print(data)
        id = data['id']
        print('idddddddddddddddddddddddd---------->',id)
        company_name = data['companyName'] if data['companyName'] != '' else None
        # course = data['course'] if data['course'] != '' else None
        industry_type = data['industryType'] if data['industryType'] != '' else None
        employment_type = data['employementType'] if data['employementType'] != '' else None
        graduation_year = int(data['graduationYear']) if data['graduationYear'] != '' else 0
        location = data['jobLocation'] if data['jobLocation'] != '' else None
        designation = data['designation'] if data['designation'] != '' else None
        date_of_arrival = data['dateOfArrival'] if data['dateOfArrival'] != '' else None
        relevantExperience = data['relevantExperience'] if data['relevantExperience'] != '' else None
        totalExperience = data['totalExperience'] if data['totalExperience'] != '' else None
        salary_range = data['salaryRange'] if data['salaryRange'] != '' else None
        placement_coordinator = data['placementCoOrdinator'] if data['placementCoOrdinator'] != '' else None
        status = None
        # batch = data['batch'] if data['batch'] != '' else None
        studentList = str(data['studentList']) if data['studentList'] != '' else None 
        required_skills = data['requiredSkills'] if data['requiredSkills'] != '' else None
        job_description = data['jobDescription'] if data['jobDescription'] != '' else None
        total_rounds = data['totalRounds'] if data['totalRounds'] != '' else None
       



        batch = json.dumps(data['batch']) if data['batch'] else None
        # course = json.dumps(data['course']) if data['course'] else None

        STMT = """UPDATE placement SET 
            company_name=%s,  
            industry_type=%s, 
            employment_type=%s, 
            graduation_year=%s, 
            location=%s, 
            designation=%s, 
            date_of_arrival=%s, 
            relevant_experience=%s,
            total_experience=%s, 
            salary_range=%s, 
            placement_coordinator=%s, 
            status=%s, 
            batch=%s, 
            required_skills=%s, 
            job_description=%s, 
            total_rounds=%s,
            studentList=%s        
          WHERE id=%s"""

        print('STMT',STMT)
        cursor.execute(STMT, (company_name, industry_type, employment_type, graduation_year, location, designation,
            date_of_arrival, relevantExperience, totalExperience, salary_range, placement_coordinator, status, batch, required_skills,
            job_description, total_rounds,studentList, id
        ))

        return jsonify({"status": 1, "success": True, "message": " Updated successfully", }) , 200
    except Exception as e:
        print(e)
        return jsonify({"status": 0}), 500
  
    finally:
        conn.close()


@student.route('/deletePlacementById', methods=['POST'])
@cross_origin()
def deletePlacementById():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json  
        id = data['id']      
        print(f"Deleting placement with ID: {id}")
        
        STMT = """DELETE FROM placement WHERE id=%s"""
        
        cursor.execute(STMT, (id))  
        conn.commit() 

        return jsonify({"status": 1, "success": True, "message": "Placement deleted successfully"})
    except Exception as e:
        print(e)
        return jsonify({"status": 0, "success": False, "message": "Failed to delete placement"})
    finally:
        cursor.close()  
        conn.close()    

 


@student.route('/getAllPlacement', methods=['GET'])
@cross_origin()
def getAllPlacement():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM placement")
        result = cursor.fetchall()
        
        final_result = []
        for row in result:
            final_result.append({
                'id': row['id'],
                "companyName": row['company_name'],
                # "course": eval(row['course']),
                "industryType": row['industry_type'],
                "employementType": row['employment_type'],
                "graduationYear": row['graduation_year'],
                "jobLocation": row['location'],
                "jobDescription": row['job_description'],
                "dateOfArrival": row['date_of_arrival'],
                # "experience": row['experience'],
                'relevantExperience':row['relevant_experience'],
                'totalExperience':row['total_experience'],
                "salaryRange": row['salary_range'],
                "placementCoOrdinator": row['placement_coordinator'],
                "status": row['status'],
                "batch": eval(row['batch']),
                "requiredSkills": row['required_skills'],
                "totalRounds": row['total_rounds'],
                'designation':row['designation']
            })
        
        return jsonify({"status": 1, "success": True, "result": final_result})
    except Exception as e:
        print(e)
        return jsonify({"status": 0, "error": str(e)})
    finally:
        cursor.close()
        conn.close()



@student.route('/getAllStudentsApplied', methods=['GET'])
@cross_origin()
def getAllStudentsApplied():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT * FROM placementStudentMaster""")
        data = cursor.fetchall()
        return jsonify({"status": 1, "success": True, "data": data})

    except Exception as e:
        print(e)
        return jsonify({"status": 0})        
 

#  Taken from chitradip 22 AUG

# @student.route('/getPlacementForStud', methods=['POST'])
# @cross_origin()
# def getPlacementForStud():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data.get('studentEmail')
#         batchstmt = """SELECT batch FROM converted_student_data WHERE EmailAddress=%s"""
#         cursor.execute(batchstmt, (email))
#         batchname = cursor.fetchone()
 
#         placement_stmt = """SELECT * FROM placement """
#         cursor.execute(placement_stmt)
#         result = cursor.fetchall()
#         # temp = []
#         # for row in result:
#         #     print(row)
#         #     if row['batch']  is not None and  json.loads(row['batch'])['batchName'] == batchname['batch'] :
#         #         temp.append(row)
#         # print(batchname['batch'])

#         temp = []
#         for row in result:
            
            
#             batch_data = json.loads(row['batch']) if row['batch'] else []
#             print('------------------------------->',batch_data)
#             batch_names = [batch['batchName'] for batch in batch_data]
#             print('=--=-=->',batch_names)
#             if batchname in batch_names:
#                 temp.append(row)
        

#         final_result = []
#         for row in temp:
#             # print("row['course']>>>>>",row['course'])
#             # print("eval(row['course'])",eval(row['course']))
 
#             final_result.append({
#                 "companyName": row['company_name'],
#                 "course":eval(row['course']) ,
#                 "industryType": row['industry_type'],
#                 "employementType": row['employment_type'],
#                 "graduationYear": row['graduation_year'],
#                 "jobLocation": row['location'],
#                 "jobDescription": row['job_description'],
#                 "dateOfArrival": row['date_of_arrival'],
#                 "experience": row['experience'],
#                 "salaryRange": row['salary_range'],
#                 "designation":row['designation'],
#                 "placementCoOrdinator": row['placement_coordinator'],
#                 "status": row['status'],
#                 "batch": eval(row['batch']),
#                 "requiredSkills": row['required_skills'],
#                 "totalRounds": row['total_rounds']
#             })
 
#         return jsonify({"status": 1, "success": True, "result": final_result})
#     # except Exception as e:
#     #     print(e)
#     #     return jsonify({"status": 0})

#     finally:
#         conn.close()


# @student.route('/getPlacementForStud', methods=['POST'])
# @cross_origin()
# def getPlacementForStud():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data.get('studentEmail')
#         cursor.execute("""SELECT batch FROM converted_student_data WHERE EmailAddress=%s""", (email,))
#         batchname = cursor.fetchone()
        
#         placement_stmt = """SELECT * FROM placement"""
#         cursor.execute(placement_stmt)
#         result = cursor.fetchall()

#         temp = []
#         for row in result:
#             print(row['company_name'], row['batch'], batchname['batch'])
#             if row['batch'] is not None:
#                 for batches in json.loads(row['batch'].replace("'", '"')):
#                     if batches['batchName'] == batchname['batch']:
#                         print('add ', row['company_name'])
#                         temp.append(row)
#                         # print(temp)
#             # print(temp, 'temptemp')

#             # if row['batch'] is not None and json.loads(row['batch'])['batch'] == batchname['batch']:
#             #     print('hiiiiiiiiiiii')
#             #     temp.append(row)

#         final_result = []
#         for row in temp:
#             final_result.append({"companyName": row['company_name'],
#                                  "industryType": row['industry_type'], "employementType": row['employment_type'],
#                                  "graduationYear": row['graduation_year'], "jobLocation": row['location'],
#                                  "jobDescription": row['job_description'], "dateOfArrival": row['date_of_arrival'],
#                                  "total_experience": row['total_experience'], "relevant_experience": row['relevant_experience'], "salaryRange": row['salary_range'], "designation": row['designation'],
#                                  "placementCoOrdinator": row['placement_coordinator'], "status": row['status'],
#                                  "batch": eval(row['batch']), "requiredSkills": row['required_skills'],
#                                  "totalRounds": row['total_rounds']})

#         return jsonify({"status": 1, "success": True, "result": final_result})
#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})

@student.route('/getPlacementForStud', methods=['POST'])
@cross_origin()
def getPlacementForStud():
    conn = connect_mysql()
    cursor = conn.cursor()  # Using dictionary=True for easier access
    try:
        data = request.json
        email = data.get('studentEmail')
               
        placement_stmt = """SELECT * FROM placement"""
        cursor.execute(placement_stmt)
        result = cursor.fetchall()
        
        print("--------->", result)
        temp = []
        for row in result:
            students_list_text = row['studentList']
            
            if students_list_text is None:
                continue
            
            students_list_text = students_list_text.replace("'", '"')
            
          
            try:
                students_list = json.loads(students_list_text)
            except json.JSONDecodeError:
                print("Error parsing studentList JSON for row:", row)
                continue
            
            for student in students_list:
                if student.get('email') == email:
                    # print('add', row['company_name'])
                    temp.append(row)
                    

        final_result = []
        for row in temp:
            final_result.append({"companyId": row['id'],
                                 "companyName": row['company_name'],
                                 "industryType": row['industry_type'], "employementType": row['employment_type'],
                                 "graduationYear": row['graduation_year'], "jobLocation": row['location'],
                                 "jobDescription": row['job_description'], "dateOfArrival": row['date_of_arrival'],
                                 "total_experience": row['total_experience'], "relevant_experience": row['relevant_experience'], "salaryRange": row['salary_range'], "designation": row['designation'],
                                 "placementCoOrdinator": row['placement_coordinator'], "status": row['status'],
                                 "batch": eval(row['batch']), "requiredSkills": row['required_skills'],
                                 "totalRounds": row['total_rounds']})

        return jsonify({"status": 1, "success": True, "result": final_result})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})

 

def preprocess_json_string(json_string):
    # Replace single quotes with double quotes and ensure proper formatting
    json_string = re.sub(r"(?<!\w)'(?!\w)", '"', json_string)  # Replace single quotes around keys and values
    json_string = re.sub(r'(?<!\\)"', '\\"', json_string)  # Escape any unescaped double quotes
    json_string = '{' + re.sub(r"(\w+):", r'"\1":', json_string) + '}'  # Add double quotes around keys
    return json_string
 
 
@student.route('/getCompanyStatusStud', methods=['POST'])
@cross_origin()
# def getCompanyStatusStud():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data['studentEmail']
 
#         cursor.execute("""SELECT companyName, status FROM placementStudentMaster WHERE email=%s""", (email))
#         applied_companies = cursor.fetchall()
#         data = {
#     company['companyName']: 1 if company['status'] in ['Applied', 'Rejected',
#                                                        'Viewed', 'Accepted'] else 0
#     for company in applied_companies
# }
 
#         return jsonify({"status": 1, "success": True, "result": data})
#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})  
 
def getCompanyStatusStud():
 
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
 
        cursor.execute("""SELECT offerLetterPath, companyName, status, roundsCleared FROM placementStudentMaster WHERE email=%s""", (email))
        applied_companies = cursor.fetchall()
        data = {}
        # print("applied_companies",applied_companies)
        for company in applied_companies:
           data[company['companyName']] = company['status']
        #    print("data....",data)
        dataRound = {company['companyName']: company['roundsCleared'] for company in applied_companies}
        offerLetter = {company['companyName']: company['offerLetterPath'] for company in applied_companies}
        # print("dataRound",dataRound)
        return jsonify({"status": 1, "success": True, "statusData": data, "roundsData": dataRound, "offerLetterData" : offerLetter})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
 
 


@student.route('/updateSingleStudent', methods=['POST'])
@cross_origin()
def updateSingleStudent():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        status = data['placementInterviewUpdates']['interviewStatus']
        rounds_cleared = data['placementInterviewUpdates']['roundsClear']
        companyName = data['companyName']
        cursor.execute("""UPDATE placementStudentMaster SET status=%s, roundsCleared=%s WHERE email=%s AND companyName=%s""",
                       (status, rounds_cleared, email,companyName))

        return jsonify({"status": 1, "success": True, "message": " Updated successfully", })
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    

    
@student.route('/deleteSingleStudent', methods=['POST'])
@cross_origin()
def deleteSingleStudent():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['id']
        cursor.execute("""DELETE FROM placementStudentMaster WHERE id=%s""", (email))
        return jsonify({"status": 1, "success": True, "message": " Deleted successfully", })
    except Exception as e:
        print(e)
        return jsonify({"status": 0})    
 


@student.route('/studentPlacementapply', methods=['POST'])
@cross_origin()
def studentPlacementapply():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.form
        file = request.files.get('file')
        email = data.get('studentEmail')
        companyId=data.get('companyId')
        print(companyId)
 
        username = email.split('@')[0]
 
        studName = ''.join(word.capitalize() for word in re.split(r'[^a-zA-Z]', username) if word)
 
        # print(studName)
        
        path = f'{UPLOAD_STUD_RESUME_PATH}/{studName}'
        if not os.path.exists(path):
            os.mkdir(path)
        file.save(f'{path}/{studName}.pdf')
        
        resumePath = f"{studName}/{studName}.pdf"
        
        sql = """SELECT FirstName, LastName, batch FROM converted_student_data WHERE EmailAddress=%s"""
        cursor.execute(sql,(email))
        
        studentDetails = cursor.fetchone()
        # print("-------------->",resumePath)
 
 
        cursor.execute(
            """INSERT INTO placementStudentMaster (studentName, email, companyName, resume_cv, status, batchName, jobTitle, roundsCleared, companyId) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (f"{studentDetails['FirstName']} {studentDetails['LastName']}", email, data.get('companyName'), resumePath,
             '0', studentDetails['batch'], data.get('jobTitle'), 0, companyId))
        conn.commit()
 
        return jsonify({"status": 1, "success": True, "message": " Inserted successfully", })
 
 
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
       conn.close()

# from taken

@student.route('/getPlacementStatus', methods=['POST'])
@cross_origin()
def getPlacementStatus():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
 
        cursor.execute("SELECT companyName, status FROM placementStudentMaster WHERE email=%s", (email,))
        placedSql = cursor.fetchall()
 
        companies = []
        placement_status = False
 
        for record in placedSql:
            if record['status'] == '4':
                companies.append(record['companyName'])
                placement_status = True
        data = {
            "companyName": ", ".join(companies),  
            "status":placement_status
                    
                    }
            
        return jsonify({"status": 1, "success": True, "statusData": data})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        cursor.close()
        conn.close()
    


# @student.route('/getAssignmentData', methods=['POST'])
# @cross_origin()
# def getAssignmentData():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data['studentEmail']
 
#         cursor.execute("SELECT batch FROM converted_student_data WHERE EmailAddress=%s", (email,))
#         batchSql = cursor.fetchall()
#         if batchSql:
#             batchSql = batchSql[0]['batch']
        
 
#         print(batchSql)
        
#         data = []
        
#         cursor.execute("SELECT * FROM assignmentMaster WHERE JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s)) = 1", (batchSql,))
#         assignmentSql = cursor.fetchall()
#         for assignment in assignmentSql:
#             data.append(assignment)      
                    
#         return jsonify({"status": 1, "success": True, "result": data})
#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})
#     finally:
#         cursor.close()
#         conn.close()
 





def parse_excel(file_path):
    json_data = pd.read_excel(file_path)
    json_data.dropna(axis=0, how='all', inplace=True)

    question_list = []
    for row in json_data.iterrows():
        question_list.append({
            "questionText": row[1]['Question Text'],
            "marks": row[1]['Marks'],
            "options": [
                {"text": row[1]['Option A'], "isAnswer": 'true' if row[1]['Answer'] == 'Option A' else 'false'},
                {"text": row[1]['Option B'], "isAnswer": 'true' if row[1]['Answer'] == 'Option B' else 'false'},
                {"text": row[1]['Option C'], "isAnswer": 'true' if row[1]['Answer'] == 'Option C' else 'false'},
                {"text": row[1]['Option D'], "isAnswer": 'true' if row[1]['Answer'] == 'Option D' else 'false'}
            ]
        })
    return question_list


def parse_excel_des(file_path):
    json_data = pd.read_excel(file_path)
    json_data.dropna(axis=0, how='all', inplace=True)

    question_list = []
    for row in json_data.iterrows():
        question_list.append({
            "uploadedQuestion": row[1]['Question Text'],
            "marks": row[1]['Marks']
            # "options": [
            #     {"text": row[1]['Option A'], "isAnswer": 'true' if row[1]['Answer'] == 'Option A' else 'false'},
            #     {"text": row[1]['Option B'], "isAnswer": 'true' if row[1]['Answer'] == 'Option B' else 'false'},
            #     {"text": row[1]['Option C'], "isAnswer": 'true' if row[1]['Answer'] == 'Option C' else 'false'},
            #     {"text": row[1]['Option D'], "isAnswer": 'true' if row[1]['Answer'] == 'Option D' else 'false'}
            # ]
        })
    return question_list


# @student.route('/insertTest', methods=['POST'])
# @cross_origin()

# def insertTest():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.form
#         addInstructions = data.get('addInstructions') or None
#         batchName = data.get('batch') or None
#         endDate = data.get('endDate') or None
#         endTime = data.get('endTime') or None
#         evaluator = data.get('evaluator') or None
#         passingMarks = data.get('passingMarks') or None
#         startDate = data.get('startDate') or None
#         startTime = data.get('startTime') or None
#         testName = data.get('testName') or None
#         totalMarks = data.get('totalMarks') or None
#         totalTime = data.get('totalTime') or None
#         updQesAndMarks = data.get('updQesAndMarks') or None
#         uploadFileMCQ = data.get('uploadFileMCQ') or None
#         uploadFileDes = data.get('uploadFileDes') or None
#         mentorEmail = data.get('mentorEmail') or None

#         if request.files.get('uploadFileMCQ') or request.files.get('uploadFileDes'):
#             file1 = request.files.get('uploadFileMCQ') or None
#             file2 = request.files.get('uploadFileDes') or None
#             if not os.path.exists('ALL_TESTS'):
#                 os.mkdir('ALL_TESTS')
#             path = 'ALL_TESTS'
#             if file1:
#                 file1.save(f'{path}/{file1.filename or None}')
#             elif file2:
#                 file2.save(f'{path}/{file2.filename or None}')

#             json_data = parse_excel(f'{path}/{file1.filename}') if file1 else parse_excel_des(
#                 f'{path}/{file2.filename}')
#             print(json.dumps(json_data))

#             cursor.execute("""INSERT INTO tests (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName, totalMarks, totalTime, updQesAndMarks, uploadFileMCQ, uploadFileDes, mentorEmail)
#                  VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
#                 addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName,
#                 totalMarks, totalTime, json.dumps(json_data), file1.filename if file1 else None,
#                 file2.filename if file2 else None, mentorEmail))
#         else:
#             cursor.execute(
#                 """INSERT INTO tests (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName, totalMarks, totalTime, updQesAndMarks, mentorEmail  ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
#                 (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName,
#                  totalMarks, totalTime, updQesAndMarks, mentorEmail))
#         conn.commit()

#         return jsonify({"status": 1, "success": True, "message": " Inserted successfully"})

#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})






@student.route('/insertTest', methods=['POST'])
@cross_origin()
def insertTest():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.form or request.json
        addInstructions = data.get('addInstructions') or None
        batchName = data.get('batch') or None
        endDate = data.get('endDate') or None
        endTime = data.get('endTime') or None
        evaluator = data.get('evaluator') or None
        passingMarks = data.get('passingMarks') or None
        startDate = data.get('startDate') or None
        startTime = data.get('startTime') or None
        testName = data.get('testName') or None
        totalMarks = data.get('totalMarks') or None
        totalTime = data.get('totalTime') or None
        updQesAndMarks = data.get('updQesAndMarks') or None
        uploadFileMCQ = data.get('uploadFileMCQ') or request.files.get('uploadFileMCQ') or None
        uploadFileDes = data.get('uploadFileDes') or request.files.get('uploadFileDes') or None
        mentorEmail = data.get('mentorEmail') or None

        # batch_data = [{"batchName": "pentation_2024", "id": 43}, {"batchName": "pentation_2025", "id": 44}]
        batch_names = [item['batchName'] for item in eval(batchName)]

        # Create directory if it does not exist
        path = 'ALL_TESTS'
        if not os.path.exists(path):
            os.mkdir(path)

        if uploadFileMCQ:
            # Save MCQ file
            file_path = os.path.join(path, uploadFileMCQ.filename)
            uploadFileMCQ.save(file_path)

            # Process and parse Excel file
            json_data = parse_excel(file_path)
            print(json.dumps(json_data))

            # Insert into database
            cursor.execute("""INSERT INTO tests (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName, totalMarks, totalTime, updQesAndMarks, uploadFileMCQ, uploadFileDes, mentorEmail)
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName,
                totalMarks, totalTime, json.dumps(json_data), uploadFileMCQ.filename, None, mentorEmail))

        elif uploadFileDes:
            # Save descriptive file
            file_path = os.path.join(path, f'{"-".join(batch_names)}-{testName}.pdf')
            uploadFileDes.save(file_path)

            # Insert into database
            cursor.execute("""INSERT INTO tests (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName, totalMarks, totalTime, uploadFileMCQ, mentorEmail, uploadFileDes)
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName,
                totalMarks, totalTime, None, mentorEmail, f'{"-".join(batch_names)}-{testName}.pdf'))

        else:
            # Insert into database without files
            cursor.execute("""INSERT INTO tests (addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName, totalMarks, totalTime, updQesAndMarks, mentorEmail)
                              VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (
                addInstructions, batchName, endDate, endTime, evaluator, passingMarks, startDate, startTime, testName,
                totalMarks, totalTime, updQesAndMarks, mentorEmail))

        placeholders = ', '.join(['%s'] * len(batch_names))  # Adjust for your DB library
        query = f"""
            SELECT FirstName, LastName, EmailAddress, batch
            FROM converted_student_data
            WHERE batch IN ({placeholders})
        """
        cursor.execute(query, batch_names)

        student_test_assigned_for = cursor.fetchall()

        adding_test = 'INSERT INTO studentTestStatus (name, email, batchName, testName, status) VALUES  (%s, %s, %s, %s, %s)'

        data = []

        for student in student_test_assigned_for:
            data.append((f'{student["FirstName"]} {student["LastName"]}', f'{student["EmailAddress"].strip(" ")}',
                         f'{student["batch"].strip(" ")}', f'{testName.strip()}', 'Not Started'))

        cursor.executemany(adding_test, data)

        conn.commit()

        return jsonify({"status": 1, "success": True, "message": " Inserted successfully"})

    except Exception as e:
        print(e)
        return jsonify({"status": 0})
 


@student.route('/getTests', methods=['POST'])
@cross_origin()
def getTests():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
 
        cursor.execute("""SELECT * FROM tests WHERE mentorEmail=%s ORDER BY id DESC""", (email,))
        data = cursor.fetchall()
 
        return jsonify({"status": 1, "success": True, "data": data})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})      
    

    #RISHU AUG 28
# @student.route('/uploadAssignment', methods=['POST'])
# @cross_origin()
# def uploadAssignment():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         file = request.files.get('file')
#         studentEmail = request.form.get('studentEmail')
#         assignmentName = request.form.get('assignmentName')
#         assignmentId = request.form.get('assignmentId')

#         # Fetch batch and student name
#         cursor.execute("SELECT batch, FirstName FROM converted_student_data WHERE EmailAddress=%s", (studentEmail,))
#         batchSql = cursor.fetchall()
#         if batchSql:
#             batch = batchSql[0]['batch']
#             studentName = batchSql[0]['FirstName']
#         else:
#             return jsonify({"status": 0, "message": "Student not found!"})

#         # Handle file upload
#         if file:
#             filename, file_extension = os.path.splitext(file.filename)
#             # new_filename = f"{studentName}_{filename}{file_extension}"
#             new_filename = f"{studentName}_{assignmentName}_{filename}{file_extension}"
#             batch_folder = os.path.join(UPLOAD_STUD_ASSIGNMENT_PATH, batch)
#             if not os.path.exists(batch_folder):
#                 os.makedirs(batch_folder)
            
#             file_path = os.path.join(batch_folder, new_filename)
#             file.save(file_path)
#             assignmentPath = f"studentAssignmentUploads/{batch}/{new_filename}"
#         else:
#             return jsonify({"status": 0, "message": "No file found!"})

#         current_timestamp = datetime.now()
#         submittedDate = current_timestamp.strftime('%Y-%m-%d')

#         activeFlag = 1

#         # Check if the record exists
#         checkSql = """
#             SELECT COUNT(*) as count FROM assignmentStudentMaster
#             WHERE assignmentId = %s AND assignmentName = %s AND email = %s
#         """
#         cursor.execute(checkSql, (assignmentId, assignmentName, studentEmail))
#         count = cursor.fetchone()['count']

#         if count > 0:
#             # Update existing record
#             updateSql = """
#                 UPDATE assignmentStudentMaster
#                 SET submittedDate = %s, assignmentPath = %s, activeFlag = %s
#                 WHERE assignmentId = %s AND assignmentName = %s AND email = %s
#             """
#             cursor.execute(updateSql, (submittedDate, assignmentPath, activeFlag, assignmentId, assignmentName, studentEmail))
#         else:
#             # Insert new record
#             insertSql = """
#                 INSERT INTO assignmentStudentMaster (assignmentId, assignmentName, submittedDate, email, batch, assignmentPath, activeFlag)
#                 VALUES (%s, %s, %s, %s, %s, %s, %s)
#             """
#             cursor.execute(insertSql, (assignmentId, assignmentName, submittedDate, studentEmail, batch, assignmentPath, activeFlag))

#         conn.commit()

#         resp = jsonify({"status": 1, "success": True, "message": "Assignment uploaded successfully"})

#     except Exception as e:
#         print(e)
#         conn.rollback()
#         resp = jsonify({"status": 0, "message": "An error occurred"})

#     finally:
#         cursor.close()
#         conn.close()

#     return resp


@student.route('/downloadAssignmentStatus', methods=['POST'])
@cross_origin()
def downloadAssignmentStatus():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        studentEmail = request.form.get('studentEmail')
        assignmentName = request.form.get('assignmentName')
        assignmentId = request.form.get('assignmentId')

        cursor.execute("SELECT batch, FirstName FROM converted_student_data WHERE EmailAddress=%s", (studentEmail,))
        batchSql = cursor.fetchall()
        if batchSql:
            batch = batchSql[0]['batch']
            # studentName = batchSql[0]['FirstName']
        else:
            return jsonify({"status": 0, "message": "Student not found!"})
        
        activeFlag = 2

        checkSql = """
            SELECT COUNT(*) as count FROM assignmentStudentMaster
            WHERE assignmentId = %s AND assignmentName = %s AND email = %s
        """
        cursor.execute(checkSql, (assignmentId, assignmentName, studentEmail))
        count = cursor.fetchone()['count']

        if count > 0:

            resp = jsonify({"status": 2, "success": False, "message": "Assignment already downloaded"})
        else:

            insertSql = """
                INSERT INTO assignmentStudentMaster (assignmentId, assignmentName, submittedDate, email, batch, assignmentPath, activeFlag)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(insertSql, (assignmentId, assignmentName, None, studentEmail, batch, None, activeFlag))

        conn.commit()

        resp = jsonify({"status": 1, "success": True, "message": "Assignment downloaded successfully"})
            
         
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "An error occurred"})

    finally:
        cursor.close()
        conn.close()

    return resp


@student.route('/uploadAssignment', methods=['POST'])
@cross_origin()
def uploadAssignment():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        file = request.files.get('file')
        studentEmail = request.form.get('studentEmail')
        assignmentName = request.form.get('assignmentName')
        assignmentId = request.form.get('assignmentId')
        endDate = request.form.get('endDate')

        cursor.execute("SELECT batch, FirstName FROM converted_student_data WHERE EmailAddress=%s", (studentEmail,))
        batchSql = cursor.fetchall()
        if batchSql:
            batch = batchSql[0]['batch']
            studentName = batchSql[0]['FirstName']
        else:
            return jsonify({"status": 0, "message": "Student not found!"})
        
        endDate = datetime.strptime(endDate, '%Y-%m-%d').date()
        current_timestamp = datetime.now().date()
        submittedDate = current_timestamp
        print('----------->',submittedDate)

        if submittedDate <= endDate:
            print("Hell YEah")
            if file:
                filename, file_extension = os.path.splitext(file.filename)
                new_filename = f"{studentName}_{assignmentName}_{filename}{file_extension}"
                batch_folder = os.path.join(UPLOAD_STUD_ASSIGNMENT_PATH, batch)
                if not os.path.exists(batch_folder):
                    os.makedirs(batch_folder)
                
                file_path = os.path.join(batch_folder, new_filename)
                file.save(file_path)
                assignmentPath = f"studentAssignmentUploads/{batch}/{new_filename}"
            else:
                return jsonify({"status": 0, "message": "No file found!"})

            

            activeFlag = 1

            checkSql = """
                SELECT COUNT(*) as count FROM assignmentStudentMaster
                WHERE assignmentId = %s AND assignmentName = %s AND email = %s
            """
            cursor.execute(checkSql, (assignmentId, assignmentName, studentEmail))
            count = cursor.fetchone()['count']

            if count > 0:

                updateSql = """
                    UPDATE assignmentStudentMaster
                    SET submittedDate = %s, assignmentPath = %s, activeFlag = %s
                    WHERE assignmentId = %s AND assignmentName = %s AND email = %s
                """
                cursor.execute(updateSql, (submittedDate, assignmentPath, activeFlag, assignmentId, assignmentName, studentEmail))
            else:

                insertSql = """
                    INSERT INTO assignmentStudentMaster (assignmentId, assignmentName, submittedDate, email, batch, assignmentPath, activeFlag)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(insertSql, (assignmentId, assignmentName, submittedDate, studentEmail, batch, assignmentPath, activeFlag))

            conn.commit()

            resp = jsonify({"status": 1, "success": True, "message": "Assignment uploaded successfully"})
            
        else:
            resp = jsonify({"status": 2, "success": False, "message": "Assignment upload date is over"})
            
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "An error occurred"})

    finally:
        cursor.close()
        conn.close()

    return resp


# @student.route('/getAssignmentData', methods=['POST'])
# @cross_origin()
# def getAssignmentData():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data['studentEmail']
        
#         cursor.execute("SELECT batch FROM converted_student_data WHERE EmailAddress=%s", (email,))
#         batchSql = cursor.fetchall()
#         if batchSql:
#             batchSql = batchSql[0]['batch']
        
#         print(batchSql)
        
#         data = []
        
#         cursor.execute("SELECT * FROM assignmentMaster WHERE JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s)) = 1", (batchSql,))
#         assignmentSql = cursor.fetchall()
#         cursor.execute("SELECT marks, assignmentId FROM assignmentStudentMaster WHERE email=%s", (email,))
#         uploadedAssignmentIdSql = cursor.fetchall()

#         uploadedAssignmentIds = [str(item['assignmentId']) for item in uploadedAssignmentIdSql]
#         marks_mapping = {str(item['assignmentId']): item['marks'] for item in uploadedAssignmentIdSql}
 
#         for assignment in assignmentSql:
#             assignment_id = str(assignment['id'])
            
#             assignment['marks'] = marks_mapping.get(assignment_id, None)
            
#             assignment['uploadStatus'] = '1' if assignment_id in uploadedAssignmentIds else '0'
            
#             data.append(assignment)

#         resp = jsonify({"status": 1, "result": data})
        
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0, "message": "An error occurred"})
        
#     finally:
#         cursor.close()
#         conn.close()

#     return resp

@student.route('/getAssignmentData', methods=['POST'])
@cross_origin()
def getAssignmentData():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
        # Get the student's batch
        cursor.execute("SELECT batch FROM converted_student_data WHERE EmailAddress=%s", (email,))
        batchSql = cursor.fetchall()
        if batchSql:
            batchSql = batchSql[0]['batch']
        data = []
        # Fetch assignments linked to the student's batch
        cursor.execute("SELECT * FROM assignmentMaster WHERE JSON_CONTAINS(batch, JSON_OBJECT('batchName', %s)) = 1", (batchSql,))
        assignmentSql = cursor.fetchall()
        print(assignmentSql)
        # Fetch assignment details along with activeFlag for uploaded assignments
        cursor.execute("""
            SELECT marks, assignmentId, activeFlag 
            FROM assignmentStudentMaster 
            WHERE email=%s
        """, (email,))
        uploadedAssignmentIdSql = cursor.fetchall()
 
        # Create mapping for uploaded assignments
        uploadedAssignmentIds = [str(item['assignmentId']) for item in uploadedAssignmentIdSql]
        marks_mapping = {str(item['assignmentId']): item['marks'] for item in uploadedAssignmentIdSql}
        active_flag_mapping = {str(item['assignmentId']): item['activeFlag'] for item in uploadedAssignmentIdSql}
        for assignment in assignmentSql:
            assignment_id = str(assignment['id'])
            # Set marks if available
            assignment['marks'] = marks_mapping.get(assignment_id, None)
            # Set uploadStatus based on the activeFlag
            if assignment_id in uploadedAssignmentIds:
                active_flag = active_flag_mapping.get(assignment_id, 0)
                if active_flag == '1':
                    assignment['uploadStatus'] = '1'
                elif active_flag == '2':
                    assignment['uploadStatus'] = '2'
                else:
                    assignment['uploadStatus'] = '0'
            else:
                assignment['uploadStatus'] = '0'
            data.append(assignment)
 
        resp = jsonify({"status": 1, "result": data})
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": "An error occurred"})
    finally:
        cursor.close()
        conn.close()
 
    return resp


@student.route('/getStudentListForPlacement', methods=['POST'])
@cross_origin()
def getStudentListForPlacement():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        batches = data['batchList'] 

        studentList = []

        for batch in batches:
            batchName = batch['batchName']
            sql = """
                SELECT 
                    id, 
                    CONCAT(FirstName, ' ', LastName) AS FullName, 
                    EmailAddress, 
                    batch 
                FROM 
                    converted_student_data 
                WHERE 
                    batch=%s
            """
            cursor.execute(sql, (batchName,))
            studentDetails = cursor.fetchall()
            
            for student in studentDetails:
                studentList.append({
                    "id": student['id'],
                    "fullName": student['FullName'],
                    "email": student['EmailAddress'],
                    "batch": student['batch']
                })

        resp = jsonify({"status": 1, "message": "Success", "result": studentList})
        
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": "An error occurred"})
        
    finally:
        cursor.close()
        conn.close()

    return resp

@student.route('/getStudTests', methods=['POST'])
@cross_origin()
def getStudTests():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
 
        cursor.execute("""SELECT batch FROM converted_student_data WHERE EmailAddress=%s ORDER BY id DESC""", (email,))
        batchName = cursor.fetchall()[0]
 
        cursor.execute("""SELECT * FROM tests ORDER BY id DESC""")
 
        testsData = []
 
        for row in cursor.fetchall():
            if row['batchName']:
                for batches in eval(row['batchName']):
                    if batches:
                        if batchName['batch'] == batches['batchName']:
                            testsData.append(row)
 
        cursor.execute("""SELECT testName FROM studentTestAnswers WHERE email=%s ORDER BY id DESC""", (email,))
        temp = cursor.fetchall()
        filter_tests = [t['testName'] for t in temp]
 
        for item in testsData:
            if item['testName'] in filter_tests:
                item['test_allowed'] = False
            else:
                item['test_allowed'] = True
 
        return jsonify({"status": 1, "success": True, "data": testsData})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})


@student.route('/getCourseListForPlacement', methods=['POST'])
@cross_origin()
def getCourseListForPlacement():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT courseName FROM courseMaster")
        courses = cursor.fetchall()
        
        courseList = [item['courseName'] for item in courses]

        return jsonify({"status": 1, "success": True, "result": courseList})
     
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        cursor.close()
        conn.close()

@student.route('/uploadResumeProfile', methods=['POST'])
@cross_origin()
def uploadResumeProfile():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.form
        file = request.files.get('file')
        email = data.get('studentEmail')
        print('email',email)
        
        
        username = email.split('@')[0]
 
        # studName = ''.join(word.capitalize() for word in re.split(r'[^a-zA-Z]', username) if word)
        studName = f"{username}"

        
        path = f'{UPLOAD_STUD_RESUME_PROFILE_PATH}'
        if not os.path.exists(path):
            os.mkdir(path)
        file.save(f'{path}/{studName}.pdf')
        
        resumePath = f"{studName}.pdf"
        
        resumeSql = "UPDATE studentRegistrationDetails SET resumePath = %s WHERE email = %s"
        cursor.execute(resumeSql, (resumePath, email))
        
        conn.commit()

        return jsonify({"status": 1, "success": True, "message": "Resume uploaded successfully"})
     
    except Exception as e:
        print(e)
        return jsonify({"status": 0, "success": False, "message": "Some error occurs!"})
    finally:
        cursor.close()
        conn.close()

@student.route('/getBatchListForPlacement', methods=['POST'])
@cross_origin()
def getBatchListForPlacement():
    try:
        conn = connect_mysql()
        cursor = conn.cursor() 
        data = request.json

        courseList = data.get('courseList', [])
        if not courseList:
            return jsonify({"status": 0, "error": "courseList is empty or not provided"})

        data = []  
        searchSql = "SELECT * FROM batchMaster WHERE courseType=%s"

        for course in courseList:
            cursor.execute(searchSql, (course,))
            results = cursor.fetchall()

            for row in results:
              
                data.append(row)

        resp = jsonify({"status": 1, "success": True, "result": data, "message": "Fetched successfully"})
        return resp

    except Exception as e:
        return jsonify({"status": 0, "error": str(e)})

    finally:
        cursor.close()
        conn.close()

@student.route('/saveMCQAns', methods=['POST'])
@cross_origin()
def saveMCQans():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        testName = data['testName']
        ques_ans = data['payload']
        marks_obtained = 0
        fname = data['firstName']
        lname = data['lastName']
        batchName = data['batchName']
        testId = data['id']
 
        cursor.execute("""SELECT totalMarks, passingMarks, updQesAndMarks FROM tests WHERE testName=%s""",
                       (testName,))
        temp = cursor.fetchall()[0]
 
        total_marks = temp['totalMarks']
        passingMarks = temp['passingMarks']
        correct_ans = json.loads(temp['updQesAndMarks'])
 
        for ans in correct_ans:
            for user_ans in ques_ans:
                if ans['questionText'] == user_ans['questionText']:
                    if [item['text'] for item in ans['options'] if
                        item['isAnswer'] == 'true' or item['isAnswer'] == True]:
                        print(ans['options'], user_ans['selectedOption'], end='\n\n')
                        if [item['text'] for item in ans['options'] if
                            item['isAnswer'] == 'true' or item['isAnswer'] == True][0].strip() == user_ans[
                            'selectedOption']:
                            marks_obtained += int(ans['marks']) if ans['marks'] else 0
        cursor.execute("""
            INSERT INTO studentTestAnswers (email, testName, answers, testAllowed, marksObtained, percentage, examCleard, name, batchName, testId)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
            email, testName, json.dumps(ques_ans), 'false', marks_obtained, (marks_obtained / int(total_marks)) * 100,
            'Passed' if marks_obtained >= int(passingMarks) else 'Failed', f'{fname} {lname}', json.dumps(batchName), testId))
 
        cursor.execute("""UPDATE studentTestStatus SET status='SUBMITTED' WHERE email=%s AND testName=%s""",
                       (email, testName))
 
        conn.commit()
 
        return jsonify({'status': 1, 'message': 'Test answers saved successfully', 'marksObtained': marks_obtained,
                        'percentage': (marks_obtained / int(total_marks)) * 100,
                        'examStatus': 'Passed' if marks_obtained >= int(passingMarks) else 'Failed'})
    except Exception as e:
        tr = traceback.format_exc()
        print(tr)
        print(e)
        return jsonify({"status": 0})

@student.route('/saveDesPDF', methods=['POST'])
@cross_origin()
def saveDesPDF():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.form
        email = data['email'] or None
        testName = data['testName'] or None
        fname = data['firstName'] or None
        lname = data['lastName'] or None
        ansFile = request.files.get('file') or data['file'] or None
        testId = data['id'] or None
        batchName = data['batchName'] or None

        path = 'ALL_STUDENT_DES_ANSWERS'
        if not os.path.exists(path):
            os.mkdir(path)

        if ansFile:
            file_path = os.path.join(path, f'{fname}-{lname}-{testName}.pdf')
            ansFile.save(file_path)

        cursor.execute("""
            INSERT INTO studentTestAnswers (email, testName, answers, testAllowed, marksObtained, percentage, examCleard, name, testId, batchName)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
            email, testName, f'{fname}-{lname}-{testName}.pdf', 'false', 0, 0, 'Failed',
            f'{fname} {lname}', testId, json.dumps(batchName)))

        cursor.execute("""UPDATE studentTestStatus SET status='SUBMITTED' WHERE email=%s AND testName=%s""",
                       (email, testName))

        conn.commit()

        return jsonify({'status': 1, 'message': 'Test answers saved successfully', 'marksObtained': 0,
                        'percentage': 0,
                        'examStatus': 'Failed'})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
        
@student.route('/studentdownload/<filename>', methods=['GET'])
@cross_origin()
def student_download_file(filename):
    return send_from_directory('ALL_STUDENT_DES_ANSWERS', filename)

@student.route('/download/<filename>', methods=['GET'])
@cross_origin()
def download_file(filename):
    return send_from_directory('ALL_TESTS', filename)

@student.route('/updateTestStatus', methods=['POST'])
@cross_origin()
def updateTestStatus():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email'] or None
        testName = data['testName'] or None

        cursor.execute("""UPDATE studentTestStatus SET status = 'IN PROGRESS' WHERE email=%s AND testName=%s""", (email, testName))
        return jsonify({'status': 1, 'message': 'updated successfully'})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})






@student.route('/uploadProfilePic', methods=['POST'])
@cross_origin()
def uploadProfilePic():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.form
        file = request.files.get('file')
        email = data.get('studentEmail')
        print('email',email)
        
        
        username = email.split('@')[0]
 
 
        studName = f"{username}"
        idStudSQL = "SELECT id from converted_student_data WHERE EmailAddress = %s"
        cursor.execute(idStudSQL, (email))
        idStud = cursor.fetchone()['id']
        
        path = f'{UPLOAD_STUD_PROFILE_PIC}'
        if not os.path.exists(path):
            os.mkdir(path)
        
        
        
    
        filename, file_extension = os.path.splitext(file.filename)
        new_filename = f"{idStud}_{studName}_{filename}{file_extension}"
        file_path = os.path.join(path, new_filename)
        file.save(file_path)
    
        profilePicPath = f"studentProfilePic/{new_filename}"
        
        resumeSql = "UPDATE studentRegistrationDetails SET profilePicture = %s,photoflag=1 WHERE email = %s"
        cursor.execute(resumeSql, (profilePicPath, email))
        
        conn.commit()

        return jsonify({"status": 1, "success": True, "message": "Profile Picture uploaded successfully"})
     
    except Exception as e:
        print(e)
        return jsonify({"status": 0, "success": False, "message": "Some error occurs!"})
    finally:
        cursor.close()
        conn.close()


@student.route('/getProfilePic', methods=['POST'])
@cross_origin()
def get_profile_pic():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
        print("Email ::::::::::::::::::",email)
        # Ensure email is passed as a tuple
        profilePicSql = """
        SELECT profilePicture FROM `studentRegistrationDetails` WHERE email=%s
        """
        cursor.execute(profilePicSql, (email,))  # Pass email as a tuple
        profilePicPath = cursor.fetchone()['profilePicture']
        
        print(profilePicPath, "<---------------------------------------------------------------------")
        
        if profilePicPath:
            # profile_pic_url = profilePicPath[0]  
            return jsonify({
                "status": 1,
                "success": True,
                "message": "Profile Picture path retrieved successfully",
                "profilePicPath": profilePicPath
            })
        else:
            return jsonify({
                "status": 0,
                "success": False,
                "message": "No profile picture found for this email"
            })
    
    except Exception as e:
        print("Error:", e)
        return jsonify({"status": 0, "success": False, "message": "Some error occurred!"})
    
    finally:
        cursor.close()
        conn.close()










# sishu    ==========================================================OCT 1
@student.route('/analyzeSentiment', methods=['POST'])
def analyze_sentiment():
    conn = connect_mysql()
    cursor = conn.cursor()  
    try:
        print('Hello')
        query = "SELECT id, feedbackComment, LeftAt FROM attendanceFeedback"
        cursor.execute(query)
        res = cursor.fetchall()

        valid_feedback = [
            {"id": row['id'], "feedbackComment": row['feedbackComment']}
            for row in res if row['feedbackComment'] and row['feedbackComment'].strip()
        ]

        if not valid_feedback:
            return jsonify({
                "status": "success",
                "message": "No valid comments found for sentiment analysis"
            }), 200

        sentiment_analysis = pipeline("sentiment-analysis")

        df = pd.DataFrame(valid_feedback)

        def get_sentiment(text):
            sentiment = sentiment_analysis(text)
            return sentiment[0]['label']  

        df['Predicted_Sentiment'] = df['feedbackComment'].apply(get_sentiment)

        for _, row in df.iterrows():
            update_query = """
                UPDATE attendanceFeedback 
                SET sentimentAnalyzed = %s 
                WHERE id = %s
            """
            cursor.execute(update_query, (row['Predicted_Sentiment'], row['id']))

        conn.commit()

        sentiment_results = df.to_dict(orient='records')

        return jsonify({
            "status": "success",
            "data": sentiment_results  
        }), 200

    except Exception as e:
        conn.rollback()  
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()




@student.route('/getFeedbackWithSentiment', methods=['GET'])
def get_feedback_with_sentiment():
    conn = connect_mysql()
    cursor = conn.cursor()
      
    try:
        countsql = """
        SELECT COUNT(id) as idCount, COUNT(sentimentAnalyzed) as analyzeCount FROM attendanceFeedback
        """
        cursor.execute(countsql)
        count = cursor.fetchone()
        print(count)
        if count['idCount'] == count['analyzeCount']:
            pass
        else:
            analyze_sentiment()
        
        query = """
            SELECT af.StudentName, af.LeftAt, af.Feedback, af.BatchName, af.mentor, af.MeetingID, 
                   af.feedbackComment, af.sentimentAnalyzed
            FROM attendanceFeedback af
        """
        cursor.execute(query)
        res = cursor.fetchall()

        if not res:
            return jsonify({
                "status": "success",
                "message": "No feedback records found"
            }), 200

        def split_name_and_id(name_with_id):
            if "_" in name_with_id:
                name, user_id = name_with_id.rsplit("_", 1) 
                return name, user_id
            return name_with_id, None  
        
        feedback_data = []
        for row in res:
            student_name, student_user_id = split_name_and_id(row["StudentName"])
            mentor_name, mentor_user_id = split_name_and_id(row["mentor"])

            feedback_item = {
                "StudentName": student_name,
                "studentUserId": student_user_id,
                "LeftAt": row["LeftAt"],
                "Feedback": row["Feedback"],
                "BatchName": row["BatchName"],
                "mentor": mentor_name,
                "mentorUserId": mentor_user_id,
                "MeetingID": row["MeetingID"],
                "feedbackComment": row["feedbackComment"],
                "sentimentAnalyzed": row["sentimentAnalyzed"]
            }

            if row["feedbackComment"] and row["feedbackComment"].strip():
                feedback_item["feedbackComment"] = row["feedbackComment"]
            else:
                feedback_item["feedbackComment"] = "No feedback provided"

            feedback_data.append(feedback_item)


        return jsonify({
            "status": "success",
            "data": feedback_data 
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()        



@student.route('/getTopInstructors', methods=['GET'])
def get_top_instructors():
    conn = connect_mysql()
    cursor = conn.cursor()  # Change to fetch rows as dictionaries

    try:
        # Define the start and end dates of the current financial year
        current_year = datetime.now().year
        current_month = datetime.now().month

        if current_month >= 4:  # April onwards
            start_of_financial_year = f"{current_year}-04-01"
            end_of_financial_year = f"{current_year + 1}-03-31"
        else:  # January to March
            start_of_financial_year = f"{current_year - 1}-04-01"
            end_of_financial_year = f"{current_year}-03-31"

        # Query with the financial year filter
        query = f"""
            SELECT 
                mentor,
                COUNT(*) AS totalFeedbackCount,
                AVG(Feedback) AS averageFeedbackRating,
                SUM(CASE WHEN sentimentAnalyzed = 'POSITIVE' THEN 1 ELSE 0 END) AS positiveFeedbackCount
            FROM 
                attendanceFeedback
            WHERE 
                LeftAt BETWEEN '{start_of_financial_year}' AND '{end_of_financial_year}' -- Filter for the current financial year
            GROUP BY 
                mentor
            ORDER BY 
                positiveFeedbackCount DESC,  -- Sort by positive feedback count first
                averageFeedbackRating DESC;  -- Then by average feedback rating
        """
        
        cursor.execute(query)
        res = cursor.fetchall()

        if not res:
            return jsonify({
                "status": "success",
                "message": "No feedback data found for the current financial year"
            }), 200

        # Function to split name and userId from the format "name_userId"
        def split_name_and_id(name_with_id):
            if "_" in name_with_id:
                name, user_id = name_with_id.rsplit("_", 1)  # Split by the last underscore
                return name, user_id
            return name_with_id, None  # If no underscore, return the name as is and user_id as None

        instructors = []

        # Process each row of feedback data (res as a list of dictionaries)
        for idx, row in enumerate(res, start=1):
            mentor_name, mentor_user_id = split_name_and_id(row['mentor'])
            
            instructor_data = {
                "serialNumber": idx,  # Add serial number to indicate ranking
                "mentorName": mentor_name,
                "mentorUserId": mentor_user_id,
                "totalFeedbackCount": row['totalFeedbackCount'],
                "averageFeedbackRating": "{:.2f}".format(row['averageFeedbackRating']),  # Format to 2 decimal places
                "positiveFeedbackCount": str(row['positiveFeedbackCount'])  # Convert count to string
            }
            
            instructors.append(instructor_data)

        return jsonify({
            "status": "success",
            "data": instructors
        }), 200

    except Exception as e:
        # Log the full exception for debugging purposes
        print("Error in /getTopInstructors API:", str(e))
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        conn.close()


@student.route('/getStudentListOnBatch', methods=['POST'])
@cross_origin()
def getStudentListOnBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        print(req)
        
        # Since the payload is a list of dictionaries, we access the first element of the list
        batch = req[0].get('Batch')  # Extracting batch from the first dictionary in the list
        
        # Handling 'status' if it exists in the payload
        status = req[0].get('status')  # Assuming 'status' could also be part of the first dict, adjust accordingly
        
        studentListSql = """
            SELECT 
                (CASE WHEN LastName is Null THEN FirstName ELSE concat(FirstName,' ',LastName) END) as Name,
                EmailAddress as emailID,
                mx_Course as Course,
                Phone as contactNo,
                status,
                id,
                batch  
            FROM `converted_student_data` 
            WHERE 1 
        """
        
        if batch:
            studentListSql += " AND batch like '" + str(batch) + "'"
        
        if status:
            studentListSql += " AND status like '" + str(status) + "'"
        
        studentListSql += " ORDER BY `converted_student_data`.`id` DESC "
        
        cursor.execute(studentListSql)
        studentResult = cursor.fetchall()

        resp = jsonify({
            "status": 1, 
            "success": True, 
            "result": studentResult, 
            "message": "Fetched successfully"
        })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    
    finally:
        cursor.close()
        return resp



################################ student portfolio ##############################################
@student.route('/getPortfolioDetails', methods=['POST'])
@cross_origin()
def getPortfolioDetails():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
        print("Email ::::::::::::::::::",email)

        portfolioDetailsSql = """
        SELECT * FROM `studentRegistrationDetails` WHERE email=%s
        """
        cursor.execute(portfolioDetailsSql, (email,))  
        portfolioDetails = cursor.fetchall()
        
        
        if portfolioDetails:
            return jsonify({
                "status": 1,
                "success": True,
                "message": "Portfolio Details path retrieved successfully",
                "portfolioDetails": portfolioDetails
            })
        else:
            return jsonify({
                "status": 0,
                "success": False,
                "message": "No Portfolio Details found for this email"
            })
    
    except Exception as e:
        print("Error:", e)
        return jsonify({"status": 0, "success": False, "message": "Some error occurred!"})
    
    finally:
        cursor.close()
        conn.close() 


@student.route('/getEducationPortfolioDetails', methods=['POST'])
@cross_origin()
def getEducationPortfolioDetails():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['studentEmail']
        print("Email ::::::::::::::::::", email)

        portfolioDetailsSql = """
        SELECT 
            doctorsInstituteCgpaGpaType, doctorsInstituteGpaorCgpa, doctorsInstituteName, doctorsYOP,
            mastersInstituteCgpaGpaType, mastersInstituteGpaorCgpa, mastersInstituteName, mastersYOP,
            graduationInstituteCgpaGpaType, graduationInstituteGpaorCgpa, graduationInstituteName, graduationYOP,
            highSchoolInstituteCgpaGpaType, highSchoolInstituteGpaorCgpa, highSchoolInstituteName, highestSchoolYOP
        FROM `studentRegistrationDetails` WHERE email=%s
        """
        cursor.execute(portfolioDetailsSql, (email,))
        portfolioDetails = cursor.fetchone()

        if portfolioDetails:
            result = []

            if all([
                portfolioDetails['doctorsInstituteCgpaGpaType'], 
                portfolioDetails['doctorsInstituteGpaorCgpa'], 
                portfolioDetails['doctorsInstituteName'], 
                portfolioDetails['doctorsYOP']
            ]):
                result.append({
                    "instituteType": "Doctorate",
                    "CgpaGpaType": portfolioDetails['doctorsInstituteCgpaGpaType'],
                    "GpaorCgpa": portfolioDetails['doctorsInstituteGpaorCgpa'],
                    "InstituteName": portfolioDetails['doctorsInstituteName'],
                    "YOP": portfolioDetails['doctorsYOP']
                })
            else:
                print("Doctorate data is incomplete, skipping.")

            if all([
                portfolioDetails['mastersInstituteCgpaGpaType'], 
                portfolioDetails['mastersInstituteGpaorCgpa'], 
                portfolioDetails['mastersInstituteName'], 
                portfolioDetails['mastersYOP']
            ]):
                result.append({
                    "instituteType": "Masters",
                    "CgpaGpaType": portfolioDetails['mastersInstituteCgpaGpaType'],
                    "GpaorCgpa": portfolioDetails['mastersInstituteGpaorCgpa'],
                    "InstituteName": portfolioDetails['mastersInstituteName'],
                    "YOP": portfolioDetails['mastersYOP']
                })
            else:
                print("Masters data is incomplete, skipping.")

            if all([
                portfolioDetails['graduationInstituteCgpaGpaType'], 
                portfolioDetails['graduationInstituteGpaorCgpa'], 
                portfolioDetails['graduationInstituteName'], 
                portfolioDetails['graduationYOP']
            ]):
                result.append({
                    "instituteType": "Graduation",
                    "CgpaGpaType": portfolioDetails['graduationInstituteCgpaGpaType'],
                    "GpaorCgpa": portfolioDetails['graduationInstituteGpaorCgpa'],
                    "InstituteName": portfolioDetails['graduationInstituteName'],
                    "YOP": portfolioDetails['graduationYOP']
                })
            else:
                print("Graduation data is incomplete, skipping.")

            if all([
                portfolioDetails['highSchoolInstituteCgpaGpaType'], 
                portfolioDetails['highSchoolInstituteGpaorCgpa'], 
                portfolioDetails['highSchoolInstituteName'], 
                portfolioDetails['highestSchoolYOP']
            ]):
                result.append({
                    "instituteType": "High School",
                    "CgpaGpaType": portfolioDetails['highSchoolInstituteCgpaGpaType'],
                    "GpaorCgpa": portfolioDetails['highSchoolInstituteGpaorCgpa'],
                    "InstituteName": portfolioDetails['highSchoolInstituteName'],
                    "YOP": portfolioDetails['highestSchoolYOP']
                })
            else:
                print("High School data is incomplete, skipping.")

            if result:
                return jsonify({
                    "status": 1,
                    "success": True,
                    "message": "Portfolio Details retrieved successfully",
                    "portfolioDetails": result
                })
            else:
                return jsonify({
                    "status": 0,
                    "success": False,
                    "message": "No complete Portfolio Details found for this email"
                })
        
        else:
            return jsonify({
                "status": 0,
                "success": False,
                "message": "No Portfolio Details found for this email"
            })

    except Exception as e:
        print("Error:", e)
        return jsonify({"status": 0, "success": False, "message": "Some error occurred!"})
    
    finally:
        cursor.close()
        conn.close()




@student.route('/getCoursesForStudent', methods=['POST'])
@cross_origin()
def getCoursesForStudent():
    try:
        userId = request.json.get('id')  # Assuming studentId is passed in the request JSON

        conn = connect_mysql()
        cursor = conn.cursor()

        sqlForStudentEmailId = """
        SELECT email from user WHERE id=%s
        """
        cursor.execute(sqlForStudentEmailId, (userId,))
        emailId = cursor.fetchone()
        print(emailId)
        
        sqlForStudentId = """
        SELECT id from converted_student_data WHERE EmailAddress=%s
        """
        cursor.execute(sqlForStudentId, (emailId['email'],))
        id = cursor.fetchone()['id']
        
        
        print(id)
        
        sql = """
            SELECT 
                csd.id, 
                csd.batch AS batchName, 
                bm.courseType, 
                bm.activeFlag AS batchActiveFlag, 
                bm.updatedOn,
                cm.courseName,
                cm.folderName,
                cm.activeFlag AS courseActiveFlag
            FROM `converted_student_data` csd
            LEFT JOIN `batchMaster` bm ON csd.batch = bm.batchName
            LEFT JOIN `courseMaster` cm ON bm.courseType = cm.courseName
            WHERE csd.id = %s
            ORDER BY 
                CASE bm.activeFlag 
                    WHEN '2' THEN 0 
                    WHEN '1' THEN 1 
                    WHEN '0' THEN 2 
                    ELSE 3
                END,
                bm.updatedOn DESC
        """
        cursor.execute(sql, (id,))
        batchList = cursor.fetchall()
        
        processed_batches = []

        for batch in batchList:
            # Process the batch details
            batch_id = batch['id']
            batch_name = batch['batchName']
            course_type = batch['courseType']
            batch_active_flag = batch['batchActiveFlag']
            updated_on = batch['updatedOn']
            course_name = batch['courseName'] if batch['courseName'] else ''
            folder_name = batch['folderName'] if batch['folderName'] else ''
            course_active_flag = batch['courseActiveFlag'] if batch['courseActiveFlag'] else ''

            # Handle None for updatedOn safely
            if updated_on:
                updated_on_str = updated_on.strftime('%Y-%m-%d')
            else:
                updated_on_str = 'N/A'  # Use 'N/A' or another placeholder for None values

            # Process folderName correctly as you want
            folder_file_dict = {}
            if folder_name:
                parts = folder_name.rsplit('/', 1)  # Split only at the last '/'
                if len(parts) == 2:
                    folder_file_dict = {
                        'path': parts[0].strip(),  # The folder path part
                        'file': parts[1].strip()   # The file part
                    }

            # Construct batch details dictionary
            batch_details = {
                'id': batch_id,
                'batchName': batch_name,
                'courseType': course_type,
                'courseName': course_name,
                'folderName': folder_file_dict,  # Updated folderName to be a dictionary
                'courseActiveFlag': course_active_flag,
                'batchActiveFlag': batch_active_flag,
                'updatedOn': updated_on_str  # Safely assign formatted or placeholder string
            }
            processed_batches.append(batch_details)

        # After processing batches, prepare response
        response = {
            'success': True,
            'message': 'Batches fetched successfully.',
            'result': processed_batches
        }
        print(processed_batches)
        return jsonify(response)

    except Exception as e:
        print(str(e))  # Print or log the exception for debugging purposes
        response = {
            'success': False,
            'message': 'Failed to fetch batches. Please try again later.'
        }
        return jsonify(response)




#21 OCT 2024

# @student.route('/getCoursesModuleForStudent', methods=['POST'])
# @cross_origin()
# def getCoursesModuleForStudent():
#     try:
#         userId = request.json.get('id')  
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         sqlForStudentEmailId = """
#         SELECT email FROM user WHERE id=%s
#         """
#         cursor.execute(sqlForStudentEmailId, (userId,))
#         emailId = cursor.fetchone()
       
#         sqlForStudentId = """
#         SELECT mx_Course, batch, id FROM converted_student_data WHERE EmailAddress=%s
#         """
#         cursor.execute(sqlForStudentId, (emailId['email'],))
#         courseName = cursor.fetchone()
        
#         sqlForCourseCode = """
#         SELECT DISTINCT(code) FROM courseMaster WHERE courseName=%s
#         """
#         cursor.execute(sqlForCourseCode, (courseName['mx_Course'],))
#         courseCode = cursor.fetchone()['code']
        
#         sqlForCourseContent = """
#         SELECT * FROM courseContentMaster WHERE courseId=%s
#         """
#         cursor.execute(sqlForCourseContent, (courseCode,))
#         courseContent = cursor.fetchall()

#         grouped_content = {}
#         for content in courseContent:
#             content_name = content['contentName']
#             if content_name not in grouped_content:
#                 grouped_content[content_name] = []
#             grouped_content[content_name].append(content)

#         for content_name in grouped_content:
#             grouped_content[content_name].sort(key=lambda x: x['id'])

#         response_content = [
#             {
#                 'contentName': content_name,
#                 'modules': grouped_content[content_name]
#             }
#             for content_name in grouped_content
#         ]

#         response = {
#             'success': True,
#             'message': 'Batches fetched successfully.',
#             'result': response_content
#         }
        
#         return jsonify(response)

#     except Exception as e:
#         print(str(e))  
#         response = {
#             'success': False,
#             'message': 'Failed to fetch batches. Please try again later.'
#         }
#         return jsonify(response)







@student.route('/getCoursesModuleForStudent', methods=['POST'])
@cross_origin()
def getCoursesModuleForStudent():
    try:
        userId = request.json.get('id')
        conn = connect_mysql()
        cursor = conn.cursor()

        sqlForStudentEmailId = """
        SELECT email FROM user WHERE id=%s
        """
        cursor.execute(sqlForStudentEmailId, (userId,))
        emailId = cursor.fetchone()

        sqlForStudentId = """
        SELECT mx_Course, batch, id FROM converted_student_data WHERE EmailAddress=%s
        """
        cursor.execute(sqlForStudentId, (emailId['email'],))
        student_data = cursor.fetchone()

        courseName = student_data['mx_Course']
        batchName = student_data['batch']  

        sqlForCourseCode = """
        SELECT DISTINCT(code) FROM courseMaster WHERE courseName=%s
        """
        cursor.execute(sqlForCourseCode, (courseName,))
        courseCode = cursor.fetchone()['code']

        sqlForCourseContent = """
        SELECT * FROM courseContentMaster WHERE courseId=%s
        """
        cursor.execute(sqlForCourseContent, (courseCode,))
        courseContent = cursor.fetchall()

        sqlForMeetings = """
        SELECT recordedMeeting 
        FROM meetings 
        WHERE courseId=%s 
        AND JSON_CONTAINS(batch, %s) 
        ORDER BY id ASC
        """
        cursor.execute(sqlForMeetings, (courseCode, json.dumps(batchName)))
        recordedMeetings = cursor.fetchall()  

        grouped_content = {}
        for content in courseContent:
            content_name = content['contentName']
            if content_name not in grouped_content:
                grouped_content[content_name] = []
            grouped_content[content_name].append(content)

        for content_name in grouped_content:
            grouped_content[content_name].sort(key=lambda x: x['id'])


        meeting_index = 0  
        response_content = []

        for content_name in grouped_content:
            modules = grouped_content[content_name]

            for module in modules:
                if meeting_index < len(recordedMeetings):
                    module['recordedMeeting'] = recordedMeetings[meeting_index]['recordedMeeting']
                    meeting_index += 1  
                else:
                    module['recordedMeeting'] = None  

            response_content.append({
                'contentName': content_name,
                'modules': modules
            })

        response = {
            'success': True,
            'message': 'Batches and meetings fetched successfully.',
            'result': response_content
        }

        return jsonify(response)

    except Exception as e:
        print(str(e))  
        response = {
            'success': False,
            'message': 'Failed to fetch batches and meetings. Please try again later.'
        }
        return jsonify(response)