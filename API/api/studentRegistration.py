from flask import Blueprint
from flask.json import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
import math, random
from api.sendMail import * 
import os

from docxtpl import DocxTemplate
from docxtpl import InlineImage
from docx.shared import Mm
from .user import addUser


engine = create_engine('mysql+pymysql://root:%s@localhost/lms' % quote_plus('password'))


reg = Blueprint('reg', __name__)


def generateOTP() :
 
    
    digits = "0123456789"
    OTP = ""
 
   
    for i in range(6) :
        OTP += digits[math.floor(random.random() * 10)]
 
    return OTP

def generate_pdf(doc_path, path):
    print("in common.py")
    subprocess.call(['soffice',
                 # '--headless',
                 '--convert-to',
                 'pdf',
                 '--outdir',
                 path,
                 doc_path])
    return doc_path



@reg.route('/getMentorList',methods=['POST'])
@cross_origin()
def getMentorList():
    try:
        req = request.json
        userId = req['userId']
        
        conn = connect_mysql()
        cursor = conn.cursor()

        mentorListSql = " "
    
    except Exception as e:
        print(e)
        return jsonify({"status":0})
    finally:
        cursor.close()



@reg.route('/getBatchList',methods=['GET'])
@cross_origin()
def getBatchList():
    try:
        
        
        conn = connect_mysql()
        cursor = conn.cursor()

        batchListSql = " SELECT * FROM `batchMaster` "
        cursor.execute(batchListSql)
        batchResult = cursor.fetchall()
        resp =  jsonify({"status":1, "success":True, "batchList": batchResult, "message": "Logged in successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp





@reg.route('/getCountryList',methods=['POST'])
@cross_origin()
def getCountryList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json

        countryPhrase = req['countryPhrase']

        if countryPhrase != '' and countryPhrase is not None:
            sql = " SELECT DISTINCT(name),id FROM `countries` WHERE name LIKE '"+str(countryPhrase)+"%'"
        else: 
            sql = " SELECT DISTINCT(name),id FROM `countries` "

        sql += " order by name"
        
        cursor.execute(sql)
        result =cursor.fetchall()

        resp =  jsonify({"status":1, "success":True, "countryList":result,  "message": "Fetched successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    

@reg.route('/getStateList',methods=['POST'])
@cross_origin()
def getStateList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        country = req['country']

        if country != [] and country is not None :
            sql = "SELECT * FROM `states` WHERE country_id  LIKE '"+str(country[0]['id'])+"'"

        else:
            sql = "SELECT * FROM `states`"
        
        sql += " order by name"
        
        cursor.execute(sql)
        result =cursor.fetchall()

        resp =  jsonify({"status":1, "success":True, "stateList":result,  "message": "Fetched successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    

@reg.route('/getCityList',methods=['POST'])
@cross_origin()
def getCityList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        state = req['state']

        if state != [] and state is not None:
            sql = "SELECT * FROM `city` WHERE state_id  LIKE '"+str(state[0]['id'])+"'"
        else : 
            sql = "select * from city "
        
        sql += " order by name"
        
        cursor.execute(sql)
        result =cursor.fetchall()

        resp =  jsonify({"status":1, "success":True, "cityList":result,  "message": "Fetched successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    


@reg.route('/getCountryCode',methods=['POST'])
@cross_origin()
def getCountryCode():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        country = req['country']

        if country != [] and country is not None :
            sql = "SELECT * FROM `countries` WHERE id  LIKE '"+str(country[0]['id'])+"'"

        else:
            sql = "SELECT * FROM `countries`"
        
        sql += " order by name"
        
        cursor.execute(sql)
        result =cursor.fetchone()

        resp =  jsonify({"status":1, "success":True, "countryCode":result['phone_code'],  "message": "Fetched successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp


@reg.route('/sendOtp',methods=['POST'])
@cross_origin()
def sendOtp():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        req =request.json
        print(req)
        emailId = req['emailId']

        emailSearchSql = "SELECT * FROM `converted_student_data`  WHERE EmailAddress like '"+str(emailId)+"' "
        cursor.execute(emailSearchSql)
        res = cursor.fetchone()
        emailIdList =[]
        if res != None and len(res) != 0:
            if res['submitFlag'] == 0:
                # print(res)
                otp = generateOTP()
                print(otp)
                otpUpdateSql = "  UPDATE `converted_student_data` SET `otp` = '"+str(otp)+"' WHERE EmailAddress LIKE  '"+str(emailId)+"'"
                cursor.execute(otpUpdateSql)
                textBody = ""
                subject = "OTP for Email Verification"
                htmlText = '''<span>Hi '''+str(res['FirstName'])+'''!<br> Your verification code is <span style ="font-weight:bold;background : yellow "> '''+str(otp)+'''</span>.<br> Enter this code in our website to proceed your registration process. Please do not share your OTP with anyone.   </span>'''
                emailIdList.append(emailId)
                obj = sendMail(emailIdList,textBody,subject,htmlText,'')
                obj.sendmail()
                resp = jsonify({"status":1, "success":True,   "message": "OTP Sent Successfully","submitFlag" : res['submitFlag'] })
            else:
               resp = jsonify({"status":1, "success":True,   "message": "Already Submitted.","submitFlag" : res['submitFlag'] }) 
        
        else:
            resp =  jsonify({"status":0, "success":False,   "message": "Email not present in crm list", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    



@reg.route('/verifyOtp',methods=['POST'])
@cross_origin()
def verifyOtp():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        req =request.json
        print(req)
        emailId = req['emailId']
        otp = req['otp']

        verifyOtpSql = "SELECT * FROM `converted_student_data`  WHERE EmailAddress like '"+str(emailId)+"' AND otp like "+str(otp)
        cursor.execute(verifyOtpSql)
        res = cursor.fetchone()
        if res != None and len(res) != 0:
            print(res)
            

            resp = jsonify({"status":1, "success":True,   "message": "Log in Succesfuly","submitFlag" : res['submitFlag']  })
        
        else:
            resp =  jsonify({"status":0, "success":False,   "message": "Please enter valid OTP", })







    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    






@reg.route('/insertBasicDetails',methods=['POST'])
@cross_origin()
def insertBasicDetails():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        req = request.json
        req = req['info']
        firstName = req['firstName']
        middleName = req['middleName']
        lastName = req['lastName']
        gender = req['gender']
        dob = req['dob']
        mobileNumber = req['mobileNumber']
        alternateNumber = req['alternateNumber']
        email = req['email']
        gurdianName = req['gurdianName']
        IdproofType = req['IdproofType']
        idproofNumber = req['idproofNumber']

        # Hardcoded value for profilePicture
        profilePicture = "fake"

        # Directly store JSON strings for country, state, and city
        country = req['country']
        state = req['state']
        city = req['city']

        zipCode = req['zipCode']
        fullAddress1 = req['fullAddress1']
        fullAddress2 = req['fullAddress2']

        # First, check if profile picture is already in the database
        select_query = """
            SELECT profilePicture FROM studentRegistrationDetails WHERE email = %s
        """
        cursor.execute(select_query, (email,))
        result = cursor.fetchone()
        print("resultttttt ",result)

        # If profile picture is already in the database and not "fake", retain it
        if result and result["profilePicture"] != "fake":
            profilePicture = result["profilePicture"]  # Keep the existing profile picture

        # Insert or update student details
        insertStudentRegistrationList = (
            """INSERT INTO studentRegistrationDetails 
                (profilePicture, firstName, middleName, lastName, gender, dob, mobileNumber, alternateNumber, email, 
                gurdianName, country, state, city, IdproofType, idproofNumber, zipCode, fullAddress1, fullAddress2) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
                ON DUPLICATE KEY UPDATE 
                profilePicture = VALUES(profilePicture),
                firstName = VALUES(firstName), 
                middleName = VALUES(middleName), 
                lastName = VALUES(lastName), 
                gender = VALUES(gender), 
                dob = VALUES(dob), 
                mobileNumber = VALUES(mobileNumber), 
                alternateNumber = VALUES(alternateNumber), 
                email = VALUES(email), 
                gurdianName = VALUES(gurdianName), 
                country = VALUES(country), 
                state = VALUES(state), 
                city = VALUES(city), 
                IdproofType = VALUES(IdproofType), 
                idproofNumber = VALUES(idproofNumber), 
                zipCode = VALUES(zipCode), 
                fullAddress1 = VALUES(fullAddress1), 
                fullAddress2 = VALUES(fullAddress2);
            """
        )

        # Execute the query with the profilePicture value decided based on the above check
        cursor.execute(insertStudentRegistrationList, (
            profilePicture, firstName, middleName, lastName, gender, dob, mobileNumber, alternateNumber, email, 
            gurdianName, country, state, city, IdproofType, idproofNumber, zipCode, fullAddress1, fullAddress2
        ))

        conn.commit()

        return jsonify({"Status": "Inserted in the database successfully!"})

    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        cursor.close()
        conn.close()




# @reg.route('/insertBasicDetails',methods=['POST'])
# @cross_origin()
# def insertBasicDetails():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         req = request.json
#         req = req['info']
#         firstName = req['firstName']
#         middleName = req['middleName']
#         lastName = req['lastName']
#         gender = req['gender']
#         dob = req['dob']
#         mobileNumber = req['mobileNumber']
#         alternateNumber = req['alternateNumber']
#         email = req['email']
#         gurdianName = req['gurdianName']
#         IdproofType = req['IdproofType']
#         idproofNumber = req['idproofNumber']
#         profilePicture = "pp"

#         # Check if country, state, and city are JSON strings or regular strings
#         def get_name(value):
#             try:
#                 value_json = json.loads(value)
#                 return value_json['name']
#             except (json.JSONDecodeError, TypeError, KeyError):
#                 return value

#         country = get_name(req['country'])
#         state = get_name(req['state'])
#         city = get_name(req['city'])

#         zipCode = req['zipCode']
#         fullAddress1 = req['fullAddress1']
#         fullAddress2 = req['fullAddress2']

#         insertStudentRegistrationList = (
#             """INSERT INTO studentRegistrationDetails 
#                 (profilePicture,firstName, middleName, lastName, gender, dob, mobileNumber, alternateNumber, email, 
#                 gurdianName, country, state, city, IdproofType, idproofNumber, zipCode, fullAddress1, fullAddress2) 
#                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s) 
#                 ON DUPLICATE KEY UPDATE 
#                 profilePicture = VALUES(profilePicture),
#                 firstName = VALUES(firstName), 
#                 middleName = VALUES(middleName), 
#                 lastName = VALUES(lastName), 
#                 gender = VALUES(gender), 
#                 dob = VALUES(dob), 
#                 mobileNumber = VALUES(mobileNumber), 
#                 alternateNumber = VALUES(alternateNumber), 
#                 email = VALUES(email), 
#                 gurdianName = VALUES(gurdianName), 
#                 country = VALUES(country), 
#                 state = VALUES(state), 
#                 city = VALUES(city), 
#                 IdproofType = VALUES(IdproofType), 
#                 idproofNumber = VALUES(idproofNumber), 
#                 zipCode = VALUES(zipCode), 
#                 fullAddress1 = VALUES(fullAddress1), 
#                 fullAddress2 = VALUES(fullAddress2);
#                 """)

#         cursor.execute(insertStudentRegistrationList, (
#             firstName, middleName, lastName, gender, dob, mobileNumber, alternateNumber, email, gurdianName, country, state,
#             city, IdproofType, idproofNumber, zipCode, fullAddress1, fullAddress2
#         ))
#         conn.commit()

#         return jsonify({"Status": "Inserted in the database successfully!"})

#     except Exception as e:
#         print(e)
#         return jsonify({"status": 0})
#     finally:
#         cursor.close()
#         conn.close()


# def insertBasicDetails():
#     try:
#         req =request.json
#         print("1111111111111",req)
#         crmId = req['crmId']
#         firstName = req['firstName']
#         middleName = req['middleName']
#         lastName = req['lastName']
#         gender = req['gender']
#         dob = req['dob']
#         mobileNumber = req['mobileNo']
#         alternateNumber = req['alternateMobileNo']
#         relation = req['relation']
#         email = req['email']
#         idNo = req['idNo']
#         fileType = req ['idProofType']
#         if fileType == 'aadhar':
#             aadharNumber = idNo
#             panNumber = None
#         else :
#             aadharNumber= None
#             panNumber = idNo
#         if req['country'] != []:

#             country = req['country'][0]['name']
#             countryId = req['country'][0]['id']
#         else:
#             country = ''
#             countryId = None

#         if req['state'] !=[]:
#             state = req['state'][0]['name']
#             stateId = req['state'][0]['id']
#         else: 
#             state = ''
#             stateId = None
#         if req['city'] !=[]:

#             city = req['city'][0]['name']
#             cityId = req['city'][0]['id']
#         else:
#             city = ''
#             cityId = None
#         zipCode = req['zipCode']
#         fullAddress1 = req['fullAddress1']
#         fullAddress2 = req['fullAddress2']
#         countryCode = req['countryCode']

#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # searchSql = ""

#         insertStudentRegistrationList = " INSERT INTO `studentRegistrationDetails`(  `firstName`, `middleName`, `lastName`, `gurdianName`, `gender`, `dob`, `mobileNumber`, `alternateNumber`, `email`, `country`,countryId,countryCode,stateId, `state`, cityId, `city`, `zipCode`, `fullAddress1`,fullAddress2,aadharNumber,panNumber) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

#         updateSql = "UPDATE `studentRegistrationDetails` SET `firstName`=%s,`middleName`= %s,`lastName`= %s,`gurdianName`=%s,`gender`= %s,`dob`= %s,`mobileNumber`= %s,`alternateNumber`= %s,`country`= %s,countryId = %s, countryCode = %s,stateId =%s,`state`= %s,cityId = %s,`city`= %s ,`zipCode`= %s,`fullAddress1`= %s,`fullAddress2`= %s,aadharNumber = %s,panNumber = %s  where  email like '"+str(email)+"' "

#         searchSql = " SELECT *  FROM `studentRegistrationDetails` WHERE `email` = '"+str(email) + "'"
#         cursor.execute(searchSql)
#         count  =cursor.fetchall()

#         if count is None or len(count) == 0 :
#             data = (firstName,middleName,lastName,relation,gender,dob,mobileNumber,alternateNumber,email,country,countryId,countryCode,stateId,state,cityId,city,zipCode,fullAddress1,fullAddress2,aadharNumber,panNumber)
#             print(data)
#             cursor.execute (insertStudentRegistrationList,data)

#             resp =  jsonify({"status":1, "success":True,  "message": "Insert successfully", })
#         else:
#             data = (firstName,middleName,lastName,relation,gender,dob,mobileNumber,alternateNumber,country,countryId,countryCode,stateId,state,cityId,city,zipCode,fullAddress1,fullAddress2,aadharNumber,panNumber)
#             cursor.execute(updateSql,data)

#             resp = jsonify({"status":1, "success":True,  "message": "Updated Successfully",})



#     except Exception as e:
#         print(e)
#         resp = jsonify({"status":0})
#     finally:
#         cursor.close()
#         return resp


        





@reg.route('/fetchBasicInfo',methods=['POST'])
@cross_origin()
def fetchBasicInfo():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        print(datetime.now())
        req = request.json
        print(req)
        email = req['email']
        
        cursor.execute("""SELECT * FROM studentRegistrationDetails WHERE email=%s""",(email))
        data = cursor.fetchone()
        print(data)
        response = data
        if response is not None and len(response) != 0:
            response["folderName"] = email.split('@')[0]
            print(response) 
            return jsonify(response)
        else : 
            resp = jsonify({"status": 1, "success": True, "dataPresent": 0})
            return resp
        # searchSql = " SELECT crmId,firstName,middleName,lastName,gurdianName,gender,dob,fullAddress1,fullAddress2,country,countryCode,countryId,stateId,state,cityId,city,zipCode,mobileNumber,alternateNumber,email,aadharNumber,panNumber,photoFlag,idProofFlag,batch  FROM `studentRegistrationDetails` WHERE `email` = '" + str(
        #     email) + "'"
        # cursor.execute(searchSql)
        # result = cursor.fetchone()

        # if result != None and len(result) != 0:
        #     if result['aadharNumber'] is None:
        #         idProofType = 'pan'
        #     else:
        #         idProofType = 'aadhar'
        #     for each in result.keys():
        #         if result[each] is None:
        #             result[each] = ''
        #     folder = email.split('@')[0]
        #     resp = jsonify({"status": 1, "success": True, "basicDetails": result, "message": "Fetched successfully",
        #                     "dataPresent": 1, 'idProofType': idProofType, "folderName": folder})
        # else:
        #     resp = jsonify({"status": 1, "success": True, "dataPresent": 0})


    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        cursor.close()
        print("end", datetime.now())
 
# def fetchBasicInfo():
#     try:
#         print(datetime.now())
#         req =request.json
#         print(req)
#         conn = connect_mysql()
#         cursor = conn.cursor()
#         email = req['email']

#         searchSql = " SELECT crmId,firstName,middleName,lastName,gurdianName,gender,dob,fullAddress1,fullAddress2,country,countryCode,countryId,stateId,state,cityId,city,zipCode,mobileNumber,alternateNumber,email,aadharNumber,panNumber,photoFlag,idProofFlag,batch  FROM `studentRegistrationDetails` WHERE `email` = '"+str(email) + "'"
#         cursor.execute(searchSql)
#         result  =cursor.fetchone()

#         if result != None and len(result) !=0 :
#             if result['aadharNumber'] is None:
#                 idProofType = 'pan'
#             else : 
#                 idProofType = 'aadhar'
#             for each in result.keys():
#                 if result[each] is None:
#                     result[each] = ''
#             folder = email.split ('@')[0]
#             resp = jsonify({"status":1, "success":True, "basicDetails":result,  "message": "Fetched successfully","dataPresent":1,'idProofType':idProofType,"folderName" : folder })
#         else:
#             resp = jsonify({"status":1, "success":True, "dataPresent":0 })


#     except Exception as e:
#         print(e)
#         resp = jsonify({"status":0})
#     finally:
#         cursor.close()
#         print("end",datetime.now())
#         return resp
    
@reg.route('/uploadDocument',methods = ['POST'])
@cross_origin()
def uploadDocument():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        # req =request.json
        # print(req)
        # print(request.form)
        # emailid = request.form.get('email')
        # email = emailid.split('@')[0]
        # fileType = request.form.get('fileType')
        # print("fileType-----", fileType)
        # file = request.files.get('file')
        # print("file", file)
        # path = UPLOAD_DOC_PATH + str(email)
        # print("path-----", path)
        # isExist = os.path.exists(path)
        # print(isExist)
        # print(path)
        # if isExist is False:
        #     os.mkdir(path)
        #
        # if fileType in ('pan', 'aadhar'):
        #     fileType = 'idProof'
        # if file.filename.split('.')[1] in ('jpeg', 'jpg'):
        #     fileName = fileType + ".png"
        # else:
        #     fileName = fileType + "." + file.filename.split('.')[1]
        # print("---filename---", fileName)
        # file.save(path + "/" + fileName)
        #
        # updateFlag = " UPDATE `studentRegistrationDetails` SET " + str(
        #     fileType) + "Flag = '1' WHERE email like '" + str(emailid) + "'"
        # print(updateFlag)

        # cursor.execute(updateFlag)

        # doctorate
        # Masters
        # Bachelor
        # Associate
        # Graduation
        # high_school

        reg = request.form
        email = reg.get('email')
        fileType = reg.get('fileType')
        file = request.files.get('file')
        temp = None

        path = UPLOAD_DOC_PATH + str(email)
        isExist = os.path.exists(path)

        if isExist is False:
            os.mkdir(path)

        if fileType == 'doctorate':
            fileType = 'DOCTORATE'
        elif fileType == 'Masters':
            fileType = 'MASTERS'
        elif fileType == 'Bachelor':
            fileType = 'BACHELOR'
        elif fileType == 'Associate':
            fileType = 'ASSOCIATE'
        elif fileType == 'Graduation':
            fileType = 'GRADUATION'
        elif fileType == 'high_school':
            fileType = 'high_school'
        elif fileType == 'professional_certificate':
            fileType = 'PROFESSIONAL_CERTIFICATE'    
        elif fileType in ['aadhar', 'Passport', 'Drivers_License', 'oth_valid_gov_id', 'pan','photo','Graduation',]:
            temp = fileType
            fileType = 'idProofFlag'    

        if file.filename.split('.')[1] in ('jpeg', 'jpg'):
            fileName = fileType + ".png"
        else:
            fileName = fileType + "." + file.filename.split('.')[1]

        file.save(path + "/" + fileName)
        if fileType != 'idProofFlag':
            updateFlag = f"UPDATE studentRegistrationDetails SET {fileType} = 1 WHERE email LIKE '{email}'"
        else:
            updateFlag = f"UPDATE studentRegistrationDetails SET {fileType} = {temp} WHERE email LIKE '{email}'"

        cursor.execute(updateFlag)

        resp = jsonify({"status": 1, "success": True, "message": "Upload Successfully", "dataPresent": 1})
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        return resp
      

@reg.route('/saveEducationDetails', methods=['POST'])
@cross_origin()
def saveEducationDetails():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        print("start", datetime.now())
        req = request.json
        print(req)
        email = req['email']
        # highestQualification = req['highestQualification']
        # passOutYearPost = req['pgYear']
        # print(passOutYearPost)
        # if passOutYearPost == '':
        #     print("11111",passOutYearPost)
        #     passOutYearPost = None
        # univercityPost =  req['pgCollege']
        # marksPost  = req['pgCGPA']
        # if marksPost == '':
        #     marksPost = None
        # passOutYearGrad = req['gYear']
        # if passOutYearGrad == '':
        #     passOutYearGrad = None
        # universityGrad =  req['gCollege']
        # marksGrad = req['gCGPA']
        # if marksGrad == '':
        #     marksGrad = None
        # passOutYear12 = req['c12Year']
        # if passOutYear12 == '':
        #     passOutYear12 = None
        # college12 = req['c12College']
        # marks12 =  req['c12CGPA']
        # if marks12 == '':
        #     marks12 = None
        # passOutYear10 = req['c10Year']
        # if passOutYear10 == '':
        #     passOutYear10 = None
        # school10 = req['c10College']
        # marks10 =  req['c10CGPA']
        # if marks10 == '':
        #     marks10 = None
        # crmId = req['crmId']

        highestSchoolYOPvar = None
        highSchoolInstituteGpaorCgpavar = None
        graduationYOPvar = None
        graduationInstituteGpaorCgpavar = None
        associateYOPvar = None
        associateInstituteGpaorCgpavar = None
        bachelorYOPvar = None
        bachelorInstituteGpaorCgpavar = None
        mastersYOPvar = None
        mastersInstituteGpaorCgpavar = None
        doctorsYOPvar = None
        doctorsInstituteGpaorCgpavar = None
        professionalCertificateName = None
        NoOfYearsOfFieldExperience = None

        if req['formData']['highestSchoolYOP'] != '':
            highestSchoolYOPvar = req['formData']['highestSchoolYOP']
        highSchoolInstituteName = req['formData']['highSchoolInstituteName']
        if highSchoolInstituteName == '':
            highSchoolInstituteName = None
        highSchoolInstituteCgpaGpaType = req['formData']['highSchoolInstituteCgpaGpaType']
        if highSchoolInstituteCgpaGpaType == '':
            highSchoolInstituteCgpaGpaType = None
        if req['formData']['highSchoolInstituteGpaorCgpa'] != '':
            highSchoolInstituteGpaorCgpavar = req['formData']['highSchoolInstituteGpaorCgpa']

        if req['formData']['graduationYOP'] != '':
            graduationYOPvar = req['formData']['graduationYOP']
        graduationInstituteName = req['formData']['graduationInstituteName']
        if graduationInstituteName == '':
            graduationInstituteName = None
        graduationInstituteCgpaGpaType = req['formData']['graduationInstituteCgpaGpaType']
        if graduationInstituteCgpaGpaType == '':
            graduationInstituteCgpaGpaType = None
        if req['formData']['graduationInstituteGpaorCgpa'] != '':
            graduationInstituteGpaorCgpavar = req['formData']['graduationInstituteGpaorCgpa']

        if req['formData']['associateYOP'] != '':
            associateYOPvar = req['formData']['associateYOP']
        associateInstituteName = req['formData']['associateInstituteName']
        if associateInstituteName == '':
            associateInstituteName = None
        associateInstituteCgpaGpaType = req['formData']['associateInstituteCgpaGpaType']
        if associateInstituteCgpaGpaType == '':
            associateInstituteCgpaGpaType = None
        if req['formData']['associateInstituteGpaorCgpa'] != '':
            associateInstituteGpaorCgpavar = req['formData']['associateInstituteGpaorCgpa']

        if req['formData']['bachelorYOP'] != '':
            bachelorYOPvar = req['formData']['bachelorYOP']
        bachelorInstituteName = req['formData']['bachelorInstituteName']
        if bachelorInstituteName == '':
            bachelorInstituteName = None
        bachelorInstituteCgpaGpaType = req['formData']['bachelorInstituteCgpaGpaType']
        if bachelorInstituteCgpaGpaType == '':
            bachelorInstituteCgpaGpaType = None
        if req['formData']['bachelorInstituteGpaorCgpa'] != '':
            bachelorInstituteGpaorCgpavar = req['formData']['bachelorInstituteGpaorCgpa']

        if req['formData']['mastersYOP'] != '':
            mastersYOPvar = req['formData']['mastersYOP']
        mastersInstituteName = req['formData']['mastersInstituteName']
        if mastersInstituteName == '':
            mastersInstituteName = None
        mastersInstituteCgpaGpaType = req['formData']['mastersInstituteCgpaGpaType']
        if mastersInstituteCgpaGpaType == '':
            mastersInstituteCgpaGpaType = None
        if req['formData']['mastersInstituteGpaorCgpa'] != '':
            mastersInstituteGpaorCgpavar = req['formData']['mastersInstituteGpaorCgpa']

        if req['formData']['doctorsYOP'] != '':
            doctorsYOPvar = req['formData']['doctorsYOP']
        doctorsInstituteName = req['formData']['doctorsInstituteName']
        if doctorsInstituteName == '':
            doctorsInstituteName = None
        doctorsInstituteCgpaGpaType = req['formData']['doctorsInstituteCgpaGpaType']
        if doctorsInstituteCgpaGpaType == '':
            doctorsInstituteCgpaGpaType = None
        if req['formData']['doctorsInstituteGpaorCgpa'] != '':
            doctorsInstituteGpaorCgpavar = req['formData']['doctorsInstituteGpaorCgpa']

        if req['formData']['professionalCertificateName'] != '':
            professionalCertificateName = req['formData']['professionalCertificateName']

        if req['formData']['NoOfYearsOfFieldExperience'] != '':
            NoOfYearsOfFieldExperience = req['formData']['NoOfYearsOfFieldExperience']

        # updateSql = "UPDATE `studentRegistrationDetails` SET `highestQualification`  = %s, `passOutYearPost` = %s,`univercityPost` = %s, `marksPost`= %s,`passOutYearGrad` = %s,`universityGrad`= %s,`marksGrad`= %s,`passOutYear12` = %s,`college12`= %s,`marks12`= %s,`passOutYear10`= %s,`school10`= %s,`marks10`= %s WHERE email like '"+str(email)+"' "
        # data=(highestQualification,passOutYearPost,univercityPost, marksPost,passOutYearGrad,universityGrad,marksGrad,passOutYear12,college12,marks12,passOutYear10,school10,marks10)

        # print(highestSchoolYOP,
        #       graduationYOP,
        #       associateYOP,
        #       bachelorYOP,
        #       mastersYOP,
        #       doctorsYOP,
        #
        #       type(highestSchoolYOP),
        #       type(graduationYOP),
        #       type(associateYOP),
        #       type(bachelorYOP),
        #       type(mastersYOP),
        #       type(doctorsYOP)
        #       )

        # updateSql = ("UPDATE `studentRegistrationDetails` SET `highestSchoolYOP` = %d, `highSchoolInstituteName` = %s, "
        #              "`highSchoolInstituteCgpaGpaType` = %s, `highSchoolInstituteGpaorCGPA` = %f, `graduationYOP` = "
        #              "%d, `graduationInstituteName` = %s, `graduationInstituteCgpaGpaType` = %s, "
        #              "`graduationInstituteGpaorCGPA` = %f, `associateYOP` = %d, `associateInstituteName` = %s, "
        #              "`associateInstituteCgpaGpaType` = %s, `associateInstituteGpaorCGPA` = %f, `bachelorYOP` = %d, "
        #              "`bachelorInstituteName` = %s, `bachelorInstituteCgpaGpaType` = %s, "
        #              "`bachelorInstituteGpaorCGPA` = %f, `mastersYOP` = %d, `mastersInstituteName` = %s, "
        #              "`mastersInstituteCgpaGpaType` = %s, `mastersInstituteGpaorCGPA` = %f, `doctorsYOP` = %d, "
        #              "`doctorsInstituteName` = %s, `doctorsInstituteCgpaGpaType` = %s, `doctorsInstituteGpaorCGPA` "
        #              "= %f WHERE email like '") + str(
        #     email) + "' "
        update_query = """
        UPDATE studentRegistrationDetails
        SET highestSchoolYOP = %s,
            highSchoolInstituteName = %s,
            highSchoolInstituteCgpaGpaType = %s,
            highSchoolInstituteGpaorCGPA = %s,
            graduationYOP = %s,
            graduationInstituteName = %s,
            graduationInstituteCgpaGpaType = %s,
            graduationInstituteGpaorCGPA = %s,
            associateYOP = %s,
            associateInstituteName = %s,
            associateInstituteCgpaGpaType = %s,
            associateInstituteGpaorCGPA = %s,
            bachelorYOP = %s,
            bachelorInstituteName = %s,
            bachelorInstituteCgpaGpaType = %s,
            bachelorInstituteGpaorCGPA = %s,
            mastersYOP = %s,
            mastersInstituteName = %s,
            mastersInstituteCgpaGpaType = %s,
            mastersInstituteGpaorCGPA = %s,
            doctorsYOP = %s,
            doctorsInstituteName = %s,
            doctorsInstituteCgpaGpaType = %s,
            doctorsInstituteGpaorCGPA = %s,
            professionalCertificateName = %s,
            NoOfYearsOfFieldExperience = %s
        WHERE email = %s
        """

# Data to be updated
        data = (
            highestSchoolYOPvar,
            highSchoolInstituteName,
            highSchoolInstituteCgpaGpaType,
            highSchoolInstituteGpaorCgpavar,
            graduationYOPvar,
            graduationInstituteName,
            graduationInstituteCgpaGpaType,
            graduationInstituteGpaorCgpavar,
            associateYOPvar,
            associateInstituteName,
            associateInstituteCgpaGpaType,
            associateInstituteGpaorCgpavar,
            bachelorYOPvar,
            bachelorInstituteName,
            bachelorInstituteCgpaGpaType,
            bachelorInstituteGpaorCgpavar,
            mastersYOPvar,
            mastersInstituteName,
            mastersInstituteCgpaGpaType,
            mastersInstituteGpaorCgpavar,
            doctorsYOPvar,
            doctorsInstituteName,
            doctorsInstituteCgpaGpaType,
            doctorsInstituteGpaorCgpavar,
            professionalCertificateName,
            NoOfYearsOfFieldExperience,
            email
        )

# Execute the query
        cursor.execute(update_query, data)

# Commit the transaction
        conn.commit()
        # data = (
        #     highestSchoolYOP, highSchoolInstituteName, highSchoolInstituteCgpaSgpaType, highSchoolInstituteSGPAorCGPA,
        #     graduationYOP, graduationInstituteName, graduationInstituteCgpaSgpaType, graduationInstituteSGPAorCGPA,
        #     associateYOP, associateInstituteName, associateInstituteCgpaSgpaType, associateInstituteSGPAorCGPA,
        #     bachelorYOP, bachelorInstituteName, bachelorInstituteCgpaSgpaType, bachelorInstituteSGPAorCGPA,
        #     mastersYOP, mastersInstituteName, mastersInstituteCgpaSgpaType, mastersInstituteSGPAorCGPA, doctorsYOP,
        #     doctorsInstituteName, doctorsInstituteCgpaSgpaType, doctorsInstituteSGPAorCGPA)

        # cursor.execute(updateSql)

        # cursor.execute(updateSql, data)
        print('sql query executed')

        resp = jsonify({"status": 1, "success": True, "message": "Education Details Updated Successfully"})
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        # print("end",datetime.now())
        return resp
 
# @reg.route('/saveEducationDetails',methods = ['POST'])
# @cross_origin()
# def saveEducationDetails():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         print("start", datetime.now())
#         req = request.json
#         print(req)
#         email = req['email']
#         # highestQualification = req['highestQualification']
#         # passOutYearPost = req['pgYear']
#         # print(passOutYearPost)
#         # if passOutYearPost == '':
#         #     print("11111",passOutYearPost)
#         #     passOutYearPost = None
#         # univercityPost =  req['pgCollege']
#         # marksPost  = req['pgCGPA']
#         # if marksPost == '':
#         #     marksPost = None
#         # passOutYearGrad = req['gYear']
#         # if passOutYearGrad == '':
#         #     passOutYearGrad = None
#         # universityGrad =  req['gCollege']
#         # marksGrad = req['gCGPA']
#         # if marksGrad == '':
#         #     marksGrad = None
#         # passOutYear12 = req['c12Year']
#         # if passOutYear12 == '':
#         #     passOutYear12 = None
#         # college12 = req['c12College']
#         # marks12 =  req['c12CGPA']
#         # if marks12 == '':
#         #     marks12 = None
#         # passOutYear10 = req['c10Year']
#         # if passOutYear10 == '':
#         #     passOutYear10 = None
#         # school10 = req['c10College']
#         # marks10 =  req['c10CGPA']
#         # if marks10 == '':
#         #     marks10 = None
#         # crmId = req['crmId']

#         highestSchoolYOP = None
#         highSchoolInstituteGpaorCgpa = None
#         graduationYOP = None
#         graduationInstituteGpaorCgpa = None
#         associateYOP = None
#         associateInstituteGpaorCgpa = None
#         bachelorYOP = None
#         bachelorInstituteGpaorCgpa = None
#         mastersYOP = None
#         mastersInstituteGpaorCgpa = None
#         doctorsYOP = None
#         doctorsInstituteGpaorCgpa = None

#         if req['formData']['highestSchoolYOP'] != '':
#             highestSchoolYOP = int(req['formData']['highestSchoolYOP'])
#         highSchoolInstituteName = req['formData']['highSchoolInstituteName']
#         if highSchoolInstituteName == '':
#             highSchoolInstituteName = None
#         highSchoolInstituteCgpaGpaType = req['formData']['highSchoolInstituteCgpaGpaType']
#         if highSchoolInstituteCgpaGpaType == '':
#             highSchoolInstituteCgpaGpaType = None
#         if req['formData']['highSchoolInstituteGpaorCgpa'] != '':
#             highSchoolInstituteGpaorCgpa = float(req['formData']['highSchoolInstituteGpaorCgpa'])

#         if req['formData']['graduationYOP'] != '':
#             graduationYOP = int(req['formData']['graduationYOP'])
#         graduationInstituteName = req['formData']['graduationInstituteName']
#         if graduationInstituteName == '':
#             graduationInstituteName = None
#         graduationInstituteCgpaGpaType = req['formData']['graduationInstituteCgpaGpaType']
#         if graduationInstituteCgpaGpaType == '':
#             graduationInstituteCgpaGpaType = None
#         if req['formData']['graduationInstituteGpaorCgpa'] != '':
#             graduationInstituteGpaorCgpa = float(req['formData']['graduationInstituteGpaorCgpa'])

#         if req['formData']['associateYOP'] != '':
#             associateYOP = int(req['formData']['associateYOP'])
#         associateInstituteName = req['formData']['associateInstituteName']
#         if associateInstituteName == '':
#             associateInstituteName = None
#         associateInstituteCgpaGpaType = req['formData']['associateInstituteCgpaGpaType']
#         if associateInstituteCgpaGpaType == '':
#             associateInstituteCgpaGpaType = None
#         if req['formData']['associateInstituteGpaorCgpa'] != '':
#             associateInstituteGpaorCgpa = float(req['formData']['associateInstituteGpaorCgpa'])

#         if req['formData']['bachelorYOP'] != '':
#             bachelorYOP = int(req['formData']['bachelorYOP'])
#         bachelorInstituteName = req['formData']['bachelorInstituteName']
#         if bachelorInstituteName == '':
#             bachelorInstituteName = None
#         bachelorInstituteCgpaGpaType = req['formData']['bachelorInstituteCgpaGpaType']
#         if bachelorInstituteCgpaGpaType == '':
#             bachelorInstituteCgpaGpaType = None
#         if req['formData']['bachelorInstituteGpaorCgpa'] != '':
#             bachelorInstituteGpaorCgpa = float(req['formData']['bachelorInstituteGpaorCgpa'])

#         if req['formData']['mastersYOP'] != '':
#             mastersYOP = int(req['formData']['mastersYOP'])
#         mastersInstituteName = req['formData']['mastersInstituteName']
#         if mastersInstituteName == '':
#             mastersInstituteName = None
#         mastersInstituteCgpaGpaType = req['formData']['mastersInstituteCgpaGpaType']
#         if mastersInstituteCgpaGpaType == '':
#             mastersInstituteCgpaGpaType = None
#         if req['formData']['mastersInstituteGpaorCgpa'] != '':
#             mastersInstituteGpaorCgpa = float(req['formData']['mastersInstituteGpaorCgpa'])

#         if req['formData']['doctorsYOP'] != '':
#             doctorsYOP = int(req['formData']['doctorsYOP'])
#         doctorsInstituteName = req['formData']['doctorsInstituteName']
#         if doctorsInstituteName == '':
#             doctorsInstituteName = None
#         doctorsInstituteCgpaGpaType = req['formData']['doctorsInstituteCgpaGpaType']
#         if doctorsInstituteCgpaGpaType == '':
#             doctorsInstituteCgpaGpaType = None
#         if req['formData']['doctorsInstituteGpaorCgpa'] != '':
#             doctorsInstituteGpaorCgpa = float(req['formData']['doctorsInstituteGpaorCgpa'])

#         # updateSql = "UPDATE `studentRegistrationDetails` SET `highestQualification`  = %s, `passOutYearPost` = %s,`univercityPost` = %s, `marksPost`= %s,`passOutYearGrad` = %s,`universityGrad`= %s,`marksGrad`= %s,`passOutYear12` = %s,`college12`= %s,`marks12`= %s,`passOutYear10`= %s,`school10`= %s,`marks10`= %s WHERE email like '"+str(email)+"' "
#         # data=(highestQualification,passOutYearPost,univercityPost, marksPost,passOutYearGrad,universityGrad,marksGrad,passOutYear12,college12,marks12,passOutYear10,school10,marks10)

#         # print(highestSchoolYOP,
#         #       graduationYOP,
#         #       associateYOP,
#         #       bachelorYOP,
#         #       mastersYOP,
#         #       doctorsYOP,
#         #
#         #       type(highestSchoolYOP),
#         #       type(graduationYOP),
#         #       type(associateYOP),
#         #       type(bachelorYOP),
#         #       type(mastersYOP),
#         #       type(doctorsYOP)
#         #       )

#         # updateSql = ("UPDATE `studentRegistrationDetails` SET `highestSchoolYOP` = %d, `highSchoolInstituteName` = %s, "
#         #              "`highSchoolInstituteCgpaGpaType` = %s, `highSchoolInstituteGpaorCGPA` = %f, `graduationYOP` = "
#         #              "%d, `graduationInstituteName` = %s, `graduationInstituteCgpaGpaType` = %s, "
#         #              "`graduationInstituteGpaorCGPA` = %f, `associateYOP` = %d, `associateInstituteName` = %s, "
#         #              "`associateInstituteCgpaGpaType` = %s, `associateInstituteGpaorCGPA` = %f, `bachelorYOP` = %d, "
#         #              "`bachelorInstituteName` = %s, `bachelorInstituteCgpaGpaType` = %s, "
#         #              "`bachelorInstituteGpaorCGPA` = %f, `mastersYOP` = %d, `mastersInstituteName` = %s, "
#         #              "`mastersInstituteCgpaGpaType` = %s, `mastersInstituteGpaorCGPA` = %f, `doctorsYOP` = %d, "
#         #              "`doctorsInstituteName` = %s, `doctorsInstituteCgpaGpaType` = %s, `doctorsInstituteGpaorCGPA` "
#         #              "= %f WHERE email like '") + str(
#         #     email) + "' "
#         updateSql = f"UPDATE studentRegistrationDetails SET  'highestSchoolYOP' = {highestSchoolYOP}, 'highSchoolInstituteName' = '{highSchoolInstituteName}', 'highSchoolInstituteCgpaGpaType' = '{highSchoolInstituteCgpaGpaType}', 'highSchoolInstituteGpaorCGPA' = '{highSchoolInstituteGpaorCgpa}',\
#         'graduationYOP' = {graduationYOP}, 'graduationInstituteName' = '{graduationInstituteName}', 'graduationInstituteCgpaGpaType' = '{graduationInstituteCgpaGpaType}', 'graduationInstituteGpaorCGPA' = {graduationInstituteGpaorCgpa}, \
#         'associateYOP' = {associateYOP}, 'associateInstituteName' = '{associateInstituteName}', 'associateInstituteCgpaGpaType' = '{associateInstituteCgpaGpaType}', 'associateInstituteGpaorCGPA' = {associateInstituteGpaorCgpa}, \
#         'bachelorYOP' = {bachelorYOP}, 'bachelorInstituteName' = '{bachelorInstituteName}', 'bachelorInstituteCgpaGpaType' = '{bachelorInstituteCgpaGpaType}', 'bachelorInstituteGpaorCGPA' = {bachelorInstituteGpaorCgpa}, \
#         'mastersYOP' = {mastersYOP}, 'mastersInstituteName' = '{mastersInstituteName}', 'mastersInstituteCgpaGpaType' = '{mastersInstituteCgpaGpaType}', 'mastersInstituteGpaorCGPA' = {mastersInstituteGpaorCgpa}, 'doctorsYOP' = {doctorsYOP}, \
#         'doctorsInstituteName' = '{doctorsInstituteName}', 'doctorsInstituteCgpaGpaType' = '{doctorsInstituteCgpaGpaType}', 'doctorsInstituteGpaorCGPA' = {doctorsInstituteGpaorCgpa} WHERE email like '{email}'"
#         # data = (
#         #     highestSchoolYOP, highSchoolInstituteName, highSchoolInstituteCgpaSgpaType, highSchoolInstituteSGPAorCGPA,
#         #     graduationYOP, graduationInstituteName, graduationInstituteCgpaSgpaType, graduationInstituteSGPAorCGPA,
#         #     associateYOP, associateInstituteName, associateInstituteCgpaSgpaType, associateInstituteSGPAorCGPA,
#         #     bachelorYOP, bachelorInstituteName, bachelorInstituteCgpaSgpaType, bachelorInstituteSGPAorCGPA,
#         #     mastersYOP, mastersInstituteName, mastersInstituteCgpaSgpaType, mastersInstituteSGPAorCGPA, doctorsYOP,
#         #     doctorsInstituteName, doctorsInstituteCgpaSgpaType, doctorsInstituteSGPAorCGPA)

#         cursor.execute(updateSql)

#         # cursor.execute(updateSql, data)
#         print('sql query executed')

#         resp = jsonify({"status": 1, "success": True, "message": "Education Details Updated Successfully"})
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0})
#     finally:
#         cursor.close()
#         # print("end",datetime.now())
#         return resp
 
       
@reg.route('/fetchEducationInfo',methods=['POST'])
@cross_origin()
def fetchEducationInfo():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        print("Start", datetime.now())
        req = request.json
        print(req)
        email = req['email']
        print(email)

        # searchSql = " SELECT highestQualification,passOutYearPost as pgYear ,univercityPost as pgCollege, marksPost as pgCGPA,passOutYearGrad as gYear,universityGrad as gCollege,marksGrad as gCGPA,passOutYear12 as c12Year,college12 as c12College,marks12 as c12CGPA,passOutYear10 as c10Year,school10 as c10College ,marks10 as c10CGPA, email,post_graduationFlag,graduationFlag,certificate_12thFlag, certificate_10thFlag FROM `studentRegistrationDetails` WHERE `email` = '" + str(
        #     email) + "'"
        sql = "SELECT * FROM `studentRegistrationDetails` WHERE email = '"+str(email)+"'"
        cursor.execute(sql)
        result = cursor.fetchone()
        print("----------------",result)


        # if result is not None and len(result) != 0:

        #     for each in result.keys():
        #         if result[each] is None:
        #             result[each] = ''

        resp = jsonify({"status": 1, "success": True, "educationDetails": result, "message": "Fetched successfully",
                        "dataPresent": 1})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})
    finally:
        cursor.close()
        print("end", datetime.now())
        return resp

     

@reg.route('/saveProfessionalDetails',methods = ['POST'])
@cross_origin()
def saveProfessionalDetails():
    try:
        req =request.json
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()
        email = req['email']
        crmId = req['crmId']
        workingStatus = req['workStatus']
        totalExperience = req['workExp']
        if totalExperience == '':
            totalExperience = None
        currentOrganization = req['workOrganization']
        designation = req['designation']
        linkedinLink =  req['linkedinProfileLink']
        # batch = req['selectedBatch'][0]['batchname']
        try :
            batch = req['selectedBatch'][0]['batchname']
        except:
            batch = None

        update_sql = "UPDATE `studentRegistrationDetails` SET `workingStatus` =%s,`totalExperience` =%s,`currentOrganization`=%s,`designation`=%s,`linkedinLink` =%s "
        if batch != None:
         update_sql += ", batch = '"+str(batch) + "' "
        update_sql += "WHERE email LIKE '"+str(email)+"' "
        print(update_sql)

        data= (workingStatus,totalExperience,currentOrganization,designation,linkedinLink)
        cursor.execute(update_sql,data)
    
        resp = jsonify({"status":1, "success":True,   "message": "Professional Details Updated Successfully" })
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    


@reg.route('/fetchProfessionalInfo',methods=['POST'])
@cross_origin()
def fetchProfessionalInfo():
    try:
        print("end",datetime.now())
        req =request.json
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()
        email = req['email']

        searchSql = " SELECT batch, workingStatus as workStatus,totalExperience as workExp,currentOrganization as workOrganization,designation,linkedinLink as linkedinProfileLink, email,resumeFlag  FROM `studentRegistrationDetails` WHERE `email` = '"+str(email) + "'"
        cursor.execute(searchSql)
        result  =cursor.fetchone()
        if result is not None and len(result) != 0:
            for each in result.keys():
                if result[each] is None:
                    result[each] = ''
        
        resp = jsonify({"status":1, "success":True, "professionalDetails":result,  "message": "Fetched successfully","dataPresent":1 })
        


    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        print("end",datetime.now())
        return resp
    

@reg.route('/submitRegistration',methods=['POST'])
@cross_origin()
def submitRegistration():
    try:
        req =request.json
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()
        email = req['email']
        # crmId = req['crmId']

        update_sql = " UPDATE `studentRegistrationDetails` SET submitFlag = 1  WHERE email LIKE '"+str(email)+"' "
        cursor.execute(update_sql)

        update_sql = "UPDATE `converted_student_data` SET submitFlag = 1,status =1  WHERE EmailAddress LIKE '"+str(email)+"'"
        cursor.execute(update_sql)

        resp = jsonify({"status":1, "success":True,   "message": "Submitted Successfully" })

        selectSql = "Select * from studentRegistrationDetails  WHERE email LIKE '"+str(email)+"' "
        cursor.execute(selectSql)
        res = cursor.fetchone()
        name = res['firstName'] +" "+ res['middleName'] + " "+ res['lastName']
        createDoc(email,name)

        print("res>>>>",res)


        emailSearchSql = "SELECT * FROM `converted_student_data`  WHERE EmailAddress like '"+str(email)+"' "
        cursor.execute(emailSearchSql)
        crmres = cursor.fetchone()


        textBody = ""

        subject = "Successfully Submitted"
        htmlText = '''<span>Hi '''+str(crmres['FirstName'])+''',<br> Thanks for your interest in ''' +str(crmres['mx_Course'])+'''. <br> You have successfully submitted your registration form. We will notify you once it gets approved.<br><br> Best wishes,<br> <B>Techno Struct Academy<B> </span>'''

        emailIdList = []
        emailIdList.append(email)
        obj = sendMail(emailIdList,textBody,subject,htmlText,'')
        obj.sendmail()

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp


def createDoc(email,name):
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        fileName = 'Document.docx'
        folder = email.split ('@')[0]
        response ={}
        print(UPLOAD_DOC_PATH)
        path = UPLOAD_DOC_PATH + str(folder) +'/'
        print(path)

        doc = DocxTemplate(TEMPLATE_DOC +fileName  )
        idProof =  InlineImage(doc, path + 'idProof' +'.png', width=Mm(170)) # width is in millimetres
        photo = InlineImage(doc, path + 'photo' +'.png', width=Mm(100)) # width is in millimetres
        graduation = InlineImage(doc, path + 'graduation' +'.png', width=Mm(170)) # width is in millimetres
        certificate_12th = InlineImage(doc, path + 'certificate_12th' +'.png', width=Mm(170)) # width is in millimetres
        certificate_10th = InlineImage(doc, path + 'certificate_10th' +'.png', width=Mm(170)) # width is in millimetres



        response['idProof'] = idProof
        response['photo'] = photo
        response['graduation'] = graduation
        response['certificate_12th'] = certificate_12th
        response['certificate_10th'] = certificate_10th
        response['name'] = name



        doc.render(response)  
        pathFile = os.path.join(path,folder+'.docx'  )
        doc.save(pathFile)  
        generate_pdf(pathFile, path)

        resp = True
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp

# createDoc('arpand@pentationanalytics.com','Arpan Das')


@reg.route('/approveReject',methods=['POST'])
@cross_origin()
def approveReject():
    try:
        req =request.json
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()
        email = req['email']
        approveReject = req['approveReject']
        # crmId = req['crmId']
        emailIdList = []
        if approveReject == 'reject':
            rejectionReason = req['rejectionReason']
        emailSearchSql = "SELECT * FROM `converted_student_data`  WHERE EmailAddress like '"+str(email)+"' "
        cursor.execute(emailSearchSql)
        res = cursor.fetchone()
        role = 'Student'
        if approveReject == 'approve' :
 
            update_sql = "UPDATE `converted_student_data` SET status = 2 , updatedAt = '"+str(datetime.today())+"'  WHERE EmailAddress LIKE '"+str(email)+"'"
 
           
        
            textBody = ""
 
            subject = "Successfully Onboarded"
            htmlText = '''<span>Hi '''+str(res['FirstName'])+''',<br> Thanks for your interest in ''' +str(res['mx_Course'])+'''. <br> You are successfully onboarded.<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            emailIdList.append(email)
            obj = sendMail(emailIdList,textBody,subject,htmlText,'')
            obj.sendmail()
            addUser(res, role)
 
        
        else :
            textBody = ""
 
            subject = "Your Registration has been rejected"
            htmlText = '''<span>Hi '''+str(res['FirstName'])+''',<br> Thanks for your interest in ''' +str(res['mx_Course'])+'''. <br> We regret to inform you that your Registration has been rejected.<br> '''+str(rejectionReason)+'''<br><br> Best wishes,<br> <B>Techno Struct Academy<B> <br>   <img src="'''+str(img_path)+'''logo.png" width="50%" alt=""> </span>'''
            emailIdList.append(email)
            obj = sendMail(emailIdList,textBody,subject,htmlText,'')
            obj.sendmail()
 
            update_sql = "UPDATE `converted_student_data` SET status = 3 , updatedAt = '"+str(datetime.today())+"'  WHERE EmailAddress LIKE '"+str(email)+"'"
 
        cursor.execute(update_sql)
 
        resp = jsonify({"status":1, "success":True,   "message": " Successfully" })
 
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    



@reg.route('/updateProfessionalDetails', methods=['POST'])
@cross_origin()
def updateProfessionalDetails():
    conn = connect_mysql()
    cursor = conn.cursor()

    try:
        req = request.json
        designation = req['designation']
        profileLink = req['linkedinProfileLink']
        workingStatus = req['workStatus']
        batch = req['selectedBatch']
        currentOrganization = req['workOrganization']
        totalExperience = req['workExp']
        email = req['email']

        stmt = """UPDATE studentRegistrationDetails SET designation = %s,linkedinLink = %s, workingStatus = %s, batch = %s,currentOrganization = %s,       totalExperience = %s WHERE email = %s"""
        cursor.execute(stmt, (designation, profileLink, workingStatus, batch, currentOrganization, totalExperience, email))

        stmt = "UPDATE studentRegistrationDetails SET batch = %s WHERE email = %s"
        cursor.execute(stmt, (batch, email))

        stmt = "UPDATE converted_student_data SET batch = %s WHERE emailAddress = %s"
        cursor.execute(stmt, (batch, email))

        return jsonify({"status": 1, "success": True, "message": "Updated successfully"})

    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    



@reg.route('/updateBasicDetails', methods=['POST'])
@cross_origin()
def udpateBasicDetails():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        req = request.json
        req = req['info']
        firstName = req['firstName']
        middleName = req['middleName']
        lastName = req['lastName']
        gender = req['gender']
        dob = req['dob']
        mobileNumber = req['mobileNumber']
        alternateNumber = req['alternateNumber']
        email = req['email']
        gurdianName = req['gurdianName']
        IdproofType = req['IdproofType']
        idproofNumber = req['idproofNumber']

        # Check if country, state, and city are JSON strings or regular strings
        def get_name(value):
            try:
                value_json = json.loads(value)
                return value_json['name']
            except (json.JSONDecodeError, TypeError, KeyError):
                return value

        country = get_name(req['country'])
        state = get_name(req['state'])
        city = get_name(req['city'])

        zipCode = req['zipCode']
        fullAddress1 = req['fullAddress1']
        fullAddress2 = req['fullAddress2']

        udpateStudentRegistrationList = """UPDATE studentRegistrationDetails SET firstName=%s, middleName=%s, lastName=%s, gender=%s, dob=%s, mobileNumber=%s, alternateNumber=%s, email=%s, gurdianName=%s, country=%s, state=%s, city=%s, IdproofType=%s, idproofNumber=%s, zipCode=%s, fullAddress1=%s, fullAddress2=%s WHERE email=%s"""

        cursor.execute(udpateStudentRegistrationList, (
        firstName, middleName, lastName, gender, dob, mobileNumber, alternateNumber, email, gurdianName, country,
        state,
        city, IdproofType, idproofNumber, zipCode, fullAddress1, fullAddress2, email
        ))
        conn.commit()
        udpateStudentRegistrationList = """UPDATE converted_student_data SET FirstName=%s, LastName=%s, EmailAddress=%s, Phone=%s WHERE EmailAddress=%s"""

        cursor.execute(udpateStudentRegistrationList, (
        firstName, lastName, email, mobileNumber, email
        ))
        conn.commit()

        return jsonify({"Status": "Inserted in the database successfully!"})
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    finally:
        cursor.close()
        conn.close()
     




@reg.route('/dropoutStudentDetails',methods = ['POST'])
@cross_origin()
def dropoutStudentDetails():
    try:
        req =request.json
        print(req)
        conn = connect_mysql()
        cursor = conn.cursor()
        email = req['email']
        dropoutStatus = req['dropoutStatus']
        
        

        update_sql = "UPDATE `studentRegistrationDetails` SET `dropoutStatus` =%s"

        update_sql += "WHERE email LIKE '"+str(email)+"' "

        cursor.execute(update_sql,dropoutStatus)
    
        resp = jsonify({"status":1, "success":True,   "message": "DroppedOut Details Updated Successfully" })
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp