from flask import Blueprint
from flask.json import jsonify
from pymysql import NULL
from db import *
from flask.globals import request
import hashlib
from flask_cors import CORS, cross_origin
# from connectAd import ConnectToAd
import json
from common import *
import numpy as np
# from interaction import interactions
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Fill
from openpyxl.utils import get_column_letter
# import bcrypt
from urllib.parse import quote_plus
from sqlalchemy import create_engine
import math, random
import os
import requests
from api.sendMail import * 
import datetime
from .user import addUser
from datetime import datetime,timezone
import traceback;
from datetime import datetime, timedelta

import pytz
import re


mentor = Blueprint('mentor', __name__)

courseFolderPath = ''

@mentor.route('/getMentorList',methods=['POST'])
@cross_origin()
def getMentorList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        # course = req['course']
        # status = req['status']
# instructor,email,phn no, course, batch, status
    # (case when status = 1 then 'Active' else 'De-Active' END )as 
 
        mentorSql = " SELECT classroommentor.id ,listOfMentor as instructor, emailId,  batchMaster.courseType as course, batchMaster.batchName,phoneNumber,status FROM `classroommentor` Left JOIN mentorMaster ON mentorId = mentorMaster.id left JOIN batchMaster ON classroommentor.batchName = batchMaster.batchName "

        # mentorSql = "Select * from mentorMaster "
        print(mentorSql)
        
        cursor.execute(mentorSql)
        mentorList = cursor.fetchall()

    

        resp =  jsonify({"status":1, "success":True, "result": mentorList, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp






# @mentor.route('/getMentorBasicList',methods=['POST'])
# @cross_origin()

# def getMentorBasicList():

#     try:

#         conn = connect_mysql()

#         cursor = conn.cursor()

#         req = request.json
 
#         try:

#             id = req['mentorID']

#         except:

#             id = 'all'
 
#         if id == 'all':

#             mentorSql = "Select *,date(dob) as abc from mentorMaster ORDER BY id DESC"

#         else:

#             mentorSql = "Select * from mentoraster where id = " + str(id) + "ORDER BY id DESC"
 
#         cursor.execute(mentorSql)

#         mentorList = cursor.fetchall()
 
#         mentor_names = [item['name'] for item in mentorList]
 
#         cursor.execute("""SELECT batchName, mentorName FROM batchMaster""")

#         data = cursor.fetchall()
 
#         all_mentors = {}
 
#         for mentor in mentor_names:

#             batches_taught_by_mentor = []

#             for batches in data:

#                 if mentor in batches['mentorName']:

#                     batches_taught_by_mentor.append(batches['batchName'])

#             # print(mentor, set(batches_taught_by_mentor))
 
#             all_mentors[mentor] = list(batches_taught_by_mentor)
 
#         for mentor in all_mentors.keys():

#             if all_mentors[mentor]:

#                 if len(all_mentors[mentor]) == 1:

#                     cursor.execute(

#                         f"""SELECT FirstName, LastName, batch FROM converted_student_data WHERE batch='{all_mentors[mentor][0]}'""")

#                     data = cursor.fetchall()

#                 else:

#                     cursor.execute(

#                         f"""SELECT FirstName, LastName, batch FROM converted_student_data WHERE batch IN {tuple(all_mentors[mentor])}""")

#                     data = cursor.fetchall()
 
#                 all_mentors.update({f'{mentor}': len(data)})

#             else:

#                 all_mentors.update({f'{mentor}': 0})
 
#             for mentor in mentorList:

#                 mentor['NoOfStudents'] = all_mentors[f"{mentor['name']}"]
 
#         resp = jsonify({"status": 1, "success": True, "result": mentorList, "message": " Fetched successfully", })

#     except Exception as e:

#         print(e)

#         resp = jsonify({"status": 0})

#     finally:

#         cursor.close()

#         return resp

 
# def getMentorBasicList():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()
#         req =request.json
#         print(req)

#         try:
#             id = req['mentorID']
#         except:
#             id = 'all'
        
#         if id == 'all':
#             mentorSql = "Select *,date(dob) as abc from mentorMaster ORDER BY id DESC"
#         else:
#             mentorSql = "Select * from mentorMaster where id = " + str(id) + "ORDER BY id DESC"
#         print(mentorSql)
        
#         cursor.execute(mentorSql)
#         mentorList = cursor.fetchall()
#         # print(mentorList)

    

#         resp =  jsonify({"status":1, "success":True, "result": mentorList, "message": " Fetched successfully", })
    
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status":0})
#     finally:
#         cursor.close()
#         return resp
    

@mentor.route('/addMentorBatch',methods=['POST'])
@cross_origin()
def addMentorBatch():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)



        insert_sql = "INSERT INTO `classroommentor`( `mentorId`, `listOfMentor`, `qualification`, `noOfBatches`, `batchName`, `classStrength`, `batchTime`, `startDate`, `endData`) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"

        listOfMentor = req['mentor']
        batchName = req['batchName']

        insert_data = (0,listOfMentor,None,None,batchName,None,None,None,None)

        cursor.execute(insert_sql,insert_data)

        resp =  jsonify({"status":1, "success":True,  "message": " Inserted successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp



@mentor.route('/addMentor',methods=['POST'])
@cross_origin()
def addMentor():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)

        

        # `firstName`, `lastName`, `emailId`, `phoneNumber`, `address1`, `address2`, `state`, `country`, 
        name = req['name']
        Gender = req['gender']
        Dob = req['dob']
        highestQualifucation = req['highest_qualification']
        universityName = req['college_name']
        # rollNo = req['roll_no']
        yearofExp = req['total_years_of_exp']
        emailId = req['mail_id']
        country = req['country']
        panNo = req['pan_no']
        phoneNumber = req['phn_no']
        linkedinLink = req['linked_in']



        inset_sql = "INSERT INTO `mentorMaster`( `name`,  `Gender`, `Dob`, `highestQualifucation`, `universityName`,  `yearofExp`, `emailId`, `panNo`, `phoneNumber`, `country`,linkedinLink	) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

        data = (name,Gender,Dob,highestQualifucation,universityName,yearofExp,emailId,panNo,phoneNumber,country,linkedinLink)
        
        cursor.execute(inset_sql,data)
        # mentorList = cursor.fetchall()

    

        resp =  jsonify({"status":1, "success":True,  "message": " Inserted successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp




@mentor.route('/editMentorBasicInfo',methods=['POST'])
@cross_origin()
def editMentorBasicInfo():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        id = req['id']
        name = req['name']
        Gender = req['Gender']
        Dob = req['Dob']
        highestQualifucation = req['highestQualifucation']
        universityName = req['universityName']
        # rollNo = req['roll_no']
        yearofExp = req['yearofExp']
        emailId = req['emailId']
        country = req['country']
        panNo = req['panNo']
        phoneNumber = req['phoneNumber']
        linkedinLink = req['linkedinLink']
        activeFlag = req['activeFlag']

        updateSql = " UPDATE mentorMaster set name = %s,Gender =%s,Dob =%s,highestQualifucation =%s,universityName =%s,yearofExp=%s,emailId=%s,country =%s,panNo =%s,phoneNumber =%s,linkedinLink=%s, activeFlag=%s  where id = %s"

        data = (name,Gender,Dob,highestQualifucation,universityName,yearofExp,emailId,country,panNo,phoneNumber,linkedinLink,activeFlag,id)

        cursor.execute(updateSql,data)
        print("NULLLL")
        updateUserSql = " UPDATE user set activeFlag = %s where email = %s"
        userData = (activeFlag, emailId)
        cursor.execute(updateUserSql,userData)

        resp = jsonify({"status":1, "success":True,  "message": " Updated successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp



@mentor.route('/editMentor',methods=['POST'])
@cross_origin()
def editMentor():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        req =request.json
        print(req)

        id = req['id']
        status = req['status']
        updateSql = " UPDATE `classroommentor`  SET status =  "+str(status)
        # if status == 'Active':
        #     updateSql += " 1 "
        # else:
        #     updateSql += " 0 "
        updateSql += " WHERE id = "+str(id)

        cursor.execute(updateSql)


    
        resp =  jsonify({"status":1, "success":True,  "message": " updated successfully", })
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    


    # SELECT DISTINCT(listOfMentor) FROM `classroommentor`;




@mentor.route('/getMentor',methods=['POST'])
@cross_origin()
def getMentor():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        # req =request.json

        # print(req)
        # mentorName = req['mentorName']
        mentorName = ''
        # status = req['status']
        

        mentorSql = " SELECT DISTINCT(listOfMentor) as mentor FROM `classroommentor` "
        
        cursor.execute(mentorSql)
        courseList = cursor.fetchall()

    

        resp =  jsonify({"status":1, "success":True, "result": courseList, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp


@mentor.route('/mentorWiseCourse',methods=['POST'])
@cross_origin()
def mentorWiseCourse():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)

        # print(req)

        mentorId = req['id']

        mentorWiseCourseSql = """SELECT batchMaster.courseType as course,batchMaster.batchName,classroommentor.status FROM `classroommentor` LEFT JOIN batchMaster ON batchMaster.batchName = classroommentor.batchName WHERE mentorId = """+str(mentorId)
        cursor.execute(mentorWiseCourseSql)
        mentorWiseCourseList = cursor.fetchall()

        print("mentorWiseCourseList",mentorWiseCourseList)


    
        resp =  jsonify({"status":1, "success":True, "result": mentorWiseCourseList, "message": " Fetched successfully", })
    
    except Exception as e:
        print(e)
        resp = jsonify({"status":0})

    finally:
        cursor.close()
        return resp




#Rishu


# @mentor.route('/courseMentorDocumentUpload', methods=['POST'])
# @cross_origin()
# def courseMentorDocumentUpload():
#     global courseFolderPath
#     try:
#         conn = connect_mysql()
#         print("Hellooo", conn)
#         cursor = conn.cursor()
#         print("-------->",request.form)
#         contentFile = request.form.get('courseName')
#         content = contentFile.split('.')[0]
        
#         file = request.files.get('file')
#         contentExtension = file.filename.split('.')[-1].lower()

#         print("File---> ", file)
#         if not file:
#             return {"error": "No file uploaded"}, 400
        

#         path = os.path.join(UPLOAD_COURSE_PATH, content)
#         # Generate current timestamp
#         timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")

#         if not os.path.exists(path):
#             os.makedirs(path)
        
#         fileName = f"{content}_{timestamp}.{contentExtension}"

#         file.save(os.path.join(path, fileName))

#         folderName = f"{content}/{fileName}"

#         print("folderName-->", folderName)

#         courseFolderPath = folderName
#         print("courseFolderPath", courseFolderPath)




#         resp =  jsonify({"success": "File uploaded successfully"}), 200
        
#     except Exception as e:
#         print(e)
#         return {"error": "An error occurred during file upload"}, 500

#     finally:
#         cursor.close()
#         return resp

# @cross_origin()
# def addCourse():
#     global courseFolderPath
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()
#         req =request.json

#         print(req)
#         print("courseFolderPath---------->", courseFolderPath)

#         courseName = req['courseName']
#         code = req['code']
#         description = req['description']

#         insertSql = "INSERT INTO `courseMaster`(`courseName`, `code`, `description`,`folderName`) VALUES (%s,%s,%s,%s)"

#         data = (courseName,code,description,courseFolderPath)
#         cursor.execute(insertSql,data)


#         resp =  jsonify({"status":1, "success":True, "message": " Inserted successfully", })
    
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status":0})
#     finally:
#         cursor.close()
#         return resp



# @mentor.route('/mentorWiseCourseList',methods=['POST'])
# @cross_origin()
# def mentorWiseCourse():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()
#         req =request.json
#         print(req)

        

#         mentorId = req['id']

#         mentorWiseCourseSql = """SELECT batchMaster.courseType as course,batchMaster.batchName,classroommentor.status FROM `classroommentor` LEFT JOIN batchMaster ON batchMaster.batchName = classroommentor.batchName WHERE mentorId = """+str(mentorId)
#         cursor.execute(mentorWiseCourseSql)
#         mentorWiseCourseList = cursor.fetchall()

#         print("mentorWiseCourseList",mentorWiseCourseList)


    
#         resp =  jsonify({"status":1, "success":True, "result": mentorWiseCourseList, "message": " Fetched successfully", })
    
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status":0})

#     finally:
#         cursor.close()
#         return resp
    






@mentor.route('/getBatchesByMentorId', methods=['POST'])
@cross_origin()
def getBatchesByMentorId():
    try:
        mentor_id = request.json.get('id')  # Assuming mentorId is passed in the request JSON

        conn = connect_mysql()
        cursor = conn.cursor()

        sql = """
            SELECT id, batchName, mentorName, mentorDetails, courseType, activeFlag, updatedOn 
            FROM `batchMaster`
            ORDER BY 
                CASE activeFlag 
                    WHEN '2' THEN 0 
                    WHEN '1' THEN 1 
                    WHEN '0' THEN 2 
                    ELSE 3
                END,
                updatedOn DESC
        """
        cursor.execute(sql)
        batchList = cursor.fetchall()
        
        processed_batches = []
        
        for batch in batchList:
            # Check if mentorDetails is None or empty string
            if batch['mentorDetails'] is None or batch['mentorDetails'] == '':
                continue  # Skip processing this batch if mentorDetails is None or empty

            # Decode mentorDetails from JSON
            try:
                mentor_details = json.loads(batch['mentorDetails'])
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON for batch id {batch['id']}: {str(e)}")
                continue  # Skip processing this batch if JSON decoding fails

            # Check if mentor_id matches (assuming mentor_id is in mentorDetails)
            mentor_ids = [mentor['id'] for mentor in mentor_details]
            if mentor_id in mentor_ids:
                # Process the batch details
                batch_id = batch['id']
                batch_name = batch['batchName']
                mentor_name = batch['mentorName']
                course_type = batch['courseType']
                active_flag = batch['activeFlag']
                updated_on = batch['updatedOn']

                # Example of processing mentorDetails (if needed)
                # Create a list of mentor names
                mentor_names = ', '.join([mentor['name'] for mentor in mentor_details])

                # Construct batch details dictionary
                batch_details = {
                    'id': batch_id,
                    'batchName': batch_name,
                    'mentorName': mentor_name,
                    'mentorNames': mentor_names,  # Updated to include all mentor names
                    'courseType': course_type,
                    'activeFlag': active_flag,
                    'updatedOn': updated_on.strftime('%Y-%m-%d')  # Format date as string
                }
                processed_batches.append(batch_details)

        # After processing batches, prepare response
        response = {
            'success': True,
            'message': 'Batches fetched successfully.',
            'result': processed_batches
        }
        print(processed_batches)
        return jsonify(response)

    except Exception as e:
        print(str(e))  # Print or log the exception for debugging purposes
        response = {
            'success': False,
            'message': 'Failed to fetch batches. Please try again later.'
        }
        return jsonify(response)

    
@mentor.route('/getCoursesForMentor', methods=['POST'])
@cross_origin()
def getCoursesForMentor():
    try:
        mentor_id = request.json.get('id')  # Assuming mentorId is passed in the request JSON

        conn = connect_mysql()
        cursor = conn.cursor()

        sql = """
            SELECT 
                bm.id, 
                bm.batchName, 
                bm.mentorName, 
                bm.mentorDetails, 
                bm.courseType, 
                bm.activeFlag AS batchActiveFlag, 
                bm.updatedOn,
                cm.courseName,
                cm.folderName,
                cm.activeFlag AS courseActiveFlag
            FROM `batchMaster` bm
            LEFT JOIN `courseMaster` cm ON bm.courseType = cm.courseName
            ORDER BY 
                CASE bm.activeFlag 
                    WHEN '2' THEN 0 
                    WHEN '1' THEN 1 
                    WHEN '0' THEN 2 
                    ELSE 3
                END,
                bm.updatedOn DESC
        """
        cursor.execute(sql)
        batchList = cursor.fetchall()
        
        processed_batches = []
        
        for batch in batchList:
            # Check if mentorDetails is None or empty string
            if batch['mentorDetails'] is None or batch['mentorDetails'] == '':
                continue  # Skip processing this batch if mentorDetails is None or empty

            # Decode mentorDetails from JSON
            try:
                mentor_details = json.loads(batch['mentorDetails'])
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON for batch id {batch['id']}: {str(e)}")
                continue  # Skip processing this batch if JSON decoding fails

            # Check if mentor_id matches (assuming mentor_id is in mentorDetails)
            mentor_ids = [mentor['id'] for mentor in mentor_details]
            if mentor_id in mentor_ids:
                # Process the batch details
                batch_id = batch['id']
                batch_name = batch['batchName']
                mentor_name = batch['mentorName']
                course_type = batch['courseType']
                batch_active_flag = batch['batchActiveFlag']
                updated_on = batch['updatedOn']
                course_name = batch['courseName'] if batch['courseName'] else ''
                folder_name = batch['folderName'] if batch['folderName'] else ''
                course_active_flag = batch['courseActiveFlag'] if batch['courseActiveFlag'] else ''

                # Create a list of mentor names
                mentor_names = ', '.join([mentor['name'] for mentor in mentor_details])

                # Process folderName
                folder_file_list = []
                if folder_name:
                    folders_and_files = folder_name.split(',')  # Split by comma to get individual entries
                    for entry in folders_and_files:
                        parts = entry.split('/')  # Split each entry by slash to get folder and file
                        if len(parts) == 2:
                            folder_file_dict = {
                                'path': entry.strip(),
                                'file': parts[1].strip()
                            }
                            folder_file_list.append(folder_file_dict)

                # Construct batch details dictionary
                batch_details = {
                    'id': batch_id,
                    'batchName': batch_name,
                    'mentorName': mentor_name,
                    'mentorNames': mentor_names,  # Updated to include all mentor names
                    'courseType': course_type,
                    'courseName': course_name,
                    'folderName': folder_file_list,  # Updated folderName to be a list of dictionaries
                    'courseActiveFlag': course_active_flag,
                    'batchActiveFlag': batch_active_flag,
                    'updatedOn': updated_on.strftime('%Y-%m-%d')  # Format date as string
                }
                processed_batches.append(batch_details)

        # After processing batches, prepare response
        response = {
            'success': True,
            'message': 'Batches fetched successfully.',
            'result': processed_batches
        }
        print(processed_batches)
        return jsonify(response)

    except Exception as e:
        print(str(e))  # Print or log the exception for debugging purposes
        response = {
            'success': False,
            'message': 'Failed to fetch batches. Please try again later.'
        }
        return jsonify(response)




# import json

# @mentor.route('/addMentorBasicDetails', methods=['POST'])
# @cross_origin()
# def addMentorBasicDetails():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()
#         req = request.json
#         print(req)

#         name = req['name']
#         gender = req['gender']
#         dob = req['dob']
#         email = req['email']
#         phone_number = req['phoneNo']
#         alt_phone_number = req['AlternatePhoneNo']
#         address1 = req['address1']
#         address2 = req['address2']
        
#         # Extract and serialize country, state, and city details
#         country = json.dumps(req['country'])
#         state = json.dumps(req['state'])
#         city = json.dumps(req['city'])
        
#         pin = req['pinNo']
#         id_type = req['IdType']
#         id_no = req['idNo']



#         # SQL insert statement
#         insert_sql = """
#         INSERT INTO mentorMaster (
#             name, Gender,Dob, emailId, phoneNumber, altPhoneNo, address1, address2, country, state, city, pinCode, idType, idNo
#         ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
#         """

#         data = (
#             name, gender, dob, email, phone_number, alt_phone_number, address1, address2, country, state, city, pin, id_type, id_no
#         )
        
#         cursor.execute(insert_sql, data)
#         conn.commit()

#         resp = jsonify({"status": 1, "success": True, "message": "Inserted successfully"})
    
#     except Exception as e:
#         print(e)
#         resp = jsonify({"status": 0, "error": str(e)})
    
#     finally:
#         cursor.close()
#         conn.close()
#         return resp




@mentor.route('/addMentorBasicDetails', methods=['POST'])
@cross_origin()
def addMentorBasicDetails():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        print(req)
        role = 'Mentor'

        # Extract details from the request
        name = req['name']
        gender = req['gender']
        dob = req['dob']
        email = req['email']
        phone_number = req['phoneNo']
        alt_phone_number = req['AlternatePhoneNo']
        address1 = req['address1']
        address2 = req['address2']
        
        # Serialize country, state, and city details
        country = json.dumps(req['country'])
        state = json.dumps(req['state'])
        city = json.dumps(req['city'])
        
        pin = req['pinNo']
        id_type = req['IdType']
        id_no = req['idNo']

        # SQL insert statement
        insert_sql = """
        INSERT INTO mentorMaster (
            name, Gender, Dob, emailId, phoneNumber, altPhoneNo, address1, address2, country, state, city, pinCode, idType, idNo
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        data = (
            name, gender, dob, email, phone_number, alt_phone_number, address1, address2, country, state, city, pin, id_type, id_no
        )

        cursor.execute(insert_sql, data)
        conn.commit()
        addUser(req, role)
        resp = jsonify({"status": 1, "success": True, "message": "Inserted successfully"})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "error": str(e)})

    finally:
        cursor.close()
        conn.close()
        return resp
    


@mentor.route('/updateMentorEducationDetails', methods=['POST'])
@cross_origin()
def updateMentorEducationDetails():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        print(req)

        email = req['email']
        grad_degree = req.get('gradDegree', '')
        grad_dept = req.get('gradDept', '')
        grad_passout = req.get('gradPassout', '')
        grad_ins_nm = req.get('gradInsNm', '')
        grad_cgpa = req.get('gradCgpa', '')
        hs_passout = req.get('hsPassout', '')
        hs_ins_nm = req.get('hsInsNm', '')
        hs_cgpa = req.get('hsCgpa', '')
        matric_passout = req.get('matricPassout', '')
        matric_nm = req.get('matricNm', '')
        matric_cgpa = req.get('matricCgpa', '')

        # SQL update statement
        update_sql = """
        UPDATE mentorMaster
        SET gradDegree = %s, gradDept = %s, gradPassYear = %s, gradInstitute = %s, gradCGPA = %s, 
            xiiPassYear = %s, xiiInstitute = %s, xiiCGPA = %s, xPassYear = %s, xInstitute = %s, xCGPA = %s
        WHERE emailId = %s
        """

        data = (
            grad_degree, grad_dept, grad_passout, grad_ins_nm, grad_cgpa, hs_passout, hs_ins_nm, hs_cgpa, matric_passout, matric_nm, matric_cgpa, email
        )

        cursor.execute(update_sql, data)
        conn.commit()

        resp = jsonify({"status": 1, "success": True, "message": "Updated successfully"})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "error": str(e)})

    finally:
        cursor.close()
        conn.close()
        return resp
    


@mentor.route('/updateMentorProfessionDetails', methods=['POST'])
@cross_origin()
def updateMentorProfessionDetails():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json
        print(req)

        email = req['email']
        total_exp = req.get('totalExp', '')
        current_org = req.get('currentOrg', '')
        current_desg = req.get('currentDesg', '')
        linkedin_url = req.get('LinkdinUrl', '')
        active_flag = req.get('activeFlag', '')

        # SQL update statement
        update_sql = """
        UPDATE mentorMaster
        SET yearofExp = %s, currentOrg = %s, currentDesignation = %s, linkedinLink = %s, activeFlag = %s
        WHERE emailId = %s
        """

        data = (total_exp, current_org, current_desg, linkedin_url, active_flag, email)

        cursor.execute(update_sql, data)
        conn.commit()

        resp = jsonify({"status": 1, "success": True, "message": "Updated successfully"})

    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "error": str(e)})

    finally:
        cursor.close()
        conn.close()
        return resp





@mentor.route('/editMentorWholeInfo',methods=['POST'])
@cross_origin()
def editMentorWholeInfo():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req =request.json
        print(req)
        id = req['id']
        name = req['name']
    
        highestQualifucation = req['highestQualifucation']
        universityName = req['universityName']
        
        gender = req['gender']
        dob = req['dob']
        email = req['email']
        phone_number = req['phoneNo']
        alt_phone_number = req['AlternatePhoneNo']
        address1 = req['address1']
        address2 = req['address2']

         # Serialize country, state, and city details
        country = json.dumps(req['country'])
        state = json.dumps(req['state'])
        city = json.dumps(req['city'])
        
        pin = req['pinNo']
        id_type = req['IdType']
        id_no = req['idNo']
        
        grad_degree = req.get('gradDegree', '')
        grad_dept = req.get('gradDept', '')
        grad_passout = req.get('gradPassout', '')
        grad_ins_nm = req.get('gradInsNm', '')
        grad_cgpa = req.get('gradCgpa', '')
        hs_passout = req.get('hsPassout', '')
        hs_ins_nm = req.get('hsInsNm', '')
        hs_cgpa = req.get('hsCgpa', '')
        matric_passout = req.get('matricPassout', '')
        matric_nm = req.get('matricNm', '')
        matric_cgpa = req.get('matricCgpa', '')
        
        total_exp = req.get('totalExp', '')
        current_org = req.get('currentOrg', '')
        current_desg = req.get('currentDesg', '')
        linkedin_url = req.get('LinkdinUrl', '')
        active_flag = req.get('activeFlag', '')


        updateSql = """
            UPDATE mentorMaster
            SET name = %s, Gender = %s, Dob = %s, highestQualifucation = %s, universityName = %s,
                gradDegree = %s, gradDept = %s, gradPassYear = %s, gradInstitute = %s, gradCgpa = %s,
                xiiPassYear = %s, xiiInstitute = %s, xiiCGPA = %s, xPassYear = %s, xInstitute = %s, 
                xCGPA = %s, yearofExp = %s, currentOrg = %s, currentDesignation = %s, linkedinLink = %s, 
                activeFlag = %s, country = %s, state = %s, city = %s, pinCode = %s, IdType = %s, 
                idNo = %s, emailId = %s, phoneNumber = %s, altPhoneNo = %s, address1 = %s, address2 = %s
            WHERE id = %s
        """
        # data = (name,Gender,Dob,highestQualifucation,universityName,yearofExp,emailId,country,panNo,phoneNumber,linkedinLink,activeFlag,id)
        data = (
            name, gender, dob, highestQualifucation, universityName,
            grad_degree, grad_dept, grad_passout, grad_ins_nm, grad_cgpa,
            hs_passout, hs_ins_nm, hs_cgpa, matric_passout, matric_nm,
            matric_cgpa, total_exp, current_org, current_desg, linkedin_url,
            active_flag, country, state, city, pin, id_type, id_no, email, phone_number, alt_phone_number, address1,address2, id)
        
        cursor.execute(updateSql,data)

        updateUserSql = " UPDATE user set activeFlag = %s where email = %s"
        userData = (active_flag, email)
        cursor.execute(updateUserSql,userData)

        resp = jsonify({"status":1, "success":True,  "message": " Updated successfully", })

    except Exception as e:
        print(e)
        resp = jsonify({"status":0})
    finally:
        cursor.close()
        return resp
    


@mentor.route('/getMentorAllBasicList', methods=['POST'])
@cross_origin()
def getMentorAllBasicList():

    FIELD_ALIAS_MAPPING = {
    "id": "id",
    "name": "name",
    "highestQualifucation": "highestQualifucation",
    "universityName": "universityName",
    "Gender": "gender",
    "Dob": "dob",
    "emailId": "email",
    "phoneNumber": "phoneNo",
    "altPhoneNo": "AlternatePhoneNo",
    "address1": "address1",
    "address2": "address2",
    "country": "country",
    "state": "state",
    "city": "city",
    "pinCode": "pinNo",
    "idType": "IdType",
    "idNo": "idNo",
    "gradDegree": "gradDegree",
    "gradDept": "gradDept",
    "gradPassYear": "gradPassout",
    "gradInstitute": "gradInsNm",
    "gradCGPA": "gradCgpa",
    "xPassYear": "matricPassout",
    "xInstitute": "matricNm",
    "xCGPA": "matricCgpa",
    "xiiPassYear": "hsPassout",
    "xiiInstitute": "hsInsNm",
    "xiiCGPA": "hsCgpa",
    "yearofExp": "totalExp",
    "currentOrg": "currentOrg",
    "currentDesignation": "currentDesg",
    "linkedinLink": "LinkdinUrl",
    "activeFlag": "activeFlag"
}


    try:
        conn = connect_mysql()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        req = request.json
        print(req)

        try:
            id = req['id']
        except KeyError:
            id = 'all'
        
        if id == 'all':
            mentorSql = "SELECT * FROM mentorMaster"
        else:
            mentorSql = "SELECT * FROM mentorMaster WHERE id = %s"
        
        print(mentorSql)
        cursor.execute(mentorSql, (id,) if id != 'all' else ())
        mentorList = cursor.fetchall()

        # Map database fields to aliases and ensure JSON fields are parsed
        result = []
        for mentor in mentorList:
            mapped_mentor = {}
            for db_field, alias in FIELD_ALIAS_MAPPING.items():
                if db_field in mentor:
                    value = mentor[db_field]
                    if isinstance(value, str):
                        # Try to parse JSON fields
                        try:
                            mapped_mentor[alias] = json.loads(value)
                        except json.JSONDecodeError:
                            mapped_mentor[alias] = value
                    else:
                        mapped_mentor[alias] = value
            result.append(mapped_mentor)

        resp = jsonify({"status": 1, "success": True, "result": result, "message": "Fetched successfully"})
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0, "message": "An error occurred"})
    
    finally:
        cursor.close()
        conn.close()  # Ensure the connection is closed

    return resp




@mentor.route('/deleteMentor', methods=['POST'])
@cross_origin()
def deleteMentor():
    conn = connect_mysql()
    cursor = conn.cursor()
 
    try:
        req = request.json
 
        stmt = "DELETE FROM mentorMaster WHERE id = " + str(req['id'])
        cursor.execute(stmt)
        conn.commit()
 
        return jsonify({"status": 1, "success": True, "message": "Deleted successfully" })
    except Exception as e:
        print(e)
        return jsonify({"status": 0})
    


@mentor.route('/addAssignment', methods=['POST'])
def addAssignment():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        
        file = request.files.get('file')
        file_path = None
        
        start_date = request.form.get('startDate')
        end_date = request.form.get('endDate')
        total_marks = request.form.get('totalMarks')
        evaluator = request.form.get('evaluator')
        add_instructions = request.form.get('addInstructions')
        mentorEmail = request.form.get('mentorEmail')
        assignmentName = request.form.get('assignmentName')
        
        
        if file:
            filename, file_extension = os.path.splitext(file.filename)
            new_filename = f"{filename}_{start_date}{file_extension}"
            file_path = os.path.join(UPLOAD_ASSIGNMENT_PATH, new_filename)
        
            file.save(file_path)
            filePathDB = f"assignmentUploads/{new_filename}"
        else:
            return jsonify({"status": 0, "message": "No file found!"})
        
        batch_data = request.form.get('batch')
        if batch_data:
            batch_data_json = json.dumps(json.loads(batch_data))  
        else:
            batch_data_json = '[]'  

        insert_assignment_query = """
            INSERT INTO assignmentMaster (startDate, endDate, totalMarks, evaluator, addInstructions, filePath, batch, mentorEmail, assignmentName)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
        """
        cursor.execute(insert_assignment_query, (start_date, end_date, total_marks, evaluator, add_instructions, filePathDB, batch_data_json, mentorEmail, assignmentName))
        conn.commit()

        resp = jsonify({"status": 1, "success": True, "message": "Assignment added successfully"})
    
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "Error adding assignment"})

    finally:
        cursor.close()
        conn.close()
    
    return resp

@mentor.route('/getMentorBasicList', methods=['POST'])
@cross_origin()
def getMentorBasicList():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        req = request.json

        try:
            id = req['mentorID']
        except:
            id = 'all'

        if id == 'all':
            mentorSql = "SELECT *, DATE(dob) AS abc FROM mentorMaster ORDER BY id DESC"
        else:
            mentorSql = f"SELECT * FROM mentorMaster WHERE id = {id} ORDER BY id DESC"

        cursor.execute(mentorSql)
        mentorList = cursor.fetchall()

        mentor_names = [item['name'] for item in mentorList]

        cursor.execute("SELECT batchName, mentorName FROM batchMaster")
        data = cursor.fetchall()

        all_mentors = {}

        for mentor in mentor_names:
            batches_taught_by_mentor = []

            for batch in data:
                if mentor in batch['mentorName']:
                    batches_taught_by_mentor.append(batch['batchName'])

            all_mentors[mentor] = batches_taught_by_mentor

        for mentor in all_mentors.keys():
            total_batches = len(all_mentors[mentor])

            if all_mentors[mentor]:
                if len(all_mentors[mentor]) == 1:
                    cursor.execute(
                        f"SELECT FirstName, LastName, batch FROM converted_student_data WHERE batch='{all_mentors[mentor][0]}'"
                    )
                else:
                    cursor.execute(
                        f"SELECT FirstName, LastName, batch FROM converted_student_data WHERE batch IN {tuple(all_mentors[mentor])}"
                    )

                data = cursor.fetchall()
                all_mentors.update({f'{mentor}': len(data)})
            else:
                all_mentors.update({f'{mentor}': 0})

            for mentor_data in mentorList:
                if mentor_data['name'] == mentor:
                    mentor_data['NoOfStudents'] = all_mentors[mentor]
                    mentor_data['NoOfBatches'] = total_batches

        resp = jsonify({"status": 1, "success": True, "result": mentorList, "message": "Fetched successfully"})
    
    except Exception as e:
        print(e)
        resp = jsonify({"status": 0})

    finally:
        cursor.close()
        conn.close()

    return resp


#Rishu

@mentor.route('/getAssignmentForMentor', methods=['POST'])
def getAssignmentForMentor():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()
        
        data = request.json
        mentorEmail = data['mentorEmail']
        
        assignmentListSql = """
            SELECT * FROM assignmentMaster WHERE mentorEmail=%s ORDER BY id DESC 
        """
        cursor.execute(assignmentListSql, (mentorEmail,))
        assignmentList = cursor.fetchall()

        for i, assignment in enumerate(assignmentList):
            batch = assignment['batch']  
            if isinstance(batch, str):
                try:
                    assignmentList[i]['batch'] = json.loads(batch)
                except json.JSONDecodeError:
                    print(f"Error decoding batch JSON for assignment {assignment['id']}")
        
        resp = jsonify({"status": 1, "success": True, "result": assignmentList})
    
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "Error getting assignment"})

    finally:
        cursor.close()
        conn.close()
    
    return resp



@mentor.route('/getStudentSubAssignData', methods=['POST'])
def getStudentSubAssignData():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']

        assignmentListSql = """
            SELECT id FROM assignmentMaster WHERE mentorEmail=%s
        """
        cursor.execute(assignmentListSql, (mentorEmail,))
        assignmentList = cursor.fetchall()

        assignmentIds = [assignment['id'] for assignment in assignmentList]
        if assignmentIds:
            format_strings = ','.join(['%s'] * len(assignmentIds))
            studentAssignmentSql = f"""
                SELECT asm.*, CONCAT(cs.FirstName, ' ', cs.LastName) as studentName, am.endDate,am.totalMarks
                FROM assignmentStudentMaster asm
                JOIN converted_student_data cs ON asm.email = cs.EmailAddress
                JOIN assignmentMaster am ON asm.assignmentId = am.id
                WHERE asm.assignmentId IN ({format_strings}) AND asm.activeFlag = 1
            """
            cursor.execute(studentAssignmentSql, tuple(assignmentIds))
            studentAssignmentData = cursor.fetchall()
        else:
            studentAssignmentData = []

        resp = jsonify({"status": 1, "success": True, "result": studentAssignmentData})

    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "Error getting assignment"})

    finally:
        cursor.close()
        conn.close()

    return resp

@mentor.route('/updateStudentAssignmentMarks', methods=['POST'])
def updateStudentAssignmentMarks():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        mentorEmail = data['mentorEmail']
        studentAssignmentId = data['studentAssignmentId']
        marks = data['marks']

        updateMarkSql ="""
                UPDATE assignmentStudentMaster
                SET marks = %s
                WHERE id = %s
            """
        cursor.execute(updateMarkSql, (marks, studentAssignmentId))
        conn.commit()
        resp = jsonify({"status": 1, "success": True, "message": "Marks updated successfully"})
        
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "Error updating marks"})

    finally:
        cursor.close()
        conn.close()

    return resp

@mentor.route('/applyLeaveRequest', methods=['POST'])
def leaveRequest():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        firstName = data['applyLeave']['firstName']
        lastName = data['applyLeave']['lastName']
        mentorEmail = data['applyLeave']['mail']
        startDate = data['applyLeave']['startDate']
        endDate = data['applyLeave']['endDate']
        leaveType = data['applyLeave']['leaveType']
        leaveReason = data['applyLeave']['reasonForLeave']
        activeFlag = 0

        if not (mentorEmail and startDate and endDate and leaveType and leaveReason):
            return jsonify({"status": 0, "success": False, "message": "All fields are required."})

        leaveRequest ="""
                INSERT INTO leaveMaster (firstName, lastName, mentorEmail, startDate, endDate,
                leaveType, leaveReason, activeFlag) 
                VALUE (%s, %s,%s, %s,%s, %s,%s, %s)
            """
        cursor.execute(leaveRequest, (firstName, lastName, mentorEmail, startDate, endDate, 
                                      leaveType, leaveReason, activeFlag))
        conn.commit()
        resp = jsonify({"status": 1, "success": True, "message": "Leave request applied!"})
        
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "success": False, "message": "Error during applying leave!"}),400

    finally:
        cursor.close()
        conn.close()

    return resp

# @mentor.route('/applyLeaveRequest', methods=['POST'])
# def leaveRequest():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.json
#         firstName = data['firstName']
#         lastName = data['lastName']
#         mentorEmail = data['mail']
#         startDate = data['startDate']
#         endDate = data['endDate']
#         leaveType = data['leaveType']
#         leaveReason = data['reasonForLeave']
#         activeFlag = 0

#         leaveRequest ="""
#                 INSERT INTO leaveMaster (firstName, lastName, mentorEmail, startDate, endDate,
#                 leaveType, leaveReason, activeFlag) 
#                 VALUE (%s, %s,%s, %s,%s, %s,%s, %s)
#             """
#         cursor.execute(leaveRequest, (firstName, lastName, mentorEmail, startDate, endDate, 
#                                       leaveType, leaveReason, activeFlag))
#         conn.commit()
#         resp = jsonify({"status": 1, "success": True, "message": "Leave request applied!"})
        
#     except Exception as e:
#         print(e)
#         conn.rollback()
#         resp = jsonify({"status": 0, "message": "Error during applying leave!"})

#     finally:
#         cursor.close()
#         conn.close()

#     return resp



@mentor.route('/getAllLeaveRequest', methods=['POST'])
def getAllLeaveRequest():
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        leaveRequest = """
            SELECT id, CONCAT(firstName, ' ', lastName) AS fullName, mentorEmail, leaveType, leaveReason, startDate, endDate, activeFlag
            FROM leaveMaster
            ORDER BY id DESC
        """
        cursor.execute(leaveRequest)
        leaveRequestList = cursor.fetchall()
        conn.commit()
        resp = jsonify({"status": 1, "success": True, "result" : leaveRequestList, "message": "Leave request fetched!"})
        
    except Exception as e:
        print(e)
        conn.rollback()
        resp = jsonify({"status": 0, "message": "Error while fetching leaves!"})

    finally:
        cursor.close()
        conn.close()

    return resp


# @mentor.route('/approvalLeaveRequest', methods=['POST'])
# def approvalLeaveRequest():
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         data = request.json
#         print(data)
#         leaveId = data['id']
#         activeFlag = data['activeFlag']

#         leaveRequest ="""UPDATE leaveMaster set activeFlag = %s WHERE id = %s"""
#         cursor.execute(leaveRequest, (activeFlag, leaveId))
#         conn.commit()
#         resp = jsonify({"status": 1, "success": True, "message": "Leave request approved!"})
        
#     except Exception as e:
#         print(e)
#         conn.rollback()
#         resp = jsonify({"status": 0, "message": "Error during approving leave!"})

#     finally:
#         cursor.close()
#         conn.close()

#     return resp

@mentor.route('/approvalLeaveRequest', methods=['POST'])
def approvalLeaveRequest():
    def send_leave_email(email, firstName, startDate, endDate, subject, textMail,header_color, header_text):
        htmlText = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f4f4f4;
                    color: #333;
                }}
                .container {{
                    width: 100%;
                    max-width: 600px;
                    margin: 20px auto;
                    background-color: #ffffff;
                    padding: 20px;
                    border-radius: 8px;
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                }}
                .header {{
                    background-color: {header_color};
                    padding: 10px 0;
                    text-align: center;
                    color: #ffffff;
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                }}
                .header h1 {{
                    margin: 0;
                    font-size: 24px;
                }}
                .content {{
                    padding: 20px;
                    line-height: 1.6;
                }}
                .content p {{
                    margin: 10px 0;
                }}
                .footer {{
                    margin-top: 20px;
                    text-align: center;
                    color: #777;
                    font-size: 12px;
                }}
                .footer p {{
                    margin: 5px 0;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>{header_text}</h1>
                </div>
                <div class="content">
                    <p>Dear {firstName},</p>
                    <p>{textMail}</p>
                    <p>If you have any questions, feel free to contact us.</p>
                    <p>Best wishes,</p>
                    <p><strong>Techno Struct Academy</strong></p>
                </div>
                <div class="footer">
                    <p>Techno Struct Academy.</p>
                    <p>This is an automated message, please do not reply directly to this email.</p>
                </div>
            </div>
        </body>
        </html>
        '''
        emailIdList = [email]
        obj = sendMail(emailIdList, "", subject, htmlText, '')
        obj.sendmail()

    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        data = request.json
        leaveId = data['id']
        activeFlag = data['activeFlag']

        leaveRequest = """UPDATE leaveMaster set activeFlag = %s WHERE id = %s"""
        cursor.execute(leaveRequest, (activeFlag, leaveId))
        conn.commit()

        cursor.execute("SELECT firstName, mentorEmail, startDate, endDate FROM leaveMaster WHERE id = %s", (leaveId,))
        res = cursor.fetchone()

        if res:
            email = res['mentorEmail']
            firstName = res['firstName']
            startDate = res['startDate']
            endDate = res['endDate']

            if activeFlag == '1':
                send_leave_email(email, firstName, startDate, endDate, "Leave request approved!",
                                 f"We are pleased to inform you that your leave request from <strong>{startDate}</strong> to <strong>{endDate}</strong> has been approved.",
                                 "#4CAF50", "Leave Request Approved")
                postpone_meetings_for_mentor(email, startDate, endDate)                 
                return jsonify({"status": 1, "success": True, "message": "Leave request approved!"})
            
            elif activeFlag == '2':
                send_leave_email(email, firstName, startDate, endDate, "Leave request rejected!",
                                 f"We regret to inform you that your leave request from <strong>{startDate}</strong> to <strong>{endDate}</strong> has been rejected.",
                                 "#f44336", "Leave Request Rejected")
                return jsonify({"status": 1, "success": True, "message": "Leave request rejected!"})

            elif activeFlag == '3':
                send_leave_email(email, firstName, startDate, endDate, "Leave request on hold!",
                                 f"Your leave request from <strong>{startDate}</strong> to <strong>{endDate}</strong> has been placed on hold.",
                                 "#ff9800", "Leave Request On Hold")
                return jsonify({"status": 1, "success": True, "message": "Leave request on hold!"})

    except Exception as e:
        print(e)
        conn.rollback()
        return jsonify({"status": 0, "success": False, "message": "Error during approving leave!"})

    finally:
        cursor.close()
        conn.close()


# def postpone_meetings_for_mentor(email, start_date, end_date):
#     print("email, start_date, end_date", email, start_date, end_date)
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # Get all meetings assigned to the mentor
#         cursor.execute("""
#             SELECT * FROM meetings 
#             WHERE mentorList LIKE %s
#             ORDER BY startDate ASC
#         """, (f"%{email}%",))
#         all_meetings = cursor.fetchall()
       
#         if not all_meetings:
#             return jsonify({"status": 1, "success": True, "message": "No meetings found for the mentor to check."})

#         # Find meetings that need to be postponed (between start_date and end_date)
#         meetings_to_postpone = [
#             meeting for meeting in all_meetings
#             if start_date <= meeting['startDate'] <= end_date
#         ]
        

#         if not meetings_to_postpone:
#             return jsonify({"status": 1, "success": True, "message": "No meetings found within the leave period to postpone."})

#         last_meeting_date = all_meetings[-1]['startDate']
        
#         new_date = last_meeting_date

#         print("meetings_to_postpone",meetings_to_postpone) 
#         for meeting in meetings_to_postpone:
#             print("meeting",meeting) 
#             new_date += timedelta(days=1)  

#             cursor.execute("""
#                 UPDATE meetings 
#                 SET startDate = %s, endDate = %s 
#                 WHERE id = %s
#             """, (new_date, new_date, meeting['id']))

#         conn.commit()

#         return jsonify({"status": 1, "success": True, "message": "Meetings postponed successfully to the end of the schedule."})

#     except Exception as e:
#         conn.rollback()
#         return jsonify({"status": 0, "success": False, "message": f"Error during postponing meetings: {str(e)}"})

#     finally:
#         cursor.close()
#         conn.close()        


# def postpone_meetings_for_mentor(email, start_date, end_date):
#     print("email, start_date, end_date", email, start_date, end_date)
#     try:
#         conn = connect_mysql()
#         cursor = conn.cursor()

#         # Get all meetings assigned to the mentor
#         cursor.execute("""
#             SELECT * FROM meetings 
#             WHERE mentorList LIKE %s
#             ORDER BY startDate ASC
#         """, (f"%{email}%",))
#         all_meetings = cursor.fetchall()
       
#         if not all_meetings:
#             return jsonify({"status": 1, "success": True, "message": "No meetings found for the mentor to check."})

#         # Convert start_date and end_date to datetime objects for comparison
#         start_date = datetime.strptime(start_date, "%Y-%m-%d")
#         end_date = datetime.strptime(end_date, "%Y-%m-%d")

#         # Find meetings that need to be postponed (between start_date and end_date)
#         meetings_to_postpone = [
#             meeting for meeting in all_meetings
#             if start_date <= datetime.strptime(meeting['startDate'], "%Y-%m-%d") <= end_date
#         ]
        
#         if not meetings_to_postpone:
#             return jsonify({"status": 1, "success": True, "message": "No meetings found within the leave period to postpone."})

#         # Get the last meeting date and convert it to a datetime object
#         last_meeting_date = datetime.strptime(all_meetings[-1]['startDate'], "%Y-%m-%d")
#         new_date = last_meeting_date

#         print("meetings_to_postpone", meetings_to_postpone) 
#         for meeting in meetings_to_postpone:
#             print("meeting", meeting)
#             new_date += timedelta(days=1)  # Add one day to the new_date

#             # Update the meeting with the new date, formatted as a string
#             cursor.execute("""
#                 UPDATE meetings 
#                 SET startDate = %s, endDate = %s 
#                 WHERE id = %s
#             """, (new_date.strftime("%Y-%m-%d"), new_date.strftime("%Y-%m-%d"), meeting['id']))

#         conn.commit()

#         return jsonify({"status": 1, "success": True, "message": "Meetings postponed successfully to the end of the schedule."})

#     except Exception as e:
#         conn.rollback()
#         return jsonify({"status": 0, "success": False, "message": f"Error during postponing meetings: {str(e)}"})

#     finally:
#         cursor.close()
#         conn.close()


def postpone_meetings_for_mentor(email, start_date, end_date):
    print("email, start_date, end_date", email, start_date, end_date)
    try:
        conn = connect_mysql()
        cursor = conn.cursor()

        # Get all meetings assigned to the mentor
        cursor.execute("""
            SELECT * FROM meetings 
            WHERE mentorList LIKE %s
            ORDER BY startDate ASC
        """, (f"%{email}%",))
        all_meetings = cursor.fetchall()
       
        if not all_meetings:
            return jsonify({"status": 1, "success": True, "message": "No meetings found for the mentor to check."})

        # Convert start_date and end_date to datetime objects for comparison
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")

        # Find meetings that need to be postponed (between start_date and end_date)
        meetings_to_postpone = [
            meeting for meeting in all_meetings
            if start_date <= datetime.strptime(meeting['startDate'], "%Y-%m-%d") <= end_date
        ]
        
        if not meetings_to_postpone:
            return jsonify({"status": 1, "success": True, "message": "No meetings found within the leave period to postpone."})

        # Get the last meeting date and convert it to a datetime object
        last_meeting_date = datetime.strptime(all_meetings[-1]['startDate'], "%Y-%m-%d")
        new_date = last_meeting_date

        print("meetings_to_postpone", meetings_to_postpone) 
        for meeting in meetings_to_postpone:
            print("meeting", meeting)
            
            # Increment new_date until it is a weekday (Monday to Friday)
            while True:
                new_date += timedelta(days=1)
                if new_date.weekday() < 5:  # Check if new_date is a weekday (0=Monday, 4=Friday)
                    break

            # Update the meeting with the new date, formatted as a string
            cursor.execute("""
                UPDATE meetings 
                SET startDate = %s, endDate = %s 
                WHERE id = %s
            """, (new_date.strftime("%Y-%m-%d"), new_date.strftime("%Y-%m-%d"), meeting['id']))

        conn.commit()

        return jsonify({"status": 1, "success": True, "message": "Meetings postponed successfully to the end of the schedule."})

    except Exception as e:
        conn.rollback()
        return jsonify({"status": 0, "success": False, "message": f"Error during postponing meetings: {str(e)}"})

    finally:
        cursor.close()
        conn.close()






# @mentor.route('/getUpcomingMeeting', methods=['POST'])
# @cross_origin()
# def getUpcomingMeeting():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data['email']

#         cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
#         mentorId = cursor.fetchone()['id']

#         mentor_batches = []
#         mentor_meetings = []

#         cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster""")
#         for batches in cursor.fetchall():
#             if batches['mentorDetails'] and batches['batchName']:
#                 for mentors in eval(batches['mentorDetails']):
#                     if mentorId == mentors['id']:
#                         mentor_batches.append(batches['batchName'])
        
#         cursor.execute("""SELECT * FROM meetings""")

#         for meetings in cursor.fetchall():
#             if meetings['batch']:
#                 print(mentor_batches)
#                 if set(mentor_batches).issubset(set(eval(meetings['batch']))) or set(eval(meetings['batch'])).issubset(set(mentor_batches)):
#                     mentor_meetings.append(meetings)
        
        
#         upcoming_meet = []
#         utc_time = datetime.now(timezone.utc)
#         ist = pytz.timezone('Asia/Kolkata')
#         ist_time = utc_time.astimezone(ist)
#         today_date = ist_time.strftime('%Y-%m-%d')

#         # for meets in mentor_meetings:
#         #     date1 = datetime.strptime(meets['startDate'], '%Y-%m-%d')
#         #     date2 = datetime.strptime(today_date, '%Y-%m-%d')
#         #     if (date1 - date2).days >= 1:
#         #      upcoming_meet.append(meets)

#         for meets in mentor_meetings:
#             try:
#                 date1 = datetime.strptime(meets['startDate'], '%Y-%m-%d')
#                 date2 = datetime.strptime(today_date, '%Y-%m-%d')
                
#                 # Check if the difference is at least 1 day
#                 if (date1 - date2).days >= 1:
#                     upcoming_meet.append(meets)
#             except ValueError as ve:
#                 # Log or print the invalid date error
#                 print(f"Invalid date found: {meets['startDate']} - {ve}")
#                 continue  # Skip this entry if there's an invalid date

 
#         return jsonify({'status': 1, 'meetings': upcoming_meet})
#     except Exception as e:
#         print(str(e))
#         return jsonify({'status': 0})



@mentor.route('/getUpcomingMeeting', methods=['POST'])
@cross_origin()
def getUpcomingMeeting():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']

        # Fetch mentor ID
        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
        mentor_id_result = cursor.fetchone()
        
        if not mentor_id_result:
            return jsonify({'status': 1, 'meetings': []})  # If no mentor found, return no meetings
        
        mentorId = mentor_id_result['id']

        mentor_batches = []
        mentor_meetings = []

        # Fetch batch details
        cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append(batches['batchName'])

        # Fetch all meetings
        cursor.execute("""SELECT * FROM meetings""")
        
        # Ensure mentor has batches assigned before processing meetings
        if mentor_batches:
            for meetings in cursor.fetchall():
                if meetings['batch']:
                    # Ensure batch is either a subset or superset
                    if set(mentor_batches).issubset(set(eval(meetings['batch']))) or set(eval(meetings['batch'])).issubset(set(mentor_batches)):
                        mentor_meetings.append(meetings)

        upcoming_meet = []
        utc_time = datetime.now(timezone.utc)
        ist = pytz.timezone('Asia/Kolkata')
        ist_time = utc_time.astimezone(ist)
        today_date = ist_time.strftime('%Y-%m-%d')

        # Filter upcoming meetings
        for meets in mentor_meetings:
            try:
                date1 = datetime.strptime(meets['startDate'], '%Y-%m-%d')
                date2 = datetime.strptime(today_date, '%Y-%m-%d')
                
                # Check if meeting is in the future (at least 1 day)
                if (date1 - date2).days >= 1:
                    upcoming_meet.append(meets)
            except ValueError as ve:
                # Log or print the invalid date error
                print(f"Invalid date found: {meets['startDate']} - {ve}")
                continue  # Skip invalid date entries

        return jsonify({'status': 1, 'meetings': upcoming_meet})

    except Exception as e:
        print(str(e))
        return jsonify({'status': 0, 'error': str(e)})




# @mentor.route('/getMentorTodayMeeting', methods=['POST'])
# @cross_origin()
# def getMentorTodayMeeting():
#     conn = connect_mysql()
#     cursor = conn.cursor()
#     try:
#         data = request.json
#         email = data['email']

#         cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
#         mentorId = cursor.fetchone()['id']

#         mentor_batches = []
#         mentor_meetings = []

#         cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster""")
#         for batches in cursor.fetchall():
#             if batches['mentorDetails'] and batches['batchName']:
#                 for mentors in eval(batches['mentorDetails']):
#                     if mentorId == mentors['id']:
#                         mentor_batches.append(batches['batchName'])

#         cursor.execute("""SELECT * FROM meetings""")

#         for meetings in cursor.fetchall():
#             if meetings['batch']:
#                 if set(mentor_batches).issubset(set(eval(meetings['batch']))) or set(eval(meetings['batch'])).issubset(
#                         set(mentor_batches)):
#                     mentor_meetings.append(meetings)

#         today_meet = []
#         utc_time = datetime.now(timezone.utc)
#         ist = pytz.timezone('Asia/Kolkata')
#         ist_time = utc_time.astimezone(ist)
#         today_date = ist_time.strftime('%Y-%m-%d')

#         for meets in mentor_meetings:
#             if meets['startDate'] == today_date:
#                 today_meet.append(meets)

#         return jsonify({'status': 1, 'meetings': today_meet})
#     except Exception as e:
#         print(str(e))
#         return jsonify({'status': 0})



@mentor.route('/getMentorTodayMeeting', methods=['POST'])
@cross_origin()
def getMentorTodayMeeting():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']

        # Get mentor ID from mentorMaster
        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
        mentor_id_result = cursor.fetchone()
        
        if not mentor_id_result:
            return jsonify({'status': 1, 'meetings': []})  # If no mentor found, return no meetings
        
        mentorId = mentor_id_result['id']

        mentor_batches = []
        mentor_meetings = []

        # Get all batch names and mentor details from batchMaster
        cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append(batches['batchName'])

        # Get all meetings
        cursor.execute("""SELECT * FROM meetings""")

        for meetings in cursor.fetchall():
            if meetings['batch']:
                # Ensure mentor_batches is not empty before checking subsets
                if mentor_batches and (set(mentor_batches).issubset(set(eval(meetings['batch']))) or set(eval(meetings['batch'])).issubset(set(mentor_batches))):
                    mentor_meetings.append(meetings)

        # Get current date in IST timezone
        today_meet = []
        utc_time = datetime.now(timezone.utc)
        ist = pytz.timezone('Asia/Kolkata')
        ist_time = utc_time.astimezone(ist)
        today_date = ist_time.strftime('%Y-%m-%d')

        # Filter today's meetings
        for meets in mentor_meetings:
            if meets['startDate'] == today_date:
                today_meet.append(meets)

        return jsonify({'status': 1, 'meetings': today_meet})

    except Exception as e:
        print(str(e))
        return jsonify({'status': 0, 'error': str(e)})



@mentor.route('/getBatchWiseMeeting', methods=['POST'])
@cross_origin()
def getBatchWiseMeeting():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']

        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
        mentorId = cursor.fetchone()['id']

        mentor_batches = []
        mentor_meetings = []

        cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append(batches['batchName'])

        cursor.execute("""SELECT * FROM meetings""")

        for meetings in cursor.fetchall():
            if meetings['batch']:
                if set(mentor_batches).issubset(set(eval(meetings['batch']))) or set(eval(meetings['batch'])).issubset(
                        set(mentor_batches)):
                    mentor_meetings.append(meetings)

        batch_wise_meet = [{"name": None, "y": None}]

        for meets in mentor_meetings:
            for batches in eval(meets['batch']):
                for i, b_meet in enumerate(batch_wise_meet):
                    temp = [dictt['name'] for dictt in batch_wise_meet]

                    if batches in temp:
                        batch_wise_meet[temp.index(batches)] = {'name': batch_wise_meet[temp.index(batches)]['name'], 'y': batch_wise_meet[temp.index(batches)]['y'] + 1}
                    else:
                        batch_wise_meet.append({'name': batches, 'y': 1})
                    break
        batch_wise_meet.pop(0) 
        return jsonify({'status': 1, 'meetings': batch_wise_meet})
    except Exception as e:
        print(str(e))
        return jsonify({'status': 0})

@mentor.route('/getBatchWiseStudent', methods=['POST'])
@cross_origin()
def getBatchWiseStudent():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']

        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s""", (email,))
        mentorId = cursor.fetchone()['id']
        mentor_batches = []

        batch_student_data = []
        
        cursor.execute("""SELECT id, batchName, mentorDetails FROM batchMaster""")
        for batches in cursor.fetchall():
            if batches['mentorDetails'] and batches['batchName']:
                for mentors in eval(batches['mentorDetails']):
                    if mentorId == mentors['id']:
                        mentor_batches.append({
                            'batchId': batches['id'],
                            'batchName': batches['batchName']
                        })
                        
        print(mentor_batches)
        for batch in mentor_batches:
            batchId = batch['batchId']
            batchName = batch['batchName']
            
            sqlForNoOfStudent = """
            SELECT COUNT(id) as NoOfStudent FROM `converted_student_data` WHERE batch=%s;
            """
            cursor.execute(sqlForNoOfStudent, (batchName,))
            NoOfStudentResult = cursor.fetchall()
            for row in NoOfStudentResult:
                NoOfStudent = row['NoOfStudent']
                batch_student_data.append({
                    'id' : batchId,
                    'name': batchName,
                    'y': NoOfStudent
                })
                    
        return jsonify({'success': True, 'message' : "no. of student fetched!", 'result': batch_student_data})
        
    except Exception as e:
        print(str(e))
        return jsonify({'status': 0})



#RISHABH
@mentor.route('/getQuestionById', methods=['POST'])
@cross_origin()
def getQuestionById():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        id = data['id']
        email = data['email']
 
        cursor.execute("""SELECT * FROM tests WHERE id=%s""", (id,))
        tests = cursor.fetchall()
 
        cursor.execute("""SELECT testName FROM studentTestAnswers WHERE email=%s""", (email,))
        temp = cursor.fetchall()
        filter_tests = [t['testName'] for t in temp]
 
        for item in tests:
            if item['testName'] in filter_tests:
                item['test_allowed'] = False
            else:
                item['test_allowed'] = True
 
            if item['updQesAndMarks']:
                item['questionType'] = 'answer'
            else:
                item['questionType'] = 'pdf'
 
 
        return jsonify({'status': 1, 'success': True, 'tests': tests})
    except Exception as e:
        print(str(e))
        return jsonify({'status': 0})

@mentor.route('/getStudentSubmittedAnswers', methods=['POST'])
@cross_origin()
def getStudentSubmittedAnswers():
    conn = connect_mysql()
    cur = conn.cursor()
    try:
        data = request.json
        email = data['email']
        mentor_tests_ans = []
 
        cur.execute(
            """SELECT id, batchName, testName,totalMarks, updQesAndMarks, uploadFileMCQ, uploadFileDes, passingMarks FROM tests WHERE mentorEmail=%s""",
            (email,))
        mentor_tests = cur.fetchall()
 
        cur.execute("""SELECT * FROM studentTestStatus""")
        all_stud_status = cur.fetchall()
 
        for mentor_test in mentor_tests:
            temp = []
            if mentor_test['updQesAndMarks']:
                mentor_test['questionType'] = 'answer'
            else:
                mentor_test['questionType'] = 'pdf'
            for status in all_stud_status:
                if mentor_test['testName'] == status['testName'] and status['batchName'] in [item['batchName'] for item
                                                                                             in eval(
                        mentor_test['batchName'])]:
                    if status['status'] == 'SUBMITTED':
                        cur.execute("""SELECT * FROM studentTestAnswers WHERE testId=%s""", (mentor_test['id'],))
                        stud_ans = cur.fetchone()
                        stud_ans['status'] = status['status']
                        stud_ans['type'] = mentor_test['questionType']
                        temp.append(stud_ans)
                    else:
                        temp.append({'batchName': status['batchName'],
                                     'email': status['email'],
                                     'id': status['id'],
                                     'name': status['name'],
                                     'testName': status['testName'],
                                     'status': status['status'],
                                     'testId': mentor_test['id'],
                                     'answers': None,
                                     'marksObtained': None,
                                     'percentage': None,
                                     'examCleard': None,
                                     'type': mentor_test['questionType']
                                     })
            mentor_test['studentAnswers'] = temp
 
        # for item in mentor_tests:
        #     cur.execute("""SELECT * FROM studentTestAnswers WHERE testName=%s""", (item['testName'],))
        #     mentor_tests_ans = cur.fetchall()
        #
        #     for tests_ans in mentor_tests_ans:
        #         if tests_ans['answers'].startswith('['):
        #             tests_ans['type'] = 'answer'
        #         else:
        #             tests_ans['type'] = 'pdf'
        #
        #     item['questions'] = item['updQesAndMarks'] or item['uploadFileMCQ'] or item['uploadFileDes']
        #     item['questionType'] = 'answer' if item['updQesAndMarks'] else 'pdf'
        #     item['studentAnswers'] = mentor_tests_ans if mentor_tests_ans else []
        #
        # for mentor_test in mentor_tests:
        #     for all_status in all_stud_status:
        #         for batch in eval(mentor_test['batchName']):
        #             if batch['batchName'] == all_status['batchName']:
        #                 if len(mentor_test['studentAnswers']) == 0:
        #                     cur.execute("""SELECT * FROM studentTestStatus WHERE testName=%s""",
        #                                 mentor_test['testName'])
        #                     all_studs = cur.fetchall()
        #                     for studs in all_studs:
        #                         mentor_test['studentAnswers'].append({"answers": None, "email": all_status['email'],
        #                                                               "testName": all_status['testName'],
        #                                                               "name": all_status['name'],
        #                                                               "marksObtained": None, "examCleard": None,
        #                                                               "percentage": None, "testAllowed": "true",
        #                                                               "type": mentor_test['questionType',
        #                                                                       "testStatus": all_status[
        #                                                                           'testStatus'], ]})
        #                 else:
        #                     for sans in mentor_test['studentAnswers']:
        #                         if sans['email'] == all_status['email'] and sans['testName'] == all_status['testName']:
        #                             sans['testStatus'] = all_status['status']
        #                         else:
        #                             mentor_test['studentAnswers'].append({"answers": None, "email": all_status['email'],
        #                                                                   "testName": all_status['testName'],
        #                                                                   "name": all_status['name'],
        #                                                                   "marksObtained": None, "examCleard": None,
        #                                                                   "percentage": None, "testAllowed": "true",
        #                                                                   "type": mentor_test['questionType']})
 
        return jsonify({'status': 1, 'data': mentor_tests})
 
 
    except Exception as e:
        print(e)
        return jsonify({'status': 0})

@mentor.route('/deleteTests', methods=['POST'])
@cross_origin()
def deleteTests():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        id = data['id']

        cursor.execute("""DELETE FROM tests WHERE id=%s""", (id,))
        conn.commit()

        return jsonify({'status': 1, 'message': 'test deleted successfully'})
    except Exception as e:
        print(e)
        return jsonify({'status': 0})

@mentor.route('/updateMarks', methods=['POST'])
@cross_origin()
def updateMarks():
    conn = connect_mysql()
    cur = conn.cursor()
    try:
        data = request.json
        email = data['email']
        testName = data['testName']
        studentMarks = data['studentMarks']
        passingMarks = data['passingMarks']

        status = 'Passed' if studentMarks >= passingMarks else 'Failed'

        cur.execute("""UPDATE studentTestAnswers SET marksObtained=%s, examCleard=%s WHERE email=%s AND testName=%s""", (studentMarks, status, email, testName))
        return jsonify({'status': 1, 'message': 'marks udpated successfully'})

    except Exception as e:
        print(e)
        return jsonify({'status': 0})

@mentor.route('/getBatchStudentTests', methods=['POST'])
@cross_origin()
def getBatchStudentTests():
    conn = connect_mysql()
    cursor = conn.cursor()
    try:
        data = request.json
        email = data['email']
        mentor_batches = []
        mentor_tests = {}
        mentor_batchwise_tests = {}
 
        cursor.execute("""SELECT id FROM mentorMaster WHERE emailId=%s ORDER BY id DESC""", (email,))
        mentorId = cursor.fetchone()['id']
 
        cursor.execute("""SELECT batchName, mentorDetails FROM batchMaster ORDER BY id DESC""")
        all_batches = cursor.fetchall()
 
        for batches in all_batches:
            if batches['mentorDetails']:
                for mentors in eval(batches['mentorDetails']):
                    if mentors['id'] == mentorId:
                        mentor_batches.append(batches['batchName'])
 
        cursor.execute("""SELECT testName, batchName, totalMarks, passingMarks FROM tests ORDER BY id DESC""")
        all_tests = cursor.fetchall()
        for tests in all_tests:
            if tests['batchName']:
                for batches in eval(tests['batchName']):
                    if batches['batchName'] in mentor_batches:
                        if batches['batchName'] in mentor_tests.keys():
                            mentor_tests[batches['batchName']].append(tests['testName'])
                        else:
                            mentor_tests[batches['batchName']] = [tests['testName']]
 
        for key, values in mentor_tests.items():
            cursor.execute("""SELECT * FROM studentTestAnswers WHERE testName iN %s""", (values,))
            mentor_batchwise_tests[key] = cursor.fetchall()
 
        for key, values in mentor_batchwise_tests.items():
            for tests in values:
                for a_tests in all_tests:
                    if a_tests['testName'] == tests['testName']:
                        tests['totalMarks'] = a_tests['totalMarks']
                        tests['passingMarks'] = a_tests['passingMarks']
 
        temp = []
 
        for mentor_, tests_ in mentor_batchwise_tests.items():
            temp.append({'batchName': mentor_, 'tests': tests_})
 
        return jsonify({'status': 1, 'data': temp})
 
    except Exception as e:
        print(traceback.format_exc())
        print(e)
        return jsonify({'status': 0})
 

